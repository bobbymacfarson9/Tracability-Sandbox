from __future__ import annotations

from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl


def round_half_up(value: float, ndigits: int) -> float:
    q = Decimal("1").scaleb(-ndigits)
    return float(Decimal(str(value)).quantize(q, rounding=ROUND_HALF_UP))


def main() -> None:
    out = Path("grade outs") / "March_19_grade_outs_2026.xlsx"

    # Match your February file header style/order (Barn 10, Barn 11, Barn 6).
    barns = [10, 11, 6]

    # March 19 transcription (Barn 11 + Barn 6 only); Barn 10 stays blank.
    data = {
        "J/XL": {10: None, 11: 2.3 + 25.5, 6: 3.0 + 26.5},  # Jumbo% + XLarge%
        "Jumbo": {10: None, 11: 2.3, 6: 3.0},
        "Xlarge": {10: None, 11: 25.5, 6: 26.5},
        "Large": {10: None, 11: 62.6, 6: 61.0},
        "Medium": {10: None, 11: 8.1, 6: 8.2},
        "Small": {10: None, 11: 0.1, 6: 0.1},
        "Peewee": {10: None, 11: 0.0, 6: 0.0},
        # Undergrade = offgrades_total percentage
        "Undergrade": {10: None, 11: 1.4, 6: 1.2},
        "Avg. Wt. (g)": {10: None, 11: 61.622, 6: 61.725},
        "Prod. (doz.)": {10: None, 11: 25915 / 12, 6: 47478 / 12},
        "Feed (kg)": {10: None, 11: None, 6: None},
        "Liquid": {10: None, 11: 0.2, 6: 0.2},
        "Blood": {10: None, 11: 0.2, 6: 0.2},
        "Crack": {10: None, 11: 0.1, 6: 0.1},
        "Dirt": {10: None, 11: 1.0, 6: 0.7},
        "Total %": {10: None, 11: 100.0, 6: 100.0},
    }

    categories = [
        "J/XL",
        "Jumbo",
        "Xlarge",
        "Large",
        "Medium",
        "Small",
        "Peewee",
        "Undergrade",
        "Avg. Wt. (g)",
        "Prod. (doz.)",
        "Feed (kg)",
        "Liquid",
        "Blood",
        "Crack",
        "Dirt",
        "Total %",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = ["Category"] + [f"Barn {b}" for b in barns]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    col_map = {barn: 2 + i for i, barn in enumerate(barns)}

    percent_rows = {
        "J/XL",
        "Jumbo",
        "Xlarge",
        "Large",
        "Medium",
        "Small",
        "Peewee",
        "Undergrade",
        "Liquid",
        "Blood",
        "Crack",
        "Dirt",
        "Total %",
    }

    for row_idx, category in enumerate(categories, start=2):
        ws.cell(row=row_idx, column=1, value=category)
        for barn in barns:
            v = data[category][barn]
            if v is None:
                continue  # leave blank cell

            if category in percent_rows:
                v = round_half_up(float(v), 1)
            elif category in {"Avg. Wt. (g)", "Prod. (doz.)"}:
                v = round_half_up(float(v), 1)

            ws.cell(row=row_idx, column=col_map[barn], value=v)

    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()

