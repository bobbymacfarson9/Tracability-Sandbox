"""
SQF Traceability Report — one script for 99%+ traceability (producer-grader, nest run).

Inputs (all in same unit, e.g. cases/boxes):
  • Production: auto from Hilly Acres slips, or fill in CSV.
  • Shipped via orders: auto from PalletLines (loading slip reports).
  • Shipped via nest run: surplus shipped nest run (fill in CSV).
  • Grade out / off grades: auto from grade outs folder, or fill in CSV.

Reconciliation: Production = Shipped_Orders + Shipped_NestRun + OffGrades_GradeOut + Variance
Traceability % = Accounted / Production × 100  (Accounted = Orders + NestRun + OffGrades)
Target: 99%+ for SQF audit.

Single input file: Reference_Data/SQF_Traceability_Inputs.csv
  Columns: Week, Production, Shipped_NestRun, OffGrades_GradeOut [, SlipWeek, HillyAcresWeek, Reallocate_To_Prior_Week ]
  Optional SlipWeek: when slip file week differs from report week (e.g. 31), set SlipWeek so Orders use that file.
  Optional Reallocate_To_Prior_Week (2025 only): move this many cases from this week's Orders to the previous week (for weeks over 100% to get to 99%+). Export includes OldDate_NewDate_ByDay sheet to justify.
  Production and OffGrades auto-filled from Hilly Acres + grade outs when CSV has 0 or blank.
Optional evidence file: Reference_Data/Traceability_Adjustments.csv
  Columns: Week, Category, Cases, Evidence, Note
  Use only for evidence-backed missing buckets (e.g. Manual NestRun, Internal Use, Writeoff).

Run: python Scripts/sqf_traceability.py
     python Scripts/sqf_traceability.py --week 59
     python Scripts/sqf_traceability.py --export   (writes Excel for audit)
"""

import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
import argparse
import json
import os
import re
import sys

SCRIPT_DIR = Path(__file__).parent
BASE_DIR = SCRIPT_DIR.parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))
# Sandbox: copy Reference_Data + Traceability_Exports under e.g. Sandbox_Traceability/; leave BASE_DIR
# (grade outs, Hilly Acres slip paths) on the real project root. Set EGGROOM_TRACEABILITY_SANDBOX to that folder.
_SANDBOX_ROOT = (os.environ.get("EGGROOM_TRACEABILITY_SANDBOX") or "").strip()
_DATA_BASE = (os.environ.get("EGGROOM_TRACEABILITY_DATA_BASE") or "").strip()
DATA_BASE_DIR = Path(_DATA_BASE).resolve() if _DATA_BASE else BASE_DIR
if _SANDBOX_ROOT:
    _SR = Path(_SANDBOX_ROOT).resolve()
    REF_DATA_DIR = _SR / "Reference_Data"
    EXPORTS_DIR = _SR / "Traceability_Exports"
else:
    REF_DATA_DIR = BASE_DIR / "Reference_Data"
    EXPORTS_DIR = BASE_DIR / "Traceability_Exports"
# Split exports into source-of-truth vs doctored/balanced outputs
ORIGINAL_EXPORTS_DIR = EXPORTS_DIR / "Original"
# BalancedWorking is the effective per-week PalletLines view for balanced runs
BALANCED_WORKING_DIR = EXPORTS_DIR / "BalancedWorking"


def slip_week_id(year, week):
    """Canonical slip-week identifier used across originals, balanced, and exports."""
    try:
        y = int(year)
    except (TypeError, ValueError):
        y = 0
    try:
        w = int(week)
    except (TypeError, ValueError):
        w = 0
    return f"{y}-W{w:02d}"
GRADE_OUTS_DIR = DATA_BASE_DIR / "grade outs"

INPUT_CSV = REF_DATA_DIR / "SQF_Traceability_Inputs.csv"
ADJUSTMENTS_CSV = REF_DATA_DIR / "Traceability_Adjustments.csv"
PRODUCTION_ADJUSTMENTS_CSV = REF_DATA_DIR / "Traceability_Production_Adjustments.csv"
PRODUCTION_WEEKLY_OVERRIDES_CSV = REF_DATA_DIR / "Traceability_Production_Weekly_Overrides.csv"
FAILSAFE_EVIDENCE_CSV = REF_DATA_DIR / "Traceability_Failsafe_Evidence.csv"
GRADEOUT_CSV = "gradeout_data_export.csv"
TARGET_PCT = 99.0

# Optional override: allow testing with a draft adjustments CSV without changing office-controlled files.
ADJUSTMENTS_CSV_OVERRIDE = None

# When no off-grade data (gradeout CSV or CSV column): use this % of Production as estimated off grades
DEFAULT_OFFGRADE_PCT = 1.2

# Days from pack/production date to BB date. Used to attribute orders: PackDate = BBDate - N days.
# Configurable via paths.json "BBToPackDays" so it's consistent. Fallback 47.
def _get_bb_to_pack_days():
    try:
        with open(REF_DATA_DIR / "paths.json", encoding="utf-8") as f:
            n = json.load(f).get("BBToPackDays")
        if n is not None and isinstance(n, (int, float)) and 1 <= n <= 90:
            return int(n)
    except Exception:
        pass
    return 47

# Standard HA case = 15 dozen = 180 eggs (same as Scripts/hilly_acres_production.EGGS_PER_CASE).
EGGS_PER_STANDARD_CASE = 180
# Nova & Loblaws 30-pack shippers only (not OC 30, not Sobeys): 8 flats × 30 eggs = 240 eggs/box (4/3 case-equiv).
EGGS_30_PACK_BOX = 8 * 30
FACTOR_30_PACK_CASE_EQUIV = EGGS_30_PACK_BOX / EGGS_PER_STANDARD_CASE  # 4/3
# Jumbo retail boxes: 12 dozen per box vs 15 dozen per standard case.
FACTOR_JUMBO_CASE_EQUIV = 12 / 15

# Exact SKU names that use the Nova/Loblaws 8×30 shipper factor (must match normalized lower()).
THIRTY_PACK_SKU_EXACT = ("Lrg 30 pack Nova", "Lob 30 Lg")

# Word-boundary patterns: avoids "innovation" matching "nova", "global" matching "lob", etc.
_RE_WORD_NOVA = re.compile(r"\bnova\b", re.IGNORECASE)
_RE_WORD_LOBLAWS = re.compile(r"\bloblaws\b", re.IGNORECASE)
_RE_WORD_LOB = re.compile(r"\blob\b", re.IGNORECASE)
# OC 30 (Own Brands / OC line) — count as 1 case-equiv per box, not the 8×30 Nova/Loblaws factor.
_RE_OC_30_ANYWHERE = re.compile(r"\boc\s*30\b", re.IGNORECASE)


def _is_nova_loblaws_30_pack_shipper(sku_norm: str) -> bool:
    """
    True only for Nova- or Loblaws-branded 30-pack retail shippers (8×30 eggs → 4/3 case-equiv per box).
    Excludes OC 30, Sobeys, and other chains even if the line has "30".
    """
    if not sku_norm:
        return False
    if "sobeys" in sku_norm:
        return False
    if _RE_OC_30_ANYWHERE.search(sku_norm):
        return False
    has_30 = "30" in sku_norm or "30 pack" in sku_norm
    if not has_30:
        return False
    if _RE_WORD_NOVA.search(sku_norm):
        return True
    if _RE_WORD_LOBLAWS.search(sku_norm):
        return True
    if _RE_WORD_LOB.search(sku_norm):
        return True
    return False

DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
PALLETLINES_SOURCE_PREFERRED = "preferred"
PALLETLINES_SOURCE_ORIGINAL_ONLY = "original_only"
PALLETLINES_SOURCE_ADJUSTED_ONLY = "adjusted_only"
_SLIP_DAY_ORDER_TOTALS_CACHE = {}


def _case_equivalent_boxes(qty, sku):
    """
    Return case-equivalent count vs standard 180-egg (15-dozen) cases.
    Order: Jumbo (12 dozen/box) first, then Nova/Loblaws 30-pack shippers (8×30 eggs/box), else 1×.
    """
    if not qty or qty <= 0:
        return 0.0
    q = float(qty)
    if sku is None or (isinstance(sku, float) and pd.isna(sku)):
        return q
    sku_norm = str(sku).strip().lower()
    if not sku_norm:
        return q
    # Jumbo retail boxes: 12 dozen per box vs 15 dozen per standard case (PalletLines only).
    if "jumbo" in sku_norm:
        return q * FACTOR_JUMBO_CASE_EQUIV
    # 8 flats × 30 eggs = 240 eggs = 240/180 case-equiv per physical box (Nova/Loblaws only).
    for name in THIRTY_PACK_SKU_EXACT:
        if name.lower() == sku_norm:
            return q * FACTOR_30_PACK_CASE_EQUIV
    if _is_nova_loblaws_30_pack_shipper(sku_norm):
        return q * FACTOR_30_PACK_CASE_EQUIV
    return q


def _case_factor_for_sku(sku):
    """Return the case-equivalent factor per physical box for this SKU (float)."""
    return float(_case_equivalent_boxes(1.0, sku))


def _apply_core_metrics(r):
    """Production vs orders + nest only (excludes off-grades, eggs stored, manual adjustments). Modifies r in place."""
    prod = float(r.get("Production") or 0)
    orders = float(r.get("Shipped_Orders") or 0)
    nest = float(r.get("Shipped_NestRun") or 0)
    core = orders + nest
    r["Core_Accounted"] = round(core, 2)
    r["Core_Variance"] = round(prod - core, 2) if prod else 0.0
    r["Core_Traceability_Pct"] = round(100.0 * core / prod, 2) if prod > 0 else None


def _is_adjusted_palletlines_file(path_obj):
    """True when a PalletLines filename is an adjusted variant."""
    return "_palletlines_adjusted_" in str(path_obj.name).lower()


def _get_latest_palletlines_file(slip_week, exports_dir=None, source_mode=PALLETLINES_SOURCE_PREFERRED):
    """
    Return the PalletLines workbook to use for a slip week.
    source_mode:
      preferred: latest adjusted file if present, else latest original
      original_only: latest non-adjusted file only
      adjusted_only: latest adjusted file only
    """
    exports_dir = Path(exports_dir) if exports_dir else EXPORTS_DIR
    if not exports_dir.exists():
        return None
    pattern = f"Week{slip_week}_AllDays_PalletLines*.xlsx"
    files = sorted(exports_dir.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        return None
    adjusted = [p for p in files if _is_adjusted_palletlines_file(p)]
    original = [p for p in files if not _is_adjusted_palletlines_file(p)]
    if source_mode == PALLETLINES_SOURCE_ORIGINAL_ONLY:
        return original[0] if original else None
    if source_mode == PALLETLINES_SOURCE_ADJUSTED_ONLY:
        return adjusted[0] if adjusted else None
    return adjusted[0] if adjusted else (original[0] if original else None)


def _load_palletlines_df(slip_week, exports_dir=None, source_mode=PALLETLINES_SOURCE_PREFERRED):
    """Load PalletLines sheet for one slip week using the requested source mode."""
    path = _get_latest_palletlines_file(slip_week, exports_dir, source_mode=source_mode)
    if path is None:
        return None, None
    try:
        return pd.read_excel(path, sheet_name="PalletLines"), path
    except Exception:
        return None, path


def _latest_palletlines_files_by_week(exports_dirs=None, source_mode=PALLETLINES_SOURCE_PREFERRED):
    """
    Return {slip_week: path} for the latest PalletLines workbook per slip week.
    exports_dirs can be a single path or an iterable of paths; later directories
    do not shadow earlier ones except when a newer file for the same slip week
    exists (based on mtime).
    """
    if exports_dirs is None:
        exports_dirs = [EXPORTS_DIR]
    else:
        # Allow caller to pass a single path or list/tuple
        if isinstance(exports_dirs, (str, Path)):
            exports_dirs = [exports_dirs]
    # Normalise and filter to existing directories
    dirs = []
    for d in exports_dirs:
        p = Path(d)
        if p.exists():
            dirs.append(p)
    if not dirs:
        return {}
    out = {}
    for exports_dir in dirs:
        for path in exports_dir.glob("Week*_AllDays_PalletLines*.xlsx"):
            if source_mode == PALLETLINES_SOURCE_ORIGINAL_ONLY and _is_adjusted_palletlines_file(path):
                continue
            if source_mode == PALLETLINES_SOURCE_ADJUSTED_ONLY and not _is_adjusted_palletlines_file(path):
                continue
            name = path.stem.replace("Week", " ", 1).split("_")
            if not name or not name[0].strip().isdigit():
                continue
            slip_week = int(name[0].strip())
            if slip_week < 1 or (53 <= slip_week <= 55):
                continue
            prev = out.get(slip_week)
            if prev is None or path.stat().st_mtime > prev.stat().st_mtime:
                out[slip_week] = path
    return out


def build_balanced_working_view():
    """
    Ensure BalancedWorking has exactly one effective PalletLines workbook per slip week.

    For each slip week with an original PalletLines workbook in ORIGINAL_EXPORTS_DIR:
      - If BalancedWorking already has one or more adjusted files for that week, keep them and
        use the latest as the effective workbook.
      - Otherwise copy the latest original workbook for that week into BalancedWorking
        so that balanced runs have full coverage.
    Returns a mapping {slip_week: Path(effective_workbook)} based on the BalancedWorking folder.
    """
    ORIGINAL_EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    BALANCED_WORKING_DIR.mkdir(parents=True, exist_ok=True)

    # Originals by slip week: search both top-level Traceability_Exports and the
    # optional Original/ subfolder so we treat historical top-level files as
    # source-of-truth without forcing a physical move.
    originals = _latest_palletlines_files_by_week(
        [EXPORTS_DIR, ORIGINAL_EXPORTS_DIR],
        source_mode=PALLETLINES_SOURCE_ORIGINAL_ONLY,
    )

    # Existing balanced files (may include adjusted + copied originals)
    balanced_all = {}
    for path in BALANCED_WORKING_DIR.glob("Week*_AllDays_PalletLines*.xlsx"):
        name = path.stem.replace("Week", " ", 1).split("_")
        if not name or not name[0].strip().isdigit():
            continue
        slip_week = int(name[0].strip())
        if slip_week < 1 or (53 <= slip_week <= 55):
            continue
        balanced_all.setdefault(slip_week, []).append(path)

    # For each original slip week, ensure BalancedWorking has at least one workbook
    for slip_week, orig_path in originals.items():
        candidates = balanced_all.get(slip_week, [])
        # Prefer adjusted variants if they exist in BalancedWorking
        adjusted = [p for p in candidates if _is_adjusted_palletlines_file(p)]
        dest = BALANCED_WORKING_DIR / orig_path.name
        if adjusted:
            # Already have adjusted files; keep them and let downstream choose the latest adjusted.
            continue
        try:
            # No adjusted files: ensure BalancedWorking has a fresh copy of the latest original.
            # If a stale original-copy exists (or an old filename), refresh it so balanced runs
            # never use outdated order totals.
            need_refresh = True
            if dest.exists():
                try:
                    need_refresh = dest.stat().st_mtime < orig_path.stat().st_mtime
                except Exception:
                    need_refresh = True
            if need_refresh or not dest.exists():
                dest.write_bytes(orig_path.read_bytes())
            # Replace any prior non-adjusted placeholders for this week with the refreshed dest.
            balanced_all[slip_week] = [dest]
        except Exception:
            # If copy fails we simply leave the week without a balanced file;
            # validation will surface the problem later.
            continue

    # Compute effective mapping from BalancedWorking
    effective = {}
    for slip_week, paths in balanced_all.items():
        if not paths:
            continue
        # Choose latest by modification time
        best = sorted(paths, key=lambda p: p.stat().st_mtime, reverse=True)[0]
        effective[slip_week] = best
    return effective


def _clear_palletlines_caches():
    """Clear derived PalletLines caches after writing adjusted files."""
    _SLIP_DAY_ORDER_TOTALS_CACHE.clear()


def _get_slip_day_order_totals(exports_dir=None, source_mode=PALLETLINES_SOURCE_PREFERRED):
    """Return {report_week: case_equiv_total} for slip-day attribution, loading each file once."""
    exports_dir = Path(exports_dir) if exports_dir else EXPORTS_DIR
    cache_key = (str(exports_dir.resolve()), source_mode)
    if cache_key in _SLIP_DAY_ORDER_TOTALS_CACHE:
        return _SLIP_DAY_ORDER_TOTALS_CACHE[cache_key]
    totals = {}
    for slip_week, path in _latest_palletlines_files_by_week(exports_dir, source_mode=source_mode).items():
        try:
            df = pd.read_excel(path, sheet_name="PalletLines")
        except Exception:
            continue
        if "QtyBoxes" not in df.columns:
            continue
        for _, row in df.iterrows():
            rw = row.get("ReportWeek")
            if rw is None or (isinstance(rw, float) and pd.isna(rw)):
                rw = _slip_day_to_report_week(row.get("WeekNumber") or slip_week, row.get("DayName") or row.get("Day"))
            try:
                report_week = int(float(rw))
            except (TypeError, ValueError):
                continue
            totals[report_week] = totals.get(report_week, 0) + _case_equivalent_boxes(_safe_float(row.get("QtyBoxes"), 0), row.get("SKU"))
    _SLIP_DAY_ORDER_TOTALS_CACHE[cache_key] = totals
    return totals


def _report_week_sequence():
    """Ordered report weeks that participate in balancing."""
    return list(range(1, 53)) + list(range(56, 62))


def _previous_report_week(week_number):
    """Previous valid report week without crossing the 52 -> 56 year boundary."""
    weeks = _report_week_sequence()
    if week_number not in weeks:
        return None
    idx = weeks.index(week_number)
    return weeks[idx - 1] if idx > 0 else None


def _next_report_week(week_number):
    """Next valid report week without crossing the 52 -> 56 year boundary."""
    weeks = _report_week_sequence()
    if week_number not in weeks:
        return None
    idx = weeks.index(week_number)
    return weeks[idx + 1] if idx < len(weeks) - 1 else None


def _get_gradeouts_path():
    """Get grade outs folder path from paths.json or fallback."""
    paths_file = REF_DATA_DIR / "paths.json"
    if paths_file.exists():
        try:
            with open(paths_file, encoding="utf-8") as f:
                config = json.load(f)
            path = config.get("GradeOutsPath")
            if path:
                p = Path(path)
                if not p.is_absolute():
                    p = DATA_BASE_DIR / p
                if p.exists():
                    return p
        except Exception:
            pass
    return GRADE_OUTS_DIR


def get_production_from_hilly_acres(week_number):
    """Get production (cases) from Hilly Acres slips. Returns None if missing."""
    try:
        import hilly_acres_production as ha
        return ha.get_production_for_week(week_number)
    except ImportError:
        return None


def _get_hilly_acres_folders_for_report_week(report_week):
    """Restrict Hilly Acres override lookups to the same report-year context when possible."""
    try:
        import hilly_acres_production as ha
        folders = ha._get_hilly_acres_paths()
        if not folders:
            return []
        if _year_for_report_week(report_week) == 2026:
            year_folders = [f for f in folders if ha._is_2026_folder(f)]
            return year_folders or folders
        year_folders = [f for f in folders if not ha._is_2026_folder(f)]
        return year_folders or folders
    except Exception:
        return []


def get_production_from_hilly_acres_override(report_week, explicit_week):
    """
    Get Hilly Acres production for an explicit week number inside the same report-year context.
    Example: report week 59 + HillyAcresWeek 4 should use the 2026 folder's Week 4 workbook.
    """
    if explicit_week is None:
        return None
    try:
        import hilly_acres_production as ha
    except ImportError:
        return None
    for folder in _get_hilly_acres_folders_for_report_week(report_week):
        try:
            production = ha.get_production_for_week(explicit_week, folder_path=folder)
            if production is not None and production > 0:
                return production
        except Exception:
            continue
    return None


def get_nest_run_from_hilly_acres_override(report_week, explicit_week):
    """Get Hilly Acres nest run for an explicit week number inside the same report-year context."""
    if explicit_week is None:
        return None
    try:
        import hilly_acres_production as ha
    except ImportError:
        return None
    for folder in _get_hilly_acres_folders_for_report_week(report_week):
        try:
            nr = ha.get_nest_run_boxes_for_week(explicit_week, folder_path=folder)
            if nr is not None and nr > 0:
                return nr
        except Exception:
            continue
    return None


def _report_week_to_iso_year_week(week_number):
    """Return (iso_year, iso_week) for date matching. 2025: (2025, 1-52); 2026 report 56-61: (2026, 1-6)."""
    if 1 <= week_number <= 52:
        return (2025, week_number)
    if 56 <= week_number <= 61:
        return (2026, week_number - 55)
    return None


def get_off_grades_from_gradeout(week_number):
    """
    Calculate off-grade cases for the week from gradeout_data_export.csv.
    Formula: production(barn, day) × (Undergrade + Liquid + Blood + Dirt) / 100 per row, sum by week.
    Uses ISO year/week for date-to-week mapping. Report weeks 56-61 map to 2026 ISO weeks 1-6.
    Returns None if no grade-out records for week.
    """
    gradeouts_dir = _get_gradeouts_path()
    csv_path = gradeouts_dir / GRADEOUT_CSV
    if not csv_path.exists():
        return None
    try:
        df = pd.read_csv(csv_path)
    except Exception:
        return None
    if df.empty or "source_date" not in df.columns or "barn" not in df.columns:
        return None
    target_iy_iw = _report_week_to_iso_year_week(week_number)
    if target_iy_iw is None:
        return None
    target_iy, target_iw = target_iy_iw
    try:
        import hilly_acres_production as ha
    except ImportError:
        return None
    prod_by_barn_day = ha.get_production_by_barn_day(week_number)
    if prod_by_barn_day is None:
        return None
    total_off = 0.0
    seen_any = False
    for _, row in df.iterrows():
        try:
            src_date = row.get("source_date")
            if pd.isna(src_date):
                continue
            dt = pd.to_datetime(src_date)
            iso_year, iso_week, iso_weekday = dt.isocalendar()
            if (iso_year, iso_week) != (target_iy, target_iw):
                continue
            seen_any = True
            day_name = DAY_NAMES[iso_weekday - 1] if 1 <= iso_weekday <= 7 else None
            if not day_name:
                continue
            barn = int(float(row["barn"])) if pd.notna(row.get("barn")) else None
            if barn is None:
                continue
            pct_ug = _safe_float(row.get("Undergrade"), 0)
            pct_liq = _safe_float(row.get("Liquid"), 0)
            pct_blood = _safe_float(row.get("Blood"), 0)
            pct_dirt = _safe_float(row.get("Dirt"), 0)
            off_pct = pct_ug + pct_liq + pct_blood + pct_dirt
            if off_pct <= 0:
                continue
            key = (day_name, barn)
            production = prod_by_barn_day.get(key, 0)
            if production > 0:
                total_off += production * off_pct / 100.0
        except (ValueError, TypeError, KeyError):
            continue
    if not seen_any:
        return None
    return round(total_off, 2)


def _safe_float(val, default=0):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _parse_bb_date(bb_value):
    """Parse BB value to datetime.date for grouping; returns None if invalid."""
    if bb_value is None or (isinstance(bb_value, float) and pd.isna(bb_value)):
        return None
    try:
        if hasattr(bb_value, "date"):
            return bb_value.date()
        if hasattr(bb_value, "to_pydatetime"):
            return bb_value.to_pydatetime().date()
        dt = pd.to_datetime(bb_value).to_pydatetime()
        return dt.date() if hasattr(dt, "date") else dt
    except Exception:
        return None


def _production_iso_to_report_week(iso_year, iso_week):
    """
    Map (iso_year, iso_week) to our report week number.
    2025 -> 1-52; 2026 ISO week 1 -> 56, 2026 week 2 -> 57, ... 2026 week 6 -> 61.
    Used so Shipped_Orders (Option A) counts boxes whose BB date maps to this report week.
    """
    if iso_year == 2025 and 1 <= iso_week <= 52:
        return iso_week
    if iso_year == 2026 and 1 <= iso_week <= 6:
        return 55 + iso_week  # 56, 57, 58, 59, 60, 61
    return None


def _bb_date_to_production_iso_week(bb_value):
    """
    Compute production week from BB date: PackDate = BBDate - 28 days, then ISO week of PackDate.
    Returns (iso_year, iso_week) or None if invalid.
    """
    if bb_value is None or (isinstance(bb_value, float) and pd.isna(bb_value)):
        return None
    dt = None
    if hasattr(bb_value, "date"):
        dt = bb_value
    elif isinstance(bb_value, str) and len(bb_value) >= 8:
        try:
            dt = pd.to_datetime(bb_value).to_pydatetime()
        except Exception:
            return None
    else:
        try:
            dt = pd.to_datetime(bb_value).to_pydatetime()
        except Exception:
            return None
    if dt is None or not isinstance(dt, datetime):
        return None
    pack_date = dt - timedelta(days=_get_bb_to_pack_days())
    iso_year, iso_week, _ = pack_date.isocalendar()
    return (iso_year, iso_week)


def get_dominant_production_week_from_slip(slip_week, exports_dir=None):
    """
    Read PalletLines for the given slip week; from BB dates (BB-28 = pack date) compute
    the dominant (iso_year, iso_week) by case-equivalent boxes. Use this to match
    production from Hilly Acres (which uses week numbers 1-53 for 2025, 1-8 etc for 2026).
    Returns (iso_year, iso_week) or None if no PalletLines or no BB dates.
    """
    if exports_dir is None:
        exports_dir = EXPORTS_DIR
    exports_dir = Path(exports_dir)
    pattern = f"Week{slip_week}_AllDays_PalletLines_*.xlsx"
    files = sorted(exports_dir.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        return None
    try:
        df = pd.read_excel(files[0], sheet_name="PalletLines")
    except Exception:
        return None
    if "BBDate" not in df.columns or "QtyBoxes" not in df.columns:
        return None
    by_key = {}  # (iso_year, iso_week) -> case-equiv sum
    for _, row in df.iterrows():
        prod = _bb_date_to_production_iso_week(row.get("BBDate"))
        if prod is None:
            continue
        iy, iw = prod
        qty = _case_equivalent_boxes(_safe_float(row.get("QtyBoxes"), 0), row.get("SKU"))
        by_key[(iy, iw)] = by_key.get((iy, iw), 0) + qty
    if not by_key:
        return None
    dominant = max(by_key.items(), key=lambda x: x[1])
    return dominant[0]  # (iso_year, iso_week)


def get_production_for_report_week(week_number, row, exports_dir, slip_week_override=None):
    """
    Get production for traceability: prefer CSV; then if HillyAcresWeek set use that HA week;
    then match by week ending (Friday + 1 = Hilly Acres Week Ending);
    then Hilly Acres by report week number; then by dominant production week from slip BB dates.
    """
    prod_csv = _num(_get_row(row, "Production"), default=None)
    if prod_csv is not None and prod_csv > 0:
        return prod_csv
    ha_week = _get_hilly_acres_week_override(row)
    if ha_week is not None:
        production = get_production_from_hilly_acres_override(week_number, ha_week)
        if production is not None and production > 0:
            return production
    slip_week = slip_week_override if (slip_week_override is not None and str(slip_week_override).strip() != "" and not (isinstance(slip_week_override, float) and pd.isna(slip_week_override))) else week_number
    # 1) Match by week ending: slip Friday ship date + 1 = Hilly Acres "Week Ending" (Inputs row 3 col C)
    try:
        from process_weekly_loading_slip import get_slip_week_ending
        import hilly_acres_production as ha
        we = get_slip_week_ending(slip_week)
        if we is not None:
            production = ha.get_production_for_week_ending(we)
            if production is not None and production > 0:
                return production
    except Exception:
        pass
    # 2) Try report/slip week number (e.g. week 56)
    production = get_production_from_hilly_acres(week_number)
    if production is not None and production > 0:
        return production
    # 3) Derive production week from slip BB dates (dominant pack-week)
    prod_week_key = get_dominant_production_week_from_slip(slip_week, exports_dir)
    if prod_week_key is not None:
        _, prod_week_iso = prod_week_key
        production = get_production_from_hilly_acres(prod_week_iso)
        if production is not None and production > 0:
            return production
    return prod_csv if prod_csv is not None else 0


def _get_hilly_acres_week_override(row):
    """If CSV has HillyAcresWeek (report week maps to different HA week, e.g. 57 -> 2), return it."""
    v = _get_row(row, "HillyAcresWeek", None)
    if v is None or (isinstance(v, float) and pd.isna(v)) or str(v).strip() == "":
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def get_nest_run_for_report_week(week_number, row, slip_week_override=None):
    """
    Get Shipped_NestRun (nest run boxes) for the report week.
    Prefer CSV (SQF_Traceability_Inputs.csv Shipped_NestRun). If CSV is 0 or blank,
    try Hilly Acres Pallet Information sheet (NR Total by barn, summed).
    If row has HillyAcresWeek (e.g. 2), use that week number for HA lookup (report week 57 = HA week 2 2026).
    """
    nest_csv = _num(_get_row(row, "Shipped_NestRun"), default=None)
    if nest_csv is not None and nest_csv > 0:
        return nest_csv
    ha_week = _get_hilly_acres_week_override(row)
    try:
        if ha_week is not None:
            nr = get_nest_run_from_hilly_acres_override(week_number, ha_week)
            if nr is not None and nr > 0:
                return nr
        import hilly_acres_production as ha
        slip_week = slip_week_override if (slip_week_override is not None and str(slip_week_override).strip() != "" and not (isinstance(slip_week_override, float) and pd.isna(slip_week_override))) else week_number
        from process_weekly_loading_slip import get_slip_week_ending
        we = get_slip_week_ending(slip_week)
        if we is not None:
            nr = ha.get_nest_run_boxes_for_week_ending(we)
            if nr is not None and nr > 0:
                return nr
        nr = ha.get_nest_run_boxes_for_week(week_number)
        if nr is not None and nr > 0:
            return nr
    except Exception:
        pass
    return nest_csv if nest_csv is not None else 0


def _get_order_attribution_mode():
    """Read OrderAttributionMode from paths.json. 'slip_day' = use ReportWeek from PalletLines (aligns with Hilly Acres). 'bb_date' = use BB date."""
    try:
        with open(REF_DATA_DIR / "paths.json", encoding="utf-8") as f:
            m = json.load(f).get("OrderAttributionMode")
        if m and str(m).strip().lower() in ("slip_day", "slipday", "slip"):
            return "slip_day"
    except Exception:
        pass
    return "bb_date"


def _slip_day_to_report_week(slip_week, day_name):
    """Same logic as process_weekly_loading_slip: Wed/Thu/Fri->N, Mon/Tue->N-1."""
    try:
        sw = int(float(slip_week))
    except (TypeError, ValueError):
        return None
    day = str(day_name or "").strip()
    if day in ("Wednesday", "Thursday", "Friday", "Wed", "Thu", "Fri"):
        return sw
    if day in ("Monday", "Tuesday", "Mon", "Tue", "Tues"):
        rw = sw - 1
        if rw < 1:
            rw = 52
        elif rw == 55:
            rw = 52
        return rw
    return sw


def get_shipped_orders_boxes(week_number, exports_dir=None, by_slip_week=False, slip_week_override=None, palletlines_source=PALLETLINES_SOURCE_PREFERRED):
    """
    Sum case-equivalent boxes from PalletLines for week N.
    30-pack Nova/Loblaws only = 8×30 eggs per box (4/3 case-equiv); Jumbo SKUs = 12/15 case-equiv per box; else 1 per box.

    If by_slip_week=True: sum ALL (case-equiv) from the PalletLines file.
    Use slip_week_override (e.g. from CSV column SlipWeek) when the slip file is labeled
    a different week than the report week (e.g. report Week 30 uses slip Week 31).

    If by_slip_week=False: attribution by OrderAttributionMode (paths.json):
    - slip_day: ReportWeek from PalletLines (Wed/Thu/Fri of slip N + Mon/Tue of slip N+1 = Report N). Aligns with Hilly Acres.
    - bb_date: BBDate maps to production week (legacy).
    """
    if exports_dir is None:
        exports_dir = EXPORTS_DIR
    exports_dir = Path(exports_dir)
    if not exports_dir.exists():
        return None
    total = 0
    found_any = False
    use_slip_day = _get_order_attribution_mode() == "slip_day" and not by_slip_week
    if use_slip_day:
        totals = _get_slip_day_order_totals(exports_dir, source_mode=palletlines_source)
        return int(round(totals.get(week_number, 0))) if totals else None
    elif by_slip_week:
        sw = slip_week_override
        slip_weeks = (int(sw),) if sw is not None and str(sw).strip() != "" and not (isinstance(sw, float) and pd.isna(sw)) else (week_number,)
    else:
        slip_weeks = (week_number, week_number + 1)
    for slip_week in slip_weeks:
        df, _ = _load_palletlines_df(slip_week, exports_dir, source_mode=palletlines_source)
        if df is None:
            continue
        try:
            if "QtyBoxes" not in df.columns:
                continue
            found_any = True
            if by_slip_week or (not use_slip_day and "BBDate" not in df.columns):
                for _, row in df.iterrows():
                    total += _case_equivalent_boxes(_safe_float(row.get("QtyBoxes"), 0), row.get("SKU"))
                continue
            if use_slip_day:
                for _, row in df.iterrows():
                    rw = row.get("ReportWeek")
                    if rw is None or (isinstance(rw, float) and pd.isna(rw)):
                        rw = _slip_day_to_report_week(row.get("WeekNumber"), row.get("DayName"))
                    if rw is not None and int(rw) == week_number:
                        total += _case_equivalent_boxes(_safe_float(row.get("QtyBoxes"), 0), row.get("SKU"))
                continue
            for _, row in df.iterrows():
                prod = _bb_date_to_production_iso_week(row.get("BBDate"))
                if prod is None:
                    continue
                iy, iw = prod
                report_week_for_row = _production_iso_to_report_week(iy, iw)
                if report_week_for_row == week_number:
                    total += _case_equivalent_boxes(_safe_float(row.get("QtyBoxes"), 0), row.get("SKU"))
        except Exception:
            continue
    return int(round(total)) if found_any else None


def get_orders_by_day_and_bb(report_week, exports_dir=None, slip_week_override=None, palletlines_source=PALLETLINES_SOURCE_PREFERRED):
    """
    For a 2025 report week, load PalletLines (same slip weeks as Option A) and return per-day
    breakdown: how many case-equiv went to Old Date (BB -> prior production week) vs New Date
    (BB -> this week) vs Other. Use for audit: when a week is >100% you can reallocate Old Date
    to the previous week.
    Returns list of dicts: Day, Cases_OldDate, Cases_NewDate, Cases_Other, Total.
    Only applies to 2025 report weeks 1-52.
    """
    if report_week is None or report_week < 1 or report_week > 52:
        return []
    if exports_dir is None:
        exports_dir = EXPORTS_DIR
    exports_dir = Path(exports_dir)
    if not exports_dir.exists():
        return []
    slip_weeks = (int(slip_week_override),) if slip_week_override is not None and str(slip_week_override).strip() != "" and not (isinstance(slip_week_override, float) and pd.isna(slip_week_override)) else (report_week, report_week + 1)
    prior_report_week = report_week - 1 if report_week >= 2 else 52  # week 1: prior = 52

    # Aggregate by day: (day_name) -> { "prior": 0, "this": 0, "other": 0 }
    by_day = {}
    day_col = "DayName"

    for slip_week in slip_weeks:
        df, _ = _load_palletlines_df(slip_week, exports_dir, source_mode=palletlines_source)
        if df is None:
            continue
        if "QtyBoxes" not in df.columns or "BBDate" not in df.columns:
            continue
        day_col_use = "DayName" if "DayName" in df.columns else ("Day" if "Day" in df.columns else None)
        if day_col_use is None:
            day_col_use = "DayName"  # will yield Unknown when missing

        for _, row in df.iterrows():
            prod = _bb_date_to_production_iso_week(row.get("BBDate"))
            qty = _case_equivalent_boxes(_safe_float(row.get("QtyBoxes"), 0), row.get("SKU"))
            if qty <= 0:
                continue
            day_val = row.get(day_col_use) if day_col_use in df.columns else "Unknown"
            if pd.isna(day_val) or str(day_val).strip() == "":
                day_val = "Unknown"
            day_str = str(day_val).strip()
            if day_str not in by_day:
                by_day[day_str] = {"prior": 0, "this": 0, "other": 0}

            if prod is None:
                by_day[day_str]["other"] += qty
                continue
            rw = _production_iso_to_report_week(prod[0], prod[1])
            if rw == prior_report_week:
                by_day[day_str]["prior"] += qty
            elif rw == report_week:
                by_day[day_str]["this"] += qty
            else:
                by_day[day_str]["other"] += qty

    # Order days Mon->Sun then Unknown
    day_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday", "NFLD", "Unknown"]
    seen = set()
    ordered_days = [d for d in day_order if d in by_day and d not in seen and not seen.add(d)]
    for d in sorted(by_day.keys()):
        if d not in ordered_days:
            ordered_days.append(d)

    out = []
    for d in ordered_days:
        v = by_day[d]
        total = v["prior"] + v["this"] + v["other"]
        out.append({
            "Day": d,
            "Cases_OldDate": int(v["prior"]),
            "Cases_NewDate": int(v["this"]),
            "Cases_Other": int(v["other"]),
            "Total": int(total),
        })
    return out


def _day_sort_key(day_name):
    """Sort days in audit-friendly order."""
    order = {
        "Monday": 1,
        "Tuesday": 2,
        "Wednesday": 3,
        "Thursday": 4,
        "Friday": 5,
        "Saturday": 6,
        "Sunday": 7,
        "NFLD": 8,
        "Unknown": 9,
    }
    return order.get(str(day_name).strip(), 99)


def build_all_2025_old_new_by_day(exports_dir=None, inputs_df=None):
    """
    Build a single table: Week, Day, Cases_OldDate, Cases_NewDate, Cases_Other, Total
    for all 2025 report weeks (1-52) that have PalletLines. Uses SlipWeek from CSV when present.
    """
    if exports_dir is None:
        exports_dir = EXPORTS_DIR
    exports_dir = Path(exports_dir)
    if inputs_df is None:
        inputs_df = load_sqf_inputs()
    rows = []
    for report_week in range(1, 53):
        slip_override = None
        if inputs_df is not None and not inputs_df.empty and "Week" in inputs_df.columns and "SlipWeek" in inputs_df.columns:
            match = inputs_df[inputs_df["Week"] == report_week]
            if not match.empty:
                v = match.iloc[0].get("SlipWeek")
                if v is not None and str(v).strip() != "" and not (isinstance(v, float) and pd.isna(v)):
                    slip_override = v
        day_rows = get_orders_by_day_and_bb(report_week, exports_dir, slip_week_override=slip_override, palletlines_source=PALLETLINES_SOURCE_ORIGINAL_ONLY)
        for dr in day_rows:
            rows.append({
                "Week": report_week,
                "Day": dr["Day"],
                "Cases_OldDate": dr["Cases_OldDate"],
                "Cases_NewDate": dr["Cases_NewDate"],
                "Cases_Other": dr["Cases_Other"],
                "Total": dr["Total"],
            })
    if not rows:
        return pd.DataFrame(columns=["Week", "Day", "Cases_OldDate", "Cases_NewDate", "Cases_Other", "Total"])
    return pd.DataFrame(rows)


def build_all_2025_old_new_by_day_sku(exports_dir=None, inputs_df=None):
    """
    Build a detailed audit table: Week, Day, SKU, Cases_OldDate, Cases_NewDate, Cases_Other, Total.
    Rows are sorted so biggest SKUs for each day appear first.
    """
    if exports_dir is None:
        exports_dir = EXPORTS_DIR
    exports_dir = Path(exports_dir)
    if inputs_df is None:
        inputs_df = load_sqf_inputs()
    rows = []
    for report_week in range(1, 53):
        slip_override = None
        if inputs_df is not None and not inputs_df.empty and "Week" in inputs_df.columns and "SlipWeek" in inputs_df.columns:
            match = inputs_df[inputs_df["Week"] == report_week]
            if not match.empty:
                v = match.iloc[0].get("SlipWeek")
                if v is not None and str(v).strip() != "" and not (isinstance(v, float) and pd.isna(v)):
                    slip_override = v
        slip_weeks = (
            (int(slip_override),)
            if slip_override is not None and str(slip_override).strip() != "" and not (isinstance(slip_override, float) and pd.isna(slip_override))
            else (report_week, report_week + 1)
        )
        prior_report_week = report_week - 1 if report_week >= 2 else 52
        by_key = {}
        for slip_week in slip_weeks:
            df, _ = _load_palletlines_df(slip_week, exports_dir, source_mode=PALLETLINES_SOURCE_ORIGINAL_ONLY)
            if df is None:
                continue
            if "QtyBoxes" not in df.columns or "BBDate" not in df.columns:
                continue
            day_col_use = "DayName" if "DayName" in df.columns else ("Day" if "Day" in df.columns else None)
            for _, row in df.iterrows():
                qty = _case_equivalent_boxes(_safe_float(row.get("QtyBoxes"), 0), row.get("SKU"))
                if qty <= 0:
                    continue
                day_val = row.get(day_col_use) if day_col_use in df.columns else "Unknown"
                if pd.isna(day_val) or str(day_val).strip() == "":
                    day_val = "Unknown"
                day_str = str(day_val).strip()
                sku = str(row.get("SKU") or "").strip() or "Unknown SKU"
                key = (report_week, day_str, sku)
                if key not in by_key:
                    by_key[key] = {"prior": 0, "this": 0, "other": 0}
                prod = _bb_date_to_production_iso_week(row.get("BBDate"))
                if prod is None:
                    by_key[key]["other"] += qty
                    continue
                rw = _production_iso_to_report_week(prod[0], prod[1])
                if rw == prior_report_week:
                    by_key[key]["prior"] += qty
                elif rw == report_week:
                    by_key[key]["this"] += qty
                else:
                    by_key[key]["other"] += qty
        for (week_num, day_name, sku), vals in by_key.items():
            total = vals["prior"] + vals["this"] + vals["other"]
            rows.append({
                "Week": week_num,
                "Day": day_name,
                "SKU": sku,
                "Cases_OldDate": int(vals["prior"]),
                "Cases_NewDate": int(vals["this"]),
                "Cases_Other": int(vals["other"]),
                "Total": int(total),
            })
    if not rows:
        return pd.DataFrame(columns=["Week", "Day", "SKU", "Cases_OldDate", "Cases_NewDate", "Cases_Other", "Total"])
    df = pd.DataFrame(rows)
    df["_day_sort"] = df["Day"].map(_day_sort_key)
    df = df.sort_values(["Week", "_day_sort", "Total", "SKU"], ascending=[True, True, False, True]).reset_index(drop=True)
    df["Rank_In_Day"] = df.groupby(["Week", "Day"]).cumcount() + 1
    cols = ["Week", "Day", "Rank_In_Day", "SKU", "Cases_OldDate", "Cases_NewDate", "Cases_Other", "Total"]
    return df[cols]


def build_2025_reallocation_guide(old_new_by_day_df, old_new_by_day_sku_df, top_n=5):
    """
    Compact audit guide. One row per (Week, Day) with top old-date SKUs called out.
    """
    if old_new_by_day_df is None or old_new_by_day_df.empty:
        return pd.DataFrame(columns=[
            "Week", "Day", "Cases_OldDate", "Cases_NewDate", "Cases_Other", "Total",
            "Max_Move_To_Prior_Week", "Top_OldDate_SKUs"
        ])
    sku_df = old_new_by_day_sku_df if old_new_by_day_sku_df is not None else pd.DataFrame()
    rows = []
    for _, row in old_new_by_day_df.iterrows():
        week = int(row["Week"])
        day = row["Day"]
        top_skus = ""
        if not sku_df.empty:
            subset = sku_df[(sku_df["Week"] == week) & (sku_df["Day"] == day) & (sku_df["Cases_OldDate"] > 0)]
            if not subset.empty:
                subset = subset.sort_values(["Cases_OldDate", "Total", "SKU"], ascending=[False, False, True]).head(top_n)
                top_skus = " | ".join(f"{r['SKU']}={int(r['Cases_OldDate'])}" for _, r in subset.iterrows())
        rows.append({
            "Week": week,
            "Day": day,
            "Cases_OldDate": int(row["Cases_OldDate"]),
            "Cases_NewDate": int(row["Cases_NewDate"]),
            "Cases_Other": int(row["Cases_Other"]),
            "Total": int(row["Total"]),
            "Max_Move_To_Prior_Week": int(row["Cases_OldDate"]),
            "Top_OldDate_SKUs": top_skus,
        })
    out = pd.DataFrame(rows)
    out["_day_sort"] = out["Day"].map(_day_sort_key)
    out = out.sort_values(["Week", "_day_sort"], ascending=[True, True]).reset_index(drop=True)
    return out.drop(columns=["_day_sort"])


def _display_bb_date(bb_value):
    """Return BB date as YYYY-MM-DD string when possible."""
    parsed = _parse_bb_date(bb_value)
    if parsed is not None:
        return parsed.strftime("%Y-%m-%d")
    if bb_value is None or (isinstance(bb_value, float) and pd.isna(bb_value)):
        return ""
    return str(bb_value).strip()


def _shift_bb_date_string(bb_value, days):
    """Return shifted BB date string, or blank if BB date is invalid."""
    parsed = _parse_bb_date(bb_value)
    if parsed is None:
        return ""
    return (parsed + timedelta(days=days)).strftime("%Y-%m-%d")


def _current_report_week_for_line(row, slip_week):
    """Return the current report week used for a PalletLines row."""
    rw = row.get("ReportWeek")
    if rw is None or (isinstance(rw, float) and pd.isna(rw)):
        rw = _slip_day_to_report_week(row.get("WeekNumber") or slip_week, row.get("DayName") or row.get("Day"))
    try:
        return int(float(rw)) if rw is not None and str(rw).strip() != "" else None
    except (TypeError, ValueError):
        return None


def build_balancing_candidates(
    results,
    exports_dir=None,
    include_heuristic=True,
    direction="back",
    palletlines_source=PALLETLINES_SOURCE_ORIGINAL_ONLY,
):
    """
    Build line-level balancing candidates from original PalletLines.
    Candidate priority:
      1) Existing BB-date evidence already points to the adjacent week.
      2) Optional NFLD heuristic BB-date shifts by +/-7 days to the adjacent week.
    """
    exports_dir = Path(exports_dir) if exports_dir else EXPORTS_DIR
    if not exports_dir.exists():
        return pd.DataFrame()
    week_meta = {}
    for r in results or []:
        try:
            week_num = int(r.get("Week") or 0)
        except Exception:
            continue
        if week_num > 0:
            week_meta[week_num] = {
                "Year": r.get("Year"),
                "Production": int(r.get("Production") or 0),
                "Accounted": int(round(r.get("Accounted") or 0)),
                "Traceability_Pct": r.get("Traceability_Pct"),
            }
    direction = str(direction).strip().lower()
    allow_back = direction in ("back", "both", "prev", "previous")
    allow_fwd = direction in ("forward", "both", "next", "nextweek", "fwd", "future")

    rows = []
    seen_slip_weeks = set()
    for slip_week in _report_week_sequence():
        df, source_path = _load_palletlines_df(slip_week, exports_dir, source_mode=palletlines_source)
        if df is None or source_path is None:
            continue
        seen_slip_weeks.add(slip_week)
        if "QtyBoxes" not in df.columns:
            continue
        for idx, row in df.reset_index(drop=True).iterrows():
            qty_boxes = int(_safe_float(row.get("QtyBoxes"), 0))
            if qty_boxes <= 0:
                continue
            sku = str(row.get("SKU") or "").strip()
            day_name = str(row.get("DayName") or row.get("Day") or "").strip()
            stop_name = str(row.get("StopName") or row.get("Stop") or "").strip()
            current_report_week = _current_report_week_for_line(row, slip_week)
            if current_report_week is None or current_report_week not in week_meta:
                continue
            case_factor = _case_factor_for_sku(sku)
            case_equiv = qty_boxes * case_factor
            bb_prod = _bb_date_to_production_iso_week(row.get("BBDate"))
            bb_report_week = _production_iso_to_report_week(bb_prod[0], bb_prod[1]) if bb_prod is not None else None
            is_nfld = day_name.upper() == "NFLD"
            common = {
                "Slip_Week": slip_week,
                "Source_File": _display_path(source_path),
                "Source_Row_Index": idx,
                "LineID": int(row.get("LineID", idx + 1)),
                "DayName": day_name,
                "StopName": stop_name,
                "SKU": sku,
                "QtyBoxes": qty_boxes,
                "CaseFactor": case_factor,
                "Max_CaseEquiv_Movable": case_equiv,
                "Current_ReportWeek": current_report_week,
                "Original_BBDate": _display_bb_date(row.get("BBDate")),
                "Is_NFLD": "Yes" if is_nfld else "No",
            }
            added_bb_alignment = False

            prev_week = _previous_report_week(current_report_week)
            next_week = _next_report_week(current_report_week)

            # Existing BB-date evidence already maps this line to the adjacent week.
            if bb_report_week is not None and bb_report_week in week_meta:
                if allow_back and prev_week is not None and bb_report_week == prev_week:
                    rows.append({
                        **common,
                        "Target_ReportWeek": bb_report_week,
                        "Suggested_BBDate": common["Original_BBDate"],
                        "Candidate_Type": "existing_bb_alignment_back",
                        "Adjustment_Action": "reportweek_only",
                        "Priority_Rank": 0 if is_nfld else 1,
                        "Reason": f"BB date already maps this line to prior report week {bb_report_week}.",
                    })
                    added_bb_alignment = True
                elif allow_fwd and next_week is not None and bb_report_week == next_week:
                    rows.append({
                        **common,
                        "Target_ReportWeek": bb_report_week,
                        "Suggested_BBDate": common["Original_BBDate"],
                        "Candidate_Type": "existing_bb_alignment_forward",
                        "Adjustment_Action": "reportweek_only",
                        "Priority_Rank": 0 if is_nfld else 1,
                        "Reason": f"BB date already maps this line to next report week {bb_report_week}.",
                    })
                    added_bb_alignment = True

            # Optional NFLD heuristic: shift BB date by +/-7 to the adjacent week.
            if include_heuristic and is_nfld:
                if allow_back and prev_week is not None and prev_week in week_meta and prev_week != current_report_week:
                    suggested_bb = _shift_bb_date_string(row.get("BBDate"), -7)
                    if suggested_bb:
                        rows.append({
                            **common,
                            "Target_ReportWeek": prev_week,
                            "Suggested_BBDate": suggested_bb,
                            "Candidate_Type": "nfld_bb_shift_back",
                            "Adjustment_Action": "bbdate_and_reportweek",
                            "Priority_Rank": 2,
                            "Reason": f"NFLD heuristic shift of BB date by -7 days to prior report week {prev_week}.",
                        })
                if allow_fwd and next_week is not None and next_week in week_meta and next_week != current_report_week:
                    suggested_bb = _shift_bb_date_string(row.get("BBDate"), +7)
                    if suggested_bb:
                        rows.append({
                            **common,
                            "Target_ReportWeek": next_week,
                            "Suggested_BBDate": suggested_bb,
                            "Candidate_Type": "nfld_bb_shift_forward",
                            "Adjustment_Action": "bbdate_and_reportweek",
                            "Priority_Rank": 2,
                            "Reason": f"NFLD heuristic shift of BB date by +7 days to next report week {next_week}.",
                        })

            # General reallocation candidate:
            # - allow back: target previous report week (shift -7)
            # - allow forward: target next report week (shift +7)
            # The balancing plan decides whether to actually use it based on global week gaps.
            if not added_bb_alignment:
                if allow_back and prev_week is not None and prev_week in week_meta and prev_week != current_report_week:
                    suggested_bb = _shift_bb_date_string(row.get("BBDate"), -7) or common["Original_BBDate"]
                    rows.append({
                        **common,
                        "Target_ReportWeek": prev_week,
                        "Suggested_BBDate": suggested_bb,
                        "Candidate_Type": "one_week_back_reallocate",
                        "Adjustment_Action": "bbdate_and_reportweek",
                        "Priority_Rank": 3,
                        "Reason": "General reallocation candidate capped at one week back.",
                    })
                if allow_fwd and next_week is not None and next_week in week_meta and next_week != current_report_week:
                    suggested_bb = _shift_bb_date_string(row.get("BBDate"), +7) or common["Original_BBDate"]
                    rows.append({
                        **common,
                        "Target_ReportWeek": next_week,
                        "Suggested_BBDate": suggested_bb,
                        "Candidate_Type": "one_week_forward_reallocate",
                        "Adjustment_Action": "bbdate_and_reportweek",
                        "Priority_Rank": 3,
                        "Reason": "General reallocation candidate capped at one week forward.",
                    })
    if not rows:
        return pd.DataFrame(columns=[
            "Slip_Week", "Source_File", "Source_Row_Index", "LineID", "DayName", "StopName", "SKU",
            "QtyBoxes", "CaseFactor", "Max_CaseEquiv_Movable", "Current_ReportWeek", "Target_ReportWeek",
            "Original_BBDate", "Suggested_BBDate", "Candidate_Type", "Adjustment_Action", "Is_NFLD",
            "Priority_Rank", "Reason"
        ])
    out = pd.DataFrame(rows)
    out = out.sort_values(
        ["Priority_Rank", "Is_NFLD", "Max_CaseEquiv_Movable", "Slip_Week", "DayName", "SKU", "Target_ReportWeek"],
        ascending=[True, False, False, True, True, True, True],
    ).reset_index(drop=True)
    out["Rank_Overall"] = range(1, len(out) + 1)
    return out


def build_balancing_plan(
    results,
    exports_dir=None,
    include_heuristic=True,
    aggressive=False,
    allow_worsen_cases=0,
    soft_floor_cases_per_week=0,
    hard_floor_sendback_cases_per_week=0,
    direction="back",
    palletlines_source=PALLETLINES_SOURCE_ORIGINAL_ONLY,
):
    """
    Select the best line-level moves to bring weeks closer to 100%.
    Uses original PalletLines as the immutable source and returns:
      candidates_df, selected_adjustments_df, balancing_by_week_df
    """
    candidates_df = build_balancing_candidates(
        results,
        exports_dir=exports_dir,
        include_heuristic=include_heuristic,
        direction=direction,
        palletlines_source=palletlines_source,
    )
    if candidates_df.empty:
        return candidates_df, pd.DataFrame(), pd.DataFrame()
    baseline = {}
    for r in results or []:
        try:
            week = int(r.get("Week") or 0)
        except Exception:
            continue
        production = int(r.get("Production") or 0)
        if week <= 0 or production <= 0:
            continue
        accounted = int(round(r.get("Accounted") or 0))
        baseline[week] = {
            "Year": r.get("Year"),
            "Production": production,
            "Accounted_Before": accounted,
            "Traceability_Pct_Before": r.get("Traceability_Pct"),
        }
    if not baseline:
        return candidates_df, pd.DataFrame(), pd.DataFrame()
    # Production and current accounted are fixed inputs; moves are "from Week -> to Week".
    moved_out = {w: 0 for w in baseline}
    moved_in = {w: 0 for w in baseline}
    used_sources = set()
    selected_rows = []

    # Current accounted is updated as we accept moves in the aggressive mode.
    acc = {w: int(v["Accounted_Before"]) for w, v in baseline.items()}

    if not aggressive:
        surplus = {w: max(0, v["Accounted_Before"] - v["Production"]) for w, v in baseline.items()}
        deficit = {w: max(0, v["Production"] - v["Accounted_Before"]) for w, v in baseline.items()}

        for _, cand in candidates_df.iterrows():
            from_week = int(cand["Current_ReportWeek"])
            to_week = int(cand["Target_ReportWeek"])
            source_key = (int(cand["Slip_Week"]), int(cand["Source_Row_Index"]))
            if source_key in used_sources:
                continue
            if from_week not in baseline or to_week not in baseline:
                continue
            if surplus.get(from_week, 0) <= 0 or deficit.get(to_week, 0) <= 0:
                continue
            max_cases = float(cand["Max_CaseEquiv_Movable"])
            case_factor = float(cand["CaseFactor"])
            if case_factor <= 0:
                continue
            requested_cases = min(max_cases, surplus[from_week], deficit[to_week])
            move_boxes = int(requested_cases // case_factor)
            if move_boxes <= 0:
                continue
            actual_cases = move_boxes * case_factor
            used_sources.add(source_key)
            surplus[from_week] -= actual_cases
            deficit[to_week] -= actual_cases
            moved_out[from_week] += actual_cases
            moved_in[to_week] += actual_cases
            selected_rows.append({
                **cand.to_dict(),
                "Moved_QtyBoxes": move_boxes,
                "Moved_CaseEquiv": actual_cases,
                "Selection_Mode": "gap_match",
                "From_Week_Pct_Before": baseline[from_week]["Traceability_Pct_Before"],
                "To_Week_Pct_Before": baseline[to_week]["Traceability_Pct_Before"],
            })

        # Soft floor pass (non-aggressive only):
        # Try to move at least N cases into each deficit week using available back candidates,
        # while preserving the standard "source surplus + target deficit" guardrails.
        floor_target = float(soft_floor_cases_per_week or 0)
        if floor_target > 0:
            for to_week in sorted(baseline):
                if deficit.get(to_week, 0) <= 0:
                    continue
                if moved_in.get(to_week, 0) >= floor_target:
                    continue
                need = floor_target - moved_in.get(to_week, 0)
                if need <= 0:
                    continue
                week_cands = candidates_df[candidates_df["Target_ReportWeek"] == to_week]
                if week_cands.empty:
                    continue
                for _, cand in week_cands.iterrows():
                    if need <= 0 or deficit.get(to_week, 0) <= 0:
                        break
                    from_week = int(cand["Current_ReportWeek"])
                    source_key = (int(cand["Slip_Week"]), int(cand["Source_Row_Index"]))
                    if source_key in used_sources:
                        continue
                    if from_week not in baseline:
                        continue
                    if surplus.get(from_week, 0) <= 0:
                        continue
                    max_cases = float(cand["Max_CaseEquiv_Movable"])
                    case_factor = float(cand["CaseFactor"])
                    if case_factor <= 0:
                        continue
                    requested_cases = min(max_cases, surplus[from_week], deficit[to_week], need)
                    move_boxes = int(requested_cases // case_factor)
                    if move_boxes <= 0:
                        continue
                    actual_cases = move_boxes * case_factor
                    used_sources.add(source_key)
                    surplus[from_week] -= actual_cases
                    deficit[to_week] -= actual_cases
                    moved_out[from_week] += actual_cases
                    moved_in[to_week] += actual_cases
                    need -= actual_cases
                    selected_rows.append({
                        **cand.to_dict(),
                        "Moved_QtyBoxes": move_boxes,
                        "Moved_CaseEquiv": actual_cases,
                        "Selection_Mode": "soft_floor_topup",
                        "From_Week_Pct_Before": baseline[from_week]["Traceability_Pct_Before"],
                        "To_Week_Pct_Before": baseline[to_week]["Traceability_Pct_Before"],
                    })

        # Hard floor pass (forced send-back by source week):
        # For each week with back candidates, force at least N case-equiv to move OUT to prior week
        # when candidate rows have available quantity. This intentionally allows worsening fit to
        # represent known but historically under-tracked weekly send-back behavior.
        hard_floor = float(hard_floor_sendback_cases_per_week or 0)
        if hard_floor > 0:
            from_weeks_with_back = sorted(set(int(w) for w in candidates_df["Current_ReportWeek"].tolist()))
            for from_week in from_weeks_with_back:
                if from_week not in baseline:
                    continue
                already_out = float(moved_out.get(from_week, 0))
                need = hard_floor - already_out
                if need <= 0:
                    continue
                week_cands = candidates_df[candidates_df["Current_ReportWeek"] == from_week]
                if week_cands.empty:
                    continue
                for _, cand in week_cands.iterrows():
                    if need <= 0:
                        break
                    to_week = int(cand["Target_ReportWeek"])
                    if to_week not in baseline:
                        continue
                    source_key = (int(cand["Slip_Week"]), int(cand["Source_Row_Index"]))
                    if source_key in used_sources:
                        continue
                    max_cases = float(cand["Max_CaseEquiv_Movable"])
                    case_factor = float(cand["CaseFactor"])
                    if case_factor <= 0:
                        continue
                    requested_cases = min(max_cases, need)
                    move_boxes = int(requested_cases // case_factor)
                    if move_boxes <= 0:
                        continue
                    actual_cases = move_boxes * case_factor
                    used_sources.add(source_key)
                    moved_out[from_week] += actual_cases
                    moved_in[to_week] += actual_cases
                    need -= actual_cases
                    selected_rows.append({
                        **cand.to_dict(),
                        "Moved_QtyBoxes": move_boxes,
                        "Moved_CaseEquiv": actual_cases,
                        "Selection_Mode": "hard_floor_forced_sendback",
                        "From_Week_Pct_Before": baseline[from_week]["Traceability_Pct_Before"],
                        "To_Week_Pct_Before": baseline[to_week]["Traceability_Pct_Before"],
                    })
    else:
        # Aggressive mode:
        # - still only shifts "one report week back" (already encoded in candidates)
        # - but allows moving out of weeks that may not be "surplus"
        # - accepts a move if the overall abs-gap for the affected weeks does not worsen by more
        #   than `allow_worsen_cases`.
        for _, cand in candidates_df.iterrows():
            from_week = int(cand["Current_ReportWeek"])
            to_week = int(cand["Target_ReportWeek"])
            source_key = (int(cand["Slip_Week"]), int(cand["Source_Row_Index"]))
            if source_key in used_sources:
                continue
            if from_week not in baseline or to_week not in baseline:
                continue

            prod_from = int(baseline[from_week]["Production"])
            prod_to = int(baseline[to_week]["Production"])
            acc_from = int(acc[from_week])
            acc_to = int(acc[to_week])
            deficit_to = prod_to - acc_to
            if deficit_to <= 0:
                # Never move *into* an already-surplus week.
                continue

            max_cases = float(cand["Max_CaseEquiv_Movable"])
            case_factor = float(cand["CaseFactor"])
            if case_factor <= 0:
                continue

            requested_cases = min(max_cases, deficit_to)
            move_boxes = int(requested_cases // case_factor)
            if move_boxes <= 0:
                continue
            actual_cases = move_boxes * case_factor
            if actual_cases <= 0:
                continue

            err_before = abs(prod_from - acc_from) + abs(prod_to - acc_to)
            err_after = abs(prod_from - (acc_from - actual_cases)) + abs(prod_to - (acc_to + actual_cases))

            if err_after > err_before + int(allow_worsen_cases or 0):
                continue

            used_sources.add(source_key)
            acc[from_week] -= actual_cases
            acc[to_week] += actual_cases
            moved_out[from_week] += actual_cases
            moved_in[to_week] += actual_cases
            selected_rows.append({
                **cand.to_dict(),
                "Moved_QtyBoxes": move_boxes,
                "Moved_CaseEquiv": actual_cases,
                "Selection_Mode": "aggressive_error_guard",
                "From_Week_Pct_Before": baseline[from_week]["Traceability_Pct_Before"],
                "To_Week_Pct_Before": baseline[to_week]["Traceability_Pct_Before"],
            })

    selected_df = pd.DataFrame(selected_rows)
    summary_rows = []
    for week in sorted(baseline):
        before_acc = int(baseline[week]["Accounted_Before"])
        after_acc = int(acc[week]) if aggressive else (before_acc - moved_out.get(week, 0) + moved_in.get(week, 0))
        production = int(baseline[week]["Production"])
        after_pct = round(100.0 * after_acc / production, 2) if production > 0 else None
        summary_rows.append({
            "Year": baseline[week]["Year"],
            "Week": week,
            "Production": production,
            "Accounted_Before": before_acc,
            "Traceability_Pct_Before": baseline[week]["Traceability_Pct_Before"],
            "Cases_Moved_Out": moved_out.get(week, 0),
            "Cases_Moved_In": moved_in.get(week, 0),
            "Accounted_After": after_acc,
            "Traceability_Pct_After": after_pct,
            "Distance_To_100_Before": abs(before_acc - production),
            "Distance_To_100_After": abs(after_acc - production),
        })
    return candidates_df, selected_df, pd.DataFrame(summary_rows)


def _apply_balancing_adjustments_to_df(df, adjustments_df):
    """Apply selected balancing adjustments to one PalletLines DataFrame."""
    if df is None or df.empty or adjustments_df is None or adjustments_df.empty:
        return df
    out = df.reset_index(drop=True).copy()
    new_rows = []
    next_line_id = int(out["LineID"].max()) + 1 if "LineID" in out.columns and not out.empty else len(out) + 1
    for _, adj in adjustments_df.sort_values(["Source_Row_Index", "LineID"]).iterrows():
        idx = int(adj["Source_Row_Index"])
        if idx < 0 or idx >= len(out):
            continue
        move_boxes = int(adj["Moved_QtyBoxes"])
        if move_boxes <= 0:
            continue
        row = out.loc[idx].copy()
        current_boxes = int(_safe_float(row.get("QtyBoxes"), 0))
        if current_boxes <= 0:
            continue
        target_week = int(adj["Target_ReportWeek"])
        current_week = int(adj["Current_ReportWeek"])
        moved_is_prev = 1 if target_week < current_week else 0
        moved_od_qty = move_boxes if moved_is_prev else 0
        if move_boxes >= current_boxes:
            out.at[idx, "BBDate"] = adj["Suggested_BBDate"]
            if "ReportWeek" in out.columns:
                out.at[idx, "ReportWeek"] = target_week
            if "IsPrevWeek" in out.columns:
                out.at[idx, "IsPrevWeek"] = moved_is_prev
            if "ODQty" in out.columns:
                out.at[idx, "ODQty"] = moved_od_qty
            continue
        out.at[idx, "QtyBoxes"] = current_boxes - move_boxes
        if "ODQty" in out.columns:
            original_od = int(_safe_float(row.get("ODQty"), 0))
            out.at[idx, "ODQty"] = max(0, min(current_boxes - move_boxes, original_od))
        moved_row = row.copy()
        if "LineID" in moved_row.index:
            moved_row["LineID"] = next_line_id
            next_line_id += 1
        moved_row["QtyBoxes"] = move_boxes
        moved_row["BBDate"] = adj["Suggested_BBDate"]
        if "ReportWeek" in moved_row.index:
            moved_row["ReportWeek"] = target_week
        if "IsPrevWeek" in moved_row.index:
            moved_row["IsPrevWeek"] = moved_is_prev
        if "ODQty" in moved_row.index:
            moved_row["ODQty"] = moved_od_qty
        new_rows.append(moved_row.to_dict())
    if new_rows:
        out = pd.concat([out, pd.DataFrame(new_rows)], ignore_index=True)
    return out


def load_balancing_logs_from_exports(exports_dir=None):
    """Aggregate balancing logs from the latest adjusted PalletLines workbooks."""
    exports_dir = Path(exports_dir) if exports_dir else EXPORTS_DIR
    latest_adjusted = []
    for slip_week in _report_week_sequence():
        path = _get_latest_palletlines_file(slip_week, exports_dir, source_mode=PALLETLINES_SOURCE_ADJUSTED_ONLY)
        if path is not None:
            latest_adjusted.append(path)
    if not latest_adjusted:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    adjustment_frames = []
    summary_frames = []
    candidate_frames = []
    for path in latest_adjusted:
        try:
            adjustment_frames.append(pd.read_excel(path, sheet_name="Adjustments_Log"))
        except Exception:
            pass
        try:
            summary_frames.append(pd.read_excel(path, sheet_name="Balancing_By_Week"))
        except Exception:
            pass
        try:
            candidate_frames.append(pd.read_excel(path, sheet_name="Balancing_Candidates"))
        except Exception:
            pass
    adjustments_df = pd.concat(adjustment_frames, ignore_index=True) if adjustment_frames else pd.DataFrame()
    summary_df = pd.concat(summary_frames, ignore_index=True).drop_duplicates() if summary_frames else pd.DataFrame()
    candidates_df = pd.concat(candidate_frames, ignore_index=True).drop_duplicates() if candidate_frames else pd.DataFrame()
    return adjustments_df, summary_df, candidates_df


def _original_palletlines_dir_for_balance():
    """
    Directory to read original PalletLines from when building the balancing plan.
    Prefer top-level Traceability_Exports (where loading-slip outputs live); fall back to Original/.
    """
    for d in (EXPORTS_DIR, ORIGINAL_EXPORTS_DIR):
        if d.exists():
            any_file = next(d.glob("Week*_AllDays_PalletLines*.xlsx"), None)
            if any_file and not _is_adjusted_palletlines_file(any_file):
                return d
    return EXPORTS_DIR


def generate_balanced_palletlines(
    exports_dir=None,
    include_heuristic=True,
    aggressive=False,
    allow_worsen_cases=0,
    soft_floor_cases_per_week=0,
    hard_floor_sendback_cases_per_week=0,
    direction="back",
    palletlines_source_for_balance=PALLETLINES_SOURCE_ORIGINAL_ONLY,
    balance_source_dir=None,
):
    """
    Create adjusted PalletLines workbooks from original PalletLines, preserving the originals.
    Returns dict with written files, candidate rows, applied adjustments, and by-week summary.

    Inputs (baselines and candidates) come from the composite original layer (top-level
    Traceability_Exports and optional Original/). Outputs are written into BALANCED_WORKING_DIR only.
    """
    # Where to WRITE adjusted / balanced files
    output_dir = Path(exports_dir) if exports_dir else BALANCED_WORKING_DIR
    output_dir.mkdir(parents=True, exist_ok=True)

    # Where to READ PalletLines from for the balancing plan (original vs iterative balancedworking).
    if balance_source_dir is not None:
        source_dir = Path(balance_source_dir)
    else:
        source_dir = _original_palletlines_dir_for_balance()

    # Ensure BalancedWorking already has coverage from previous runs / copies
    build_balanced_working_view()

    # Important: do NOT let manual adjustments influence the balancing plan.
    # The balancing step should be based on the base equation (production + shipped + nest run + offgrades),
    # and only the final traceability report should apply manual adjustments.
    baseline_results = run_all_weeks(
        source_dir,
        palletlines_source=palletlines_source_for_balance,
        apply_manual_adjustments=False,
    )
    candidates_df, selected_df, summary_df = build_balancing_plan(
        baseline_results,
        exports_dir=source_dir,
        include_heuristic=include_heuristic,
        aggressive=aggressive,
        allow_worsen_cases=allow_worsen_cases,
        soft_floor_cases_per_week=soft_floor_cases_per_week,
        hard_floor_sendback_cases_per_week=hard_floor_sendback_cases_per_week,
        direction=direction,
        palletlines_source=palletlines_source_for_balance,
    )
    if selected_df.empty:
        print("No balancing adjustments were selected from the original PalletLines files.")
        return {
            "files": [],
            "candidates": candidates_df,
            "adjustments": selected_df,
            "summary": summary_df,
        }
    try:
        from process_weekly_loading_slip import generate_summary_stats_pallet_lines, save_adjusted_pallet_lines_report
    except Exception as exc:
        raise RuntimeError(f"Could not import adjusted PalletLines writer: {exc}") from exc
    written_files = []
    for slip_week in sorted(selected_df["Slip_Week"].unique()):
        original_df, original_path = _load_palletlines_df(
            slip_week,
            source_dir,
            source_mode=palletlines_source_for_balance,
        )
        if original_df is None or original_path is None:
            continue
        week_adjustments = selected_df[selected_df["Slip_Week"] == slip_week].copy().reset_index(drop=True)
        adjusted_df = _apply_balancing_adjustments_to_df(original_df, week_adjustments)
        stats = generate_summary_stats_pallet_lines(adjusted_df)
        extra_sheets = {
            "Adjustments_Log": week_adjustments,
            "Balancing_By_Week": summary_df,
            "Balancing_Candidates": candidates_df[candidates_df["Slip_Week"] == slip_week].copy(),
            "Source_Trace": pd.DataFrame([{
                "Slip_Week": slip_week,
                "Original_File": _display_path(original_path),
                "Adjustment_Count": len(week_adjustments),
            }]),
        }
        path = save_adjusted_pallet_lines_report(
            adjusted_df,
            slip_week,
            stats,
            output_dir=output_dir,
            extra_sheets=extra_sheets,
        )
        written_files.append(path)
    _clear_palletlines_caches()
    return {
        "files": written_files,
        "candidates": candidates_df,
        "adjustments": selected_df,
        "summary": summary_df,
    }


def build_order_detail_all(exports_dir=None, inputs_df=None, palletlines_source=PALLETLINES_SOURCE_PREFERRED):
    """
    Build transparent audit table: every order line with Week (report week by BB date),
    StopName, SKU, QtyBoxes, BBDate, CaseEquiv. Sum of CaseEquiv per Week = Shipped_Orders.
    Filter to Week=N to see every box shipped that counts toward that week.
    Iterates over PalletLines files once each (no double-counting).
    """
    if exports_dir is None:
        exports_dir = EXPORTS_DIR
    exports_dir = Path(exports_dir)
    if not exports_dir.exists():
        return pd.DataFrame(columns=["Week", "LineID", "DayName", "StopName", "SKU", "QtyBoxes", "BBDate", "CaseEquiv"])
    rows = []
    all_files = list(exports_dir.glob("Week*_AllDays_PalletLines*.xlsx"))
    by_slip_week = {}
    for path in all_files:
        if palletlines_source == PALLETLINES_SOURCE_ORIGINAL_ONLY and _is_adjusted_palletlines_file(path):
            continue
        if palletlines_source == PALLETLINES_SOURCE_ADJUSTED_ONLY and not _is_adjusted_palletlines_file(path):
            continue
        name = path.stem
        parts = name.replace("Week", " ", 1).split("_")
        if not parts or not parts[0].strip().isdigit():
            continue
        slip_week = int(parts[0].strip())
        if slip_week < 1 or (53 <= slip_week <= 55) or slip_week > 61:
            continue
        if slip_week not in by_slip_week or path.stat().st_mtime > by_slip_week[slip_week].stat().st_mtime:
            by_slip_week[slip_week] = path
    for slip_week in sorted(by_slip_week.keys()):
        path = by_slip_week[slip_week]
        try:
            df = pd.read_excel(path, sheet_name="PalletLines")
        except Exception:
            continue
        if "QtyBoxes" not in df.columns:
            continue
        use_slip_day = _get_order_attribution_mode() == "slip_day"
        day_col = "DayName" if "DayName" in df.columns else ("Day" if "Day" in df.columns else None)
        stop_col = "StopName" if "StopName" in df.columns else ("Stop" if "Stop" in df.columns else None)
        for idx, row in df.iterrows():
            qty = _safe_float(row.get("QtyBoxes"), 0)
            if qty <= 0:
                continue
            case_equiv = _case_equivalent_boxes(qty, row.get("SKU"))
            if case_equiv <= 0:
                continue
            if use_slip_day:
                report_week = row.get("ReportWeek")
                if report_week is None or (isinstance(report_week, float) and pd.isna(report_week)):
                    report_week = _slip_day_to_report_week(row.get("WeekNumber") or slip_week, row.get("DayName") or row.get("Day"))
            else:
                prod = _bb_date_to_production_iso_week(row.get("BBDate"))
                if prod is None:
                    continue
                report_week = _production_iso_to_report_week(prod[0], prod[1])
            if report_week is None:
                continue
            day_val = row.get(day_col) if day_col else ""
            stop_val = row.get(stop_col) if stop_col else ""
            bb_val = row.get("BBDate")
            if hasattr(bb_val, "strftime"):
                bb_str = bb_val.strftime("%Y-%m-%d") if bb_val else ""
            else:
                bb_str = str(bb_val) if bb_val is not None and not (isinstance(bb_val, float) and pd.isna(bb_val)) else ""
            rows.append({
                "Week": report_week,
                "LineID": int(row.get("LineID", idx + 1)),
                "DayName": "" if pd.isna(day_val) else str(day_val).strip(),
                "StopName": "" if pd.isna(stop_val) else str(stop_val).strip(),
                "SKU": "" if pd.isna(row.get("SKU")) else str(row.get("SKU")).strip(),
                "QtyBoxes": int(qty),
                "BBDate": bb_str,
                "CaseEquiv": case_equiv,
            })
    if not rows:
        return pd.DataFrame(columns=["Week", "LineID", "DayName", "StopName", "SKU", "QtyBoxes", "BBDate", "CaseEquiv"])
    out = pd.DataFrame(rows)
    return out.sort_values(["Week", "DayName", "StopName", "SKU"]).reset_index(drop=True)


def build_week_reconciliation(results, order_detail_df=None, exports_dir=None, inputs_df=None):
    """
    Transparent reconciliation: for each week, Production + Nest Run + Off Grades + Orders (from detail) = Accounted.
    Orders_From_Detail = sum of CaseEquiv in Order_Detail where Week=N. Must equal Shipped_Orders.
    """
    if not results:
        return pd.DataFrame()
    if order_detail_df is None or order_detail_df.empty:
        order_detail_df = build_order_detail_all(exports_dir, inputs_df)
    by_week = {r["Week"]: dict(r) for r in results}
    rows = []
    for week in sorted(by_week.keys()):
        r = by_week[week]
        orders_from_detail = 0
        if not order_detail_df.empty and "Week" in order_detail_df.columns:
            subset = order_detail_df[order_detail_df["Week"] == week]
            orders_from_detail = int(round(subset["CaseEquiv"].sum())) if not subset.empty else 0
        prod = int(r.get("Production") or 0)
        orders = int(r.get("Shipped_Orders") or 0)
        nest = int(r.get("Shipped_NestRun") or 0)
        off = int(round(r.get("OffGrades_GradeOut") or 0))
        eggs_stored = int(r.get("Eggs_Stored") or 0)
        accounted = int(round(r.get("Accounted") or 0))
        pct = r.get("Traceability_Pct")
        rows.append({
            "Week": week,
            "Production": prod,
            "Shipped_NestRun": nest,
            "OffGrades_GradeOut": off,
            "Eggs_Stored": eggs_stored,
            "Shipped_Orders": orders,
            "Orders_From_Detail_Sum": orders_from_detail,
            "Orders_Match": "Yes" if abs(orders - orders_from_detail) <= 1 else "No",
            "Accounted": accounted,
            "Traceability_Pct": pct,
        })
    return pd.DataFrame(rows)


def _summarize_categories(series_like, weights=None):
    """Return sorted category summary like 'Manual NestRun=120 | Internal Use=8'."""
    counts = {}
    if weights is None:
        weights = [1] * len(series_like)
    for val, weight in zip(series_like, weights):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            continue
        s = str(val).strip()
        if not s:
            continue
        counts[s] = counts.get(s, 0) + int(round(_num(weight, default=1)))
    if not counts:
        return ""
    items = sorted(counts.items(), key=lambda x: (x[0].lower(), -x[1]))
    return " | ".join(f"{k}={v}" for k, v in items)


def load_sqf_inputs():
    """Load Reference_Data/SQF_Traceability_Inputs.csv."""
    if not INPUT_CSV.exists():
        return None
    try:
        return pd.read_csv(INPUT_CSV)
    except Exception as e:
        print(f"Error reading {INPUT_CSV}: {e}", file=sys.stderr)
        return None


def load_traceability_adjustments():
    """
    Load optional manual adjustments file.
    Columns: Week, Category, Cases, Evidence, Note
    """
    adj_path = ADJUSTMENTS_CSV_OVERRIDE if ADJUSTMENTS_CSV_OVERRIDE else ADJUSTMENTS_CSV
    adj_path = Path(adj_path) if not isinstance(adj_path, Path) else adj_path
    if not adj_path.exists():
        return pd.DataFrame(columns=["Week", "Category", "Cases", "Evidence", "Note"])
    try:
        df = pd.read_csv(adj_path)
    except Exception:
        return pd.DataFrame(columns=["Week", "Category", "Cases", "Evidence", "Note"])
    if df is None or df.empty:
        return pd.DataFrame(columns=["Week", "Category", "Cases", "Evidence", "Note"])
    for col in ("Week", "Category", "Cases", "Evidence", "Note"):
        if col not in df.columns:
            df[col] = ""
    return df


def _num(val, default=0):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _get_row(row, key, default=0):
    if hasattr(row, "get"):
        return row.get(key, default)
    if key in row.index:
        return row[key]
    return default


def _has_palletlines_file(week_number, exports_dir, slip_override=None, palletlines_source=PALLETLINES_SOURCE_PREFERRED):
    """True if a Week{N}_AllDays_PalletLines_*.xlsx file exists for this week (or SlipWeek)."""
    exports_dir = Path(exports_dir)
    slip_week = int(slip_override) if slip_override is not None and str(slip_override).strip() != "" and not (isinstance(slip_override, float) and pd.isna(slip_override)) else week_number
    return _get_latest_palletlines_file(slip_week, exports_dir, source_mode=palletlines_source) is not None


def _get_slip_override_from_inputs(inputs_df, week_number):
    """Return SlipWeek override from SQF_Traceability_Inputs.csv for a given report week."""
    if inputs_df is None or inputs_df.empty or "Week" not in inputs_df.columns or "SlipWeek" not in inputs_df.columns:
        return None
    match = inputs_df[inputs_df["Week"] == week_number]
    if match.empty:
        return None
    v = match.iloc[0].get("SlipWeek")
    if v is None or str(v).strip() == "" or (isinstance(v, float) and pd.isna(v)):
        return None
    return v


def _production_source(week_number, row):
    """Return ('CSV' | 'Hilly Acres' | 'Missing', production_value)."""
    prod_csv = _num(_get_row(row, "Production"), default=None)
    if prod_csv is not None and prod_csv > 0:
        return "CSV", prod_csv
    ha = get_production_from_hilly_acres(week_number)
    if ha is not None and ha > 0:
        return "Hilly Acres", ha
    return "Missing", (prod_csv if prod_csv is not None else 0)


def _gradeout_source(week_number, row, production):
    """Return string: where off-grades came from (or what's missing)."""
    off_csv = _num(_get_row(row, "OffGrades_GradeOut"), default=None)
    if off_csv is not None and off_csv > 0:
        return "CSV column"
    from_gradeout = get_off_grades_from_gradeout(week_number)
    if from_gradeout is not None and from_gradeout > 0:
        return "gradeout CSV"
    if production and production > 0:
        return f"{DEFAULT_OFFGRADE_PCT}% default (no gradeout data)"
    return "Missing (no production for gradeout)"


def run_inputs_report(exports_dir=None, by_slip_week=False):
    """
    For each week in SQF_Traceability_Inputs.csv: compute traceability % and list missing inputs.
    Returns list of dicts: Week, Traceability_Pct, Production_Source, PalletLines, Gradeout_Source, Missing_Inputs.
    """
    if exports_dir is None:
        exports_dir = EXPORTS_DIR
    exports_dir = Path(exports_dir)
    df = load_sqf_inputs()
    if df is None or df.empty or "Week" not in df.columns:
        return []
    rows_out = []
    for _, row in df.iterrows():
        w = _get_row(row, "Week")
        if w is None or (isinstance(w, float) and pd.isna(w)):
            continue
        week_str = str(int(w)) if isinstance(w, (int, float)) else str(w).strip()
        if not week_str.isdigit():
            continue
        week_num = int(week_str)
        slip_override = _get_row(row, "SlipWeek", None)
        if slip_override is not None and isinstance(slip_override, float) and pd.isna(slip_override):
            slip_override = None

        r = compute_week(week_num, row, exports_dir, by_slip_week=by_slip_week)
        prod_src, prod_val = _production_source(week_num, row)
        has_pl = _has_palletlines_file(week_num, exports_dir, slip_override)
        grade_src = _gradeout_source(week_num, row, prod_val)

        missing = []
        if prod_src == "Missing":
            missing.append("Hilly Acres / Production")
        if not has_pl:
            missing.append("Loading slip (PalletLines)")
        if prod_val and prod_val > 0:
            off_csv = _num(_get_row(row, "OffGrades_GradeOut"), default=None)
            from_go = get_off_grades_from_gradeout(week_num)
            if (off_csv is None or off_csv == 0) and (from_go is None or from_go == 0):
                missing.append("Gradeout data (using 1.2% default)")

        pct = r["Traceability_Pct"]
        pct_str = f"{pct}%" if pct is not None else "N/A (no Production)"
        rows_out.append({
            "Week": week_num,
            "Traceability_Pct": pct_str,
            "Production_Source": prod_src,
            "PalletLines": "Yes" if has_pl else "Missing",
            "Gradeout_Source": grade_src,
            "Missing_Inputs": "; ".join(missing) if missing else "None",
        })
    return rows_out


def export_inputs_report(report_rows, path=None):
    """Write inputs report to Excel (and CSV) in Traceability_Exports."""
    if path is None:
        path = EXPORTS_DIR / f"Traceability_Inputs_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(report_rows)
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="Inputs Report", index=False)
    csv_path = path.with_suffix(".csv")
    df.to_csv(csv_path, index=False)
    print(f"Inputs report: {path}")
    print(f"Also: {csv_path}")
    return path


def compute_week(week_number, row, exports_dir, by_slip_week=False, palletlines_source=PALLETLINES_SOURCE_PREFERRED):
    """Compute accounted, variance, and traceability % for one week."""
    # Support both dict and Series (row from DataFrame)
    def get(key, default=0):
        if hasattr(row, "get"):
            return row.get(key, default)
        if key in row.index:
            return row[key]
        return default
    prod_csv = _num(get("Production"), default=None)
    slip_override = get("SlipWeek", None)
    nest_run = get_nest_run_for_report_week(week_number, row, slip_week_override=slip_override)
    off_csv = _num(get("OffGrades_GradeOut"), default=None)
    if slip_override is not None and isinstance(slip_override, float) and pd.isna(slip_override):
        slip_override = None
    production = get_production_for_report_week(week_number, row, exports_dir, slip_week_override=slip_override)
    if production is None:
        production = prod_csv if prod_csv is not None else 0
    off_grades = get_off_grades_from_gradeout(week_number) if (off_csv is None or off_csv == 0) else off_csv
    if off_grades is None:
        off_grades = off_csv if off_csv is not None else 0
    # When we have no off-grade data for the week, use default 1.2% of Production
    if off_grades == 0 and production and production > 0:
        off_grades = round(production * DEFAULT_OFFGRADE_PCT / 100.0, 2)
    # Shipped_Orders_Override: use manual count from CSV when set (to match verified manual counts)
    orders_override = _num(get("Shipped_Orders_Override"), default=None)
    if orders_override is not None and orders_override > 0:
        orders = int(orders_override)
    else:
        orders = get_shipped_orders_boxes(week_number, exports_dir, by_slip_week=by_slip_week, slip_week_override=slip_override, palletlines_source=palletlines_source)
        # Fallback: if Option A gave 0 but we have a PalletLines file for this week, use slip-week sum
        # (slip "Week N" is often shipping week, not ISO week N, so BB dates map to other ISO weeks)
        if (orders is None or orders == 0) and not by_slip_week and _has_palletlines_file(week_number, exports_dir, slip_override, palletlines_source=palletlines_source):
            orders_slip = get_shipped_orders_boxes(week_number, exports_dir, by_slip_week=True, slip_week_override=slip_override, palletlines_source=palletlines_source)
            if orders_slip and orders_slip > 0:
                orders = orders_slip
    production_adjustment = 0
    try:
        import hilly_acres_production as ha
        production_adjustment = ha.get_production_adjustment_cases_for_week(week_number)
    except Exception:
        production_adjustment = 0
    production = int(round((production or 0) + production_adjustment))
    if orders is None:
        orders = 0
    eggs_stored = _num(get("Eggs_Stored"), default=0) or 0
    accounted = orders + nest_run + off_grades + eggs_stored
    variance = production - accounted if production else 0
    if production and production > 0:
        pct = round(100.0 * accounted / production, 2)
    else:
        pct = None
    out = {
        "Year": _year_for_report_week(week_number),
        "Week": week_number,
        "Production": production,
        "Production_Adjustment": production_adjustment,
        "Shipped_Orders": orders,
        "Shipped_NestRun": nest_run,
        "OffGrades_GradeOut": off_grades,
        "Eggs_Stored": eggs_stored,
        "Manual_Adjustments": 0,
        "Adjustment_Categories": "",
        "Accounted": accounted,
        "Variance": variance,
        "Traceability_Pct": pct,
        "Meets_Target": pct is not None and pct >= TARGET_PCT,
    }
    _apply_core_metrics(out)
    return out


def _apply_reallocate_to_prior_week(results, inputs_df):
    """
    For 2025 only: if CSV has Reallocate_To_Prior_Week > 0 for a week, subtract that from
    this week's Shipped_Orders and add to prior week's. Recompute Accounted, Variance,
    Traceability_Pct, Meets_Target for both weeks. Modifies results in place.
    """
    if inputs_df is None or inputs_df.empty or "Week" not in inputs_df.columns:
        return
    realloc_col = "Reallocate_To_Prior_Week"
    if realloc_col not in inputs_df.columns:
        return
    by_week = {r["Week"]: r for r in results}
    for _, row in inputs_df.iterrows():
        w = row.get("Week")
        if w is None or (isinstance(w, float) and pd.isna(w)):
            continue
        week_num = int(float(w))
        if week_num < 2 or week_num > 52:
            continue
        val = _num(row.get(realloc_col), default=None)
        if val is None or val <= 0:
            continue
        move = int(round(val))
        curr = by_week.get(week_num)
        prev = by_week.get(week_num - 1)
        if curr is None or prev is None:
            continue
        curr["Shipped_Orders"] = max(0, curr["Shipped_Orders"] - move)
        prev["Shipped_Orders"] = prev["Shipped_Orders"] + move
        for r in (curr, prev):
            prod = r.get("Production") or 0
            r["Accounted"] = (r["Shipped_Orders"] or 0) + (r["Shipped_NestRun"] or 0) + (r["OffGrades_GradeOut"] or 0) + (r.get("Eggs_Stored") or 0)
            r["Variance"] = prod - r["Accounted"] if prod else 0
            r["Traceability_Pct"] = round(100.0 * r["Accounted"] / prod, 2) if prod and prod > 0 else None
            r["Meets_Target"] = r["Traceability_Pct"] is not None and r["Traceability_Pct"] >= TARGET_PCT
            _apply_core_metrics(r)


def _apply_manual_adjustments(results, adjustments_df):
    """
    Apply optional evidence-backed adjustments from Traceability_Adjustments.csv.
    Each row adds Cases to Accounted for the specified Week and records the categories used.
    """
    if adjustments_df is None or adjustments_df.empty or "Week" not in adjustments_df.columns:
        return
    by_week = {r["Week"]: r for r in results}
    for r in results:
        r["Manual_Adjustments"] = r.get("Manual_Adjustments") or 0
        r["Adjustment_Categories"] = r.get("Adjustment_Categories") or ""
    for week, grp in adjustments_df.groupby("Week"):
        try:
            week_num = int(float(week))
        except (TypeError, ValueError):
            continue
        result = by_week.get(week_num)
        if result is None:
            continue
        cases_total = 0
        for _, row in grp.iterrows():
            cases = _num(row.get("Cases"), default=None)
            if cases is None or cases == 0:
                continue
            cases_total += int(round(cases))
        if cases_total == 0:
            continue
        result["Manual_Adjustments"] = (result.get("Manual_Adjustments") or 0) + cases_total
        result["Adjustment_Categories"] = _summarize_categories(grp["Category"], grp["Cases"])
        prod = result.get("Production") or 0
        result["Accounted"] = (result.get("Shipped_Orders") or 0) + (result.get("Shipped_NestRun") or 0) + (result.get("OffGrades_GradeOut") or 0) + (result.get("Eggs_Stored") or 0) + (result.get("Manual_Adjustments") or 0)
        result["Variance"] = prod - result["Accounted"] if prod else 0
        result["Traceability_Pct"] = round(100.0 * result["Accounted"] / prod, 2) if prod and prod > 0 else None
        result["Meets_Target"] = result["Traceability_Pct"] is not None and result["Traceability_Pct"] >= TARGET_PCT
        _apply_core_metrics(result)


def load_weekly_production_overrides():
    """
    Optional weekly production overrides.
    CSV columns: Week, Delta_Cases [, Evidence, Note]
    Positive Delta_Cases increases production; negative decreases production.
    """
    cols = ["Week", "Delta_Cases", "Evidence", "Note"]
    if not PRODUCTION_WEEKLY_OVERRIDES_CSV.exists():
        return pd.DataFrame(columns=cols)
    try:
        df = pd.read_csv(PRODUCTION_WEEKLY_OVERRIDES_CSV)
    except Exception:
        return pd.DataFrame(columns=cols)
    if df is None or df.empty:
        return pd.DataFrame(columns=cols)
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    rows = []
    for _, row in df.iterrows():
        try:
            week = int(float(row.get("Week")))
            delta = float(row.get("Delta_Cases"))
        except (TypeError, ValueError):
            continue
        if week <= 0:
            continue
        rows.append({
            "Week": week,
            "Delta_Cases": delta,
            "Evidence": str(row.get("Evidence") or "").strip(),
            "Note": str(row.get("Note") or "").strip(),
        })
    return pd.DataFrame(rows, columns=cols)


def _apply_weekly_production_overrides(results, overrides_df):
    """Apply weekly production deltas in-place and recompute dependent metrics."""
    if not results:
        return
    overrides_df = overrides_df if overrides_df is not None else load_weekly_production_overrides()
    if overrides_df is None or overrides_df.empty:
        return
    week_delta = {}
    for _, row in overrides_df.iterrows():
        try:
            week = int(float(row.get("Week")))
            delta = float(row.get("Delta_Cases"))
        except (TypeError, ValueError):
            continue
        week_delta[week] = week_delta.get(week, 0.0) + delta
    for r in results:
        try:
            week = int(r.get("Week") or 0)
        except Exception:
            continue
        if week <= 0 or week not in week_delta:
            continue
        prod_before = float(r.get("Production") or 0)
        prod_after = max(0.0, prod_before + float(week_delta[week]))
        r["Production"] = int(round(prod_after))
        r["Production_Weekly_Override"] = int(round(week_delta[week]))
        accounted = (r.get("Shipped_Orders") or 0) + (r.get("Shipped_NestRun") or 0) + (r.get("OffGrades_GradeOut") or 0) + (r.get("Eggs_Stored") or 0) + (r.get("Manual_Adjustments") or 0)
        r["Accounted"] = accounted
        r["Variance"] = r["Production"] - accounted if r["Production"] else 0
        r["Traceability_Pct"] = round(100.0 * accounted / r["Production"], 2) if r["Production"] > 0 else None
        r["Meets_Target"] = r["Traceability_Pct"] is not None and r["Traceability_Pct"] >= TARGET_PCT
        _apply_core_metrics(r)


def _failsafe_to_100_enabled():
    """Background / CLI switch: EGGROOM_FAILSAFE_TO_100=1|true|yes|on."""
    return (os.environ.get("EGGROOM_FAILSAFE_TO_100") or "").strip().lower() in ("1", "true", "yes", "on")


def _resolve_evidence_file(path_str):
    """Resolve optional evidence file path (absolute or relative to Reference_Data / project / data base)."""
    s = (path_str or "").strip()
    if not s:
        return None
    p = Path(s)
    if p.is_file():
        return p.resolve()
    for base in (REF_DATA_DIR, BASE_DIR, DATA_BASE_DIR):
        cand = (base / s).resolve()
        if cand.is_file():
            return cand
    return None


def load_failsafe_evidence():
    """
    Rows authorize failsafe top-up to 100% for a week.
    Columns: Week, Evidence (required), Evidence_File (optional; must exist if set), Case_Cap (optional), Note (optional)
    """
    cols = ["Week", "Evidence", "Evidence_File", "Case_Cap", "Note"]
    if not FAILSAFE_EVIDENCE_CSV.exists():
        return pd.DataFrame(columns=cols)
    try:
        df = pd.read_csv(FAILSAFE_EVIDENCE_CSV)
    except Exception:
        return pd.DataFrame(columns=cols)
    if df is None or df.empty:
        return pd.DataFrame(columns=cols)
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    return df


def _apply_failsafe_topup_to_100(results):
    """
    For weeks below 100% traceability with Production > 0, add accounted top-up equal to remaining gap
    when a matching evidence row exists. Skips weeks already at/above 100%.
    Does not run during pallet balancing (caller must not invoke when apply_manual_adjustments=False).
    """
    ev_df = load_failsafe_evidence()
    if ev_df is None or ev_df.empty:
        return
    by_week = {int(r["Week"]): r for r in results if r.get("Week") is not None}
    for _, row in ev_df.iterrows():
        try:
            week_num = int(float(row.get("Week")))
        except (TypeError, ValueError):
            continue
        evidence = str(row.get("Evidence") or "").strip()
        if not evidence:
            continue
        ev_file = str(row.get("Evidence_File") or "").strip()
        if ev_file:
            resolved = _resolve_evidence_file(ev_file)
            if resolved is None:
                print(
                    f"Failsafe skip week {week_num}: Evidence_File not found ({ev_file}).",
                    file=sys.stderr,
                )
                continue
        cap = _num(row.get("Case_Cap"), default=None)
        note = str(row.get("Note") or "").strip()
        result = by_week.get(week_num)
        if result is None:
            continue
        prod = int(result.get("Production") or 0)
        if prod <= 0:
            continue
        accounted = float(result.get("Accounted") or 0)
        gap = float(prod) - accounted
        if gap <= 0:
            continue
        add_amt = gap
        if cap is not None and cap > 0:
            add_amt = min(add_amt, float(cap))
        add_amt = int(round(add_amt))
        if add_amt <= 0:
            continue
        result["Manual_Adjustments"] = (result.get("Manual_Adjustments") or 0) + add_amt
        prev_fs = int(result.get("Failsafe_Topup_Cases") or 0)
        result["Failsafe_Topup_Cases"] = prev_fs + add_amt
        bits = [evidence]
        if ev_file:
            bits.append(f"file={ev_file}")
        if note:
            bits.append(note)
        result["Failsafe_Evidence_Summary"] = " | ".join(
            x for x in [result.get("Failsafe_Evidence_Summary"), "; ".join(bits)] if x
        )
        cats = (result.get("Adjustment_Categories") or "").strip()
        tail = f"Failsafe_To100={add_amt}"
        result["Adjustment_Categories"] = f"{cats} | {tail}" if cats else tail

        result["Accounted"] = (
            (result.get("Shipped_Orders") or 0)
            + (result.get("Shipped_NestRun") or 0)
            + (result.get("OffGrades_GradeOut") or 0)
            + (result.get("Eggs_Stored") or 0)
            + (result.get("Manual_Adjustments") or 0)
        )
        result["Variance"] = prod - result["Accounted"] if prod else 0
        result["Traceability_Pct"] = (
            round(100.0 * result["Accounted"] / prod, 2) if prod > 0 else None
        )
        result["Meets_Target"] = result["Traceability_Pct"] is not None and result["Traceability_Pct"] >= TARGET_PCT
        _apply_core_metrics(result)


def build_gap_analysis(results, exports_dir=None, inputs_df=None, adjustments_df=None, palletlines_source=PALLETLINES_SOURCE_PREFERRED):
    """
    Build week-by-week gap analysis showing what explains the difference to 99%.
    """
    if not results:
        return pd.DataFrame()
    exports_dir = Path(exports_dir) if exports_dir else EXPORTS_DIR
    inputs_df = inputs_df if inputs_df is not None else load_sqf_inputs()
    adjustments_df = adjustments_df if adjustments_df is not None else load_traceability_adjustments()
    rows = []
    old_new_df = build_all_2025_old_new_by_day(exports_dir, inputs_df)
    olddate_by_week = {}
    if old_new_df is not None and not old_new_df.empty:
        olddate_by_week = old_new_df.groupby("Week")["Cases_OldDate"].sum().to_dict()
    adj_by_week = {}
    adj_cats_by_week = {}
    adj_evidence_by_week = {}
    if adjustments_df is not None and not adjustments_df.empty:
        for week, grp in adjustments_df.groupby("Week"):
            try:
                week_num = int(float(week))
            except (TypeError, ValueError):
                continue
            adj_by_week[week_num] = int(round(sum(_num(v, default=0) for v in grp["Cases"])))
            adj_cats_by_week[week_num] = _summarize_categories(grp["Category"], grp["Cases"])
            evidence_bits = []
            for _, row in grp.iterrows():
                ev = str(row.get("Evidence") or "").strip()
                note = str(row.get("Note") or "").strip()
                cat = str(row.get("Category") or "").strip()
                if ev or note or cat:
                    text = " - ".join(x for x in (cat, ev, note) if x)
                    if text:
                        evidence_bits.append(text)
            adj_evidence_by_week[week_num] = " | ".join(evidence_bits[:5])
    for r in results:
        week = r["Week"]
        slip_override = _get_slip_override_from_inputs(inputs_df, week)
        option_a_orders = get_shipped_orders_boxes(week, exports_dir, by_slip_week=False, slip_week_override=slip_override, palletlines_source=palletlines_source) or 0
        raw_slip_total = get_shipped_orders_boxes(week, exports_dir, by_slip_week=True, slip_week_override=slip_override, palletlines_source=palletlines_source) or 0
        olddate_move = int(olddate_by_week.get(week, 0))
        manual_adj = int(adj_by_week.get(week, r.get("Manual_Adjustments") or 0))
        production = int(r.get("Production") or 0)
        accounted = int(round(r.get("Accounted") or 0))
        gap_to_99 = 0
        if production > 0:
            needed = int((production * TARGET_PCT + 99) // 100) if TARGET_PCT.is_integer() else int(round((TARGET_PCT / 100.0) * production))
            gap_to_99 = max(0, needed - accounted)
        orders_gap = raw_slip_total - option_a_orders
        likely = "At or above target"
        if production <= 0:
            likely = "Missing production"
        elif gap_to_99 > 0:
            if orders_gap >= max(100, int(gap_to_99 * 0.6)):
                likely = "Orders exist on slips but are landing in other weeks"
            elif olddate_move >= max(50, int(gap_to_99 * 0.5)):
                likely = "Old-date movement can justify shifting cases to prior week"
            elif (r.get("Shipped_NestRun") or 0) < max(200, int(gap_to_99 * 0.2)):
                likely = "Likely missing nest run or other non-order movement"
            elif (r.get("OffGrades_GradeOut") or 0) <= round(production * DEFAULT_OFFGRADE_PCT / 100.0, 2):
                likely = "Likely missing gradeout, shrink, or internal-use bucket"
            else:
                likely = "Production may be overstated or another accounted bucket is missing"
        rows.append({
            "Year": r.get("Year"),
            "Week": week,
            "Production": production,
            "Orders_Reported": int(r.get("Shipped_Orders") or 0),
            "Orders_OptionA_Base": int(option_a_orders),
            "Orders_RawSlipTotal": int(raw_slip_total),
            "Orders_Not_In_OptionA": int(orders_gap),
            "OldDate_Move_Available": olddate_move,
            "NestRun": int(r.get("Shipped_NestRun") or 0),
            "OffGrades": int(round(r.get("OffGrades_GradeOut") or 0)),
            "Manual_Adjustments": manual_adj,
            "Adjustment_Categories": adj_cats_by_week.get(week, r.get("Adjustment_Categories") or ""),
            "Accounted_Reported": accounted,
            "Gap_To_99": int(gap_to_99),
            "Likely_Missing_Source": likely,
            "Adjustment_Evidence": adj_evidence_by_week.get(week, ""),
        })
    return pd.DataFrame(rows)


def _target_cases_for_pct(production, pct=TARGET_PCT):
    """Smallest accounted integer that meets target pct for a production total."""
    if not production or production <= 0:
        return 0
    raw = production * pct / 100.0
    return int(raw) if raw == int(raw) else int(raw) + 1


def build_suggested_reallocations(results, gap_df=None):
    """
    Suggest 2025 week-to-prior-week reallocations using only weeks that are above 99%
    and have provable old-date volume available.
    """
    if not results:
        return pd.DataFrame(columns=[
            "From_Week", "To_Week", "Suggested_Reallocate_To_Prior_Week",
            "From_Week_Pct_Before", "To_Week_Pct_Before", "Reason"
        ])
    gap_df = gap_df if gap_df is not None else pd.DataFrame()
    gap_old = {}
    if gap_df is not None and not gap_df.empty and "Week" in gap_df.columns and "OldDate_Move_Available" in gap_df.columns:
        gap_old = dict(zip(gap_df["Week"], gap_df["OldDate_Move_Available"]))
    by_week = {r["Week"]: dict(r) for r in results if r.get("Year") == 2025 and (r.get("Production") or 0) > 0}
    rows = []
    for week in range(2, 53):
        curr = by_week.get(week)
        prev = by_week.get(week - 1)
        if curr is None or prev is None:
            continue
        curr_target = _target_cases_for_pct(curr["Production"])
        prev_target = _target_cases_for_pct(prev["Production"])
        curr_excess = max(0, int((curr["Accounted"] or 0) - curr_target))
        prev_gap = max(0, int(prev_target - (prev["Accounted"] or 0)))
        olddate_available = int(gap_old.get(week, 0) or 0)
        move = min(curr_excess, prev_gap, olddate_available)
        if move <= 0:
            continue
        rows.append({
            "From_Week": week,
            "To_Week": week - 1,
            "Suggested_Reallocate_To_Prior_Week": move,
            "From_Week_Pct_Before": curr.get("Traceability_Pct"),
            "To_Week_Pct_Before": prev.get("Traceability_Pct"),
            "Reason": f"Week {week} has {olddate_available} provable old-date cases available; moving {move} gets both weeks closer to {TARGET_PCT}%.",
        })
    return pd.DataFrame(rows)


def build_suggested_adjustments(results, gap_df=None, realloc_df=None):
    """
    Suggest first-pass evidence-backed adjustments after applying suggested reallocations virtually.
    Prioritizes weeks below 90%, but also includes small below-target weeks.
    """
    if not results:
        return pd.DataFrame(columns=[
            "Week", "Priority", "Suggested_Category", "Suggested_Cases",
            "Traceability_Pct_After_Reallocation", "Likely_Missing_Source", "Evidence", "Suggested_Note"
        ])
    gap_df = gap_df if gap_df is not None else pd.DataFrame()
    realloc_df = realloc_df if realloc_df is not None else pd.DataFrame()
    merged = pd.DataFrame(results)
    if merged.empty:
        return pd.DataFrame()
    if gap_df is not None and not gap_df.empty:
        keep_cols = [c for c in ("Week", "Orders_Not_In_OptionA", "OldDate_Move_Available", "Likely_Missing_Source", "Gap_To_99") if c in gap_df.columns]
        merged = merged.merge(gap_df[keep_cols], on="Week", how="left")
    by_week = {int(r["Week"]): dict(r) for _, r in merged.iterrows() if r.get("Year") == 2025 and (r.get("Production") or 0) > 0}
    if realloc_df is not None and not realloc_df.empty:
        for _, row in realloc_df.iterrows():
            src = int(row["From_Week"])
            dst = int(row["To_Week"])
            move = int(row["Suggested_Reallocate_To_Prior_Week"])
            if src in by_week:
                by_week[src]["Shipped_Orders"] = (by_week[src].get("Shipped_Orders") or 0) - move
                by_week[src]["Accounted"] = (by_week[src].get("Accounted") or 0) - move
            if dst in by_week:
                by_week[dst]["Shipped_Orders"] = (by_week[dst].get("Shipped_Orders") or 0) + move
                by_week[dst]["Accounted"] = (by_week[dst].get("Accounted") or 0) + move
    rows = []
    for week in sorted(by_week):
        r = by_week[week]
        production = int(r.get("Production") or 0)
        if production <= 0:
            continue
        accounted = float(r.get("Accounted") or 0)
        pct = round(100.0 * accounted / production, 2)
        target = _target_cases_for_pct(production)
        gap = max(0, target - int(round(accounted)))
        if gap <= 0:
            continue
        likely = str(r.get("Likely_Missing_Source") or "").strip()
        if "orders exist on slips" in likely.lower():
            category = "Missing Orders"
        elif "nest run" in likely.lower():
            category = "Manual NestRun"
        elif "gradeout" in likely.lower() or "internal-use" in likely.lower() or "shrink" in likely.lower():
            category = "Shrink / Internal Use / Writeoff"
        elif "production" in likely.lower():
            category = "Production Review Needed"
        else:
            category = "Traceability Review Adjustment"
        priority = "High" if pct < 90 else "Medium"
        evidence_parts = []
        if pd.notna(r.get("Orders_Not_In_OptionA")):
            evidence_parts.append(f"Orders_Not_In_OptionA={int(r.get('Orders_Not_In_OptionA') or 0)}")
        if pd.notna(r.get("OldDate_Move_Available")):
            evidence_parts.append(f"OldDate_Move_Available={int(r.get('OldDate_Move_Available') or 0)}")
        note = f"Add {gap} cases to reach about {TARGET_PCT}% after suggested reallocation."
        rows.append({
            "Week": week,
            "Priority": priority,
            "Suggested_Category": category,
            "Suggested_Cases": int(gap),
            "Traceability_Pct_After_Reallocation": pct,
            "Likely_Missing_Source": likely,
            "Evidence": " | ".join(evidence_parts),
            "Suggested_Note": note,
        })
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    return out.sort_values(["Priority", "Week"], ascending=[True, True]).reset_index(drop=True)


def build_suggested_eggs_stored(results, gap_df=None):
    """
    Suggest Eggs_Stored for weeks leading up to flock-removal / zero-barn weeks.

    This is intentionally conservative:
    - only weeks in a zero-barn transition window are considered
    - the report shows both the remaining gap to 99% and the full remaining variance
    - users should only copy values into SQF_Traceability_Inputs.csv when storage/rotation
      actually happened and the quantity is supportable
    """
    cols = [
        "Year", "Week", "Transition_Stage", "Zero_Barn_Week", "Zero_Barns",
        "Traceability_Pct", "Production", "Accounted", "Current_Eggs_Stored",
        "Remaining_Gap_To_99", "Remaining_Unaccounted_To_100",
        "Likely_Missing_Source", "Suggested_Note"
    ]
    if not results:
        return pd.DataFrame(columns=cols)
    try:
        import hilly_acres_production as ha
    except ImportError:
        return pd.DataFrame(columns=cols)

    gap_df = gap_df if gap_df is not None else pd.DataFrame()
    likely_by_week = {}
    gap99_by_week = {}
    if gap_df is not None and not gap_df.empty and "Week" in gap_df.columns:
        if "Likely_Missing_Source" in gap_df.columns:
            likely_by_week = dict(zip(gap_df["Week"], gap_df["Likely_Missing_Source"]))
        if "Gap_To_99" in gap_df.columns:
            gap99_by_week = dict(zip(gap_df["Week"], gap_df["Gap_To_99"]))

    zero_barns_by_week = {}
    for r in results:
        week = int(r.get("Week") or 0)
        if week <= 0:
            continue
        barn_stacks = ha.get_barn_stacks_for_week(week)
        if not barn_stacks:
            continue
        zero_barns = sorted(
            int(b) for b, stacks in barn_stacks.items()
            if _num(stacks, default=0) <= 0
        )
        if zero_barns:
            zero_barns_by_week[week] = zero_barns

    if not zero_barns_by_week:
        return pd.DataFrame(columns=cols)

    transition_meta = {}
    for zero_week, zero_barns in zero_barns_by_week.items():
        for prior_weeks, stage in (
            (4, "4 weeks before removal"),
            (3, "3 weeks before removal"),
            (2, "2 weeks before removal"),
            (1, "1 week before removal"),
            (0, "Removal week"),
        ):
            week = zero_week - prior_weeks
            if week < 1:
                continue
            meta = transition_meta.setdefault(
                week,
                {"Zero_Barn_Weeks": set(), "Zero_Barns": set(), "Stages": []},
            )
            meta["Zero_Barn_Weeks"].add(zero_week)
            meta["Zero_Barns"].update(zero_barns)
            meta["Stages"].append((prior_weeks, stage))

    rows = []
    by_week = {int(r["Week"]): r for r in results}
    for week in sorted(transition_meta):
        r = by_week.get(week)
        if r is None:
            continue
        production = int(r.get("Production") or 0)
        accounted = int(round(r.get("Accounted") or 0))
        current_eggs_stored = int(r.get("Eggs_Stored") or 0)
        variance = max(0, production - accounted)
        gap_to_99 = int(gap99_by_week.get(week, 0) or 0)
        remaining_gap_to_99 = max(0, gap_to_99)
        remaining_unaccounted = max(0, variance)
        if remaining_gap_to_99 <= 0 and remaining_unaccounted <= 0 and current_eggs_stored <= 0:
            continue
        stages = sorted(transition_meta[week]["Stages"], key=lambda x: x[0], reverse=True)
        stage = stages[0][1]
        zero_barn_weeks = sorted(transition_meta[week]["Zero_Barn_Weeks"])
        zero_barns = sorted(transition_meta[week]["Zero_Barns"])
        likely = str(likely_by_week.get(week, "") or "").strip()
        zero_week_text = ", ".join(str(w) for w in zero_barn_weeks)
        zero_barn_text = ", ".join(f"Barn {b}" for b in zero_barns)
        note = (
            f"Week {week} is in the lead-up to zero-barn week(s) {zero_week_text} "
            f"({zero_barn_text}). If eggs were stored/rotated before flock removal, "
            f"use up to {remaining_unaccounted:,} cases for Eggs_Stored; "
            f"{remaining_gap_to_99:,} is enough to reach about {TARGET_PCT}%."
        )
        rows.append({
            "Year": r.get("Year"),
            "Week": week,
            "Transition_Stage": stage,
            "Zero_Barn_Week": zero_week_text,
            "Zero_Barns": zero_barn_text,
            "Traceability_Pct": r.get("Traceability_Pct"),
            "Production": production,
            "Accounted": accounted,
            "Current_Eggs_Stored": current_eggs_stored,
            "Remaining_Gap_To_99": remaining_gap_to_99,
            "Remaining_Unaccounted_To_100": remaining_unaccounted,
            "Likely_Missing_Source": likely,
            "Suggested_Note": note,
        })
    out = pd.DataFrame(rows, columns=cols)
    if out.empty:
        return out
    return out.sort_values(["Year", "Week"]).reset_index(drop=True)


def _display_path(path_obj):
    """Return readable path string relative to BASE_DIR when possible."""
    if not path_obj:
        return ""
    try:
        p = Path(path_obj)
        return str(p.relative_to(BASE_DIR)) if BASE_DIR in p.parents else str(p)
    except Exception:
        return str(path_obj)


def _ha_file_details_from_path(path_obj):
    """Read a Hilly Acres workbook's Inputs week and Week Ending."""
    if not path_obj:
        return {"path": "", "inputs_week": None, "week_ending": None}
    try:
        import hilly_acres_production as ha
        df = pd.read_excel(path_obj, sheet_name=ha.INPUTS_SHEET, header=None)
        return {
            "path": _display_path(path_obj),
            "inputs_week": ha._read_inputs_week_number(path_obj),
            "week_ending": ha._get_week_ending_from_inputs(df),
        }
    except Exception:
        return {"path": _display_path(path_obj), "inputs_week": None, "week_ending": None}


def _ha_file_details_from_report_week(report_week):
    """Locate the Hilly Acres workbook that week-number matching would use for this report week."""
    try:
        import hilly_acres_production as ha
    except ImportError:
        return {"path": "", "inputs_week": None, "week_ending": None, "lookup_week": None}
    for folder in ha._get_hilly_acres_paths():
        try:
            lookup_week = ha._get_week_to_look_for_in_folder(folder, report_week)
            if lookup_week is None:
                continue
            path = ha._find_file_for_week(folder, lookup_week)
            if path is not None:
                out = _ha_file_details_from_path(path)
                out["lookup_week"] = lookup_week
                return out
        except Exception:
            continue
    return {"path": "", "inputs_week": None, "week_ending": None, "lookup_week": None}


def _ha_file_details_from_override(report_week, explicit_week):
    """Locate the Hilly Acres workbook used by HillyAcresWeek override in the correct report-year context."""
    if explicit_week is None:
        return {"path": "", "inputs_week": None, "week_ending": None, "lookup_week": None}
    try:
        import hilly_acres_production as ha
    except ImportError:
        return {"path": "", "inputs_week": None, "week_ending": None, "lookup_week": None}
    for folder in _get_hilly_acres_folders_for_report_week(report_week):
        try:
            path = ha._find_file_for_week(folder, explicit_week)
            if path is not None:
                out = _ha_file_details_from_path(path)
                out["lookup_week"] = explicit_week
                return out
        except Exception:
            continue
    return {"path": "", "inputs_week": None, "week_ending": None, "lookup_week": None}


def _ha_file_details_from_week_ending(week_ending):
    """Locate the Hilly Acres workbook matched by Week Ending date."""
    if week_ending is None:
        return {"path": "", "inputs_week": None, "week_ending": None}
    try:
        import hilly_acres_production as ha
    except ImportError:
        return {"path": "", "inputs_week": None, "week_ending": None}
    for folder in ha._get_hilly_acres_paths():
        try:
            for f in sorted(folder.glob("*.xlsx")):
                if "copy of" in f.name.lower():
                    continue
                try:
                    df = pd.read_excel(f, sheet_name=ha.INPUTS_SHEET, header=None)
                except Exception:
                    continue
                we = ha._get_week_ending_from_inputs(df)
                if we == week_ending:
                    return {
                        "path": _display_path(f),
                        "inputs_week": ha._read_inputs_week_number(f),
                        "week_ending": we,
                    }
        except Exception:
            continue
    return {"path": "", "inputs_week": None, "week_ending": None}


def _production_match_details(report_week, row, exports_dir=None):
    """
    Explain which production source/file current logic would use for a report week.
    Returns selected source plus week-ending and week-number matches for audit review.
    """
    exports_dir = Path(exports_dir) if exports_dir else EXPORTS_DIR
    slip_override = _get_row(row, "SlipWeek", None)
    if slip_override is not None and isinstance(slip_override, float) and pd.isna(slip_override):
        slip_override = None
    slip_week = int(float(slip_override)) if slip_override is not None and str(slip_override).strip() else report_week
    ha_override = _get_hilly_acres_week_override(row)

    slip_file = ""
    slip_week_ending = None
    try:
        from process_weekly_loading_slip import _find_slip_file_for_week, get_slip_week_ending
        slip_path = _find_slip_file_for_week(slip_week)
        slip_file = _display_path(slip_path) if slip_path else ""
        slip_week_ending = get_slip_week_ending(slip_week)
    except Exception:
        pass

    week_ending_match = _ha_file_details_from_week_ending(slip_week_ending)
    report_week_match = _ha_file_details_from_report_week(report_week)
    override_match = _ha_file_details_from_override(report_week, ha_override) if ha_override is not None else {
        "path": "", "inputs_week": None, "week_ending": None, "lookup_week": None
    }

    dominant_prod_key = get_dominant_production_week_from_slip(slip_week, exports_dir)
    dominant_match = {"path": "", "inputs_week": None, "week_ending": None, "lookup_week": None}
    if dominant_prod_key is not None:
        _, dominant_iso_week = dominant_prod_key
        dominant_match = _ha_file_details_from_report_week(dominant_iso_week)

    prod_csv = _num(_get_row(row, "Production"), default=None)
    selected_source = "Missing"
    selected_file = ""
    selected_lookup = None
    if prod_csv is not None and prod_csv > 0:
        selected_source = "CSV Production"
    elif ha_override is not None and override_match.get("path"):
        selected_source = "HillyAcresWeek override"
        selected_file = override_match.get("path") or ""
        selected_lookup = ha_override
    elif week_ending_match.get("path"):
        selected_source = "WeekEnding match"
        selected_file = week_ending_match.get("path") or ""
    elif report_week_match.get("path"):
        selected_source = "ReportWeek match"
        selected_file = report_week_match.get("path") or ""
        selected_lookup = report_week_match.get("lookup_week")
    elif dominant_match.get("path"):
        selected_source = "Dominant BB week"
        selected_file = dominant_match.get("path") or ""
        selected_lookup = dominant_match.get("lookup_week")

    selected_details = {"path": "", "inputs_week": None, "week_ending": None}
    if selected_source == "HillyAcresWeek override":
        selected_details = override_match
    elif selected_source == "WeekEnding match":
        selected_details = week_ending_match
    elif selected_source == "ReportWeek match":
        selected_details = report_week_match
    elif selected_source == "Dominant BB week":
        selected_details = dominant_match

    aligned = ""
    if slip_week_ending is not None and selected_details.get("week_ending") is not None:
        aligned = "Yes" if slip_week_ending == selected_details.get("week_ending") else "No"

    review_note = ""
    if not slip_file:
        review_note = "Missing loading slip file"
    elif not selected_file and selected_source != "CSV Production":
        review_note = "No Hilly Acres workbook matched by current logic"
    elif selected_source == "HillyAcresWeek override":
        review_note = "Production is using manual HillyAcresWeek override"
    elif selected_source == "ReportWeek match" and week_ending_match.get("path"):
        review_note = "WeekEnding workbook exists but current logic fell back to week-number match"
    elif aligned == "No":
        review_note = "Slip week-ending and selected Hilly Acres week-ending do not match"
    elif selected_source == "WeekEnding match":
        review_note = "Date alignment looks good"

    return {
        "Slip_Week_Used": slip_week,
        "Slip_File": slip_file,
        "Slip_Week_Ending": slip_week_ending,
        "HillyAcresWeek_Override": ha_override,
        "HA_File_By_WeekEnding": week_ending_match.get("path") or "",
        "HA_WeekEnding_By_WeekEnding": week_ending_match.get("week_ending"),
        "HA_InputsWeek_By_WeekEnding": week_ending_match.get("inputs_week"),
        "HA_File_By_ReportWeek": report_week_match.get("path") or "",
        "HA_LookupWeek_By_ReportWeek": report_week_match.get("lookup_week"),
        "HA_WeekEnding_By_ReportWeek": report_week_match.get("week_ending"),
        "HA_InputsWeek_By_ReportWeek": report_week_match.get("inputs_week"),
        "HA_File_By_Override": override_match.get("path") or "",
        "HA_WeekEnding_By_Override": override_match.get("week_ending"),
        "HA_InputsWeek_By_Override": override_match.get("inputs_week"),
        "Dominant_BB_ISO_Week": dominant_prod_key[1] if dominant_prod_key else None,
        "HA_File_By_Dominant_BB_Week": dominant_match.get("path") or "",
        "Production_Source_Selected": selected_source,
        "HA_File_Selected": selected_file,
        "HA_Selected_LookupWeek": selected_lookup,
        "HA_Selected_InputsWeek": selected_details.get("inputs_week"),
        "HA_Selected_WeekEnding": selected_details.get("week_ending"),
        "WeekEnding_Aligned": aligned,
        "Review_Note": review_note,
    }


def build_week_crosswalk(results, exports_dir=None, inputs_df=None):
    """
    Build audit-style crosswalk showing which slip/Hilly Acres files and dates are used.
    Focuses on 2025 week 15 onward plus 2026 report weeks.
    """
    cols = [
        "Year", "Week", "Traceability_Pct", "Production_Used", "Shipped_Orders", "Shipped_NestRun",
        "OffGrades_GradeOut", "Eggs_Stored", "Accounted", "Variance",
        "Slip_Week_Used", "Slip_File", "Slip_Week_Ending",
        "HillyAcresWeek_Override",
        "HA_File_By_WeekEnding", "HA_WeekEnding_By_WeekEnding", "HA_InputsWeek_By_WeekEnding",
        "HA_File_By_ReportWeek", "HA_LookupWeek_By_ReportWeek", "HA_WeekEnding_By_ReportWeek", "HA_InputsWeek_By_ReportWeek",
        "HA_File_By_Override", "HA_WeekEnding_By_Override", "HA_InputsWeek_By_Override",
        "Dominant_BB_ISO_Week", "HA_File_By_Dominant_BB_Week",
        "Production_Source_Selected", "HA_File_Selected", "HA_Selected_LookupWeek",
        "HA_Selected_InputsWeek", "HA_Selected_WeekEnding", "WeekEnding_Aligned", "Review_Note"
    ]
    if not results:
        return pd.DataFrame(columns=cols)
    exports_dir = Path(exports_dir) if exports_dir else EXPORTS_DIR
    inputs_df = inputs_df if inputs_df is not None else load_sqf_inputs()
    input_by_week = {}
    if inputs_df is not None and not inputs_df.empty and "Week" in inputs_df.columns:
        for _, row in inputs_df.iterrows():
            try:
                input_by_week[int(float(row.get("Week")))] = row
            except Exception:
                continue
    rows = []
    for r in results:
        week = int(r.get("Week") or 0)
        year = int(r.get("Year") or 0)
        if year == 2025 and week < 15:
            continue
        row = input_by_week.get(week, {})
        match = _production_match_details(week, row, exports_dir)
        out = {
            "Year": year,
            "Week": week,
            "Traceability_Pct": r.get("Traceability_Pct"),
            "Production_Used": int(r.get("Production") or 0),
            "Shipped_Orders": int(r.get("Shipped_Orders") or 0),
            "Shipped_NestRun": int(r.get("Shipped_NestRun") or 0),
            "OffGrades_GradeOut": int(round(r.get("OffGrades_GradeOut") or 0)),
            "Eggs_Stored": int(r.get("Eggs_Stored") or 0),
            "Accounted": int(round(r.get("Accounted") or 0)),
            "Variance": int(round(r.get("Variance") or 0)),
        }
        out.update(match)
        rows.append(out)
    return pd.DataFrame(rows, columns=cols)


def export_week_crosswalk(path=None, exports_dir=None):
    """Export the week crosswalk as Excel + CSV for audit review."""
    exports_dir = Path(exports_dir) if exports_dir else EXPORTS_DIR
    exports_dir.mkdir(parents=True, exist_ok=True)
    if path is None:
        path = exports_dir / f"Week_Crosswalk_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = Path(path)
    results = run_all_weeks(exports_dir)
    inputs_df = load_sqf_inputs()
    crosswalk_df = build_week_crosswalk(results, exports_dir, inputs_df)
    if crosswalk_df.empty:
        print("No crosswalk rows to export.")
        return None
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        crosswalk_df.to_excel(w, sheet_name="Week_Crosswalk", index=False)
    csv_path = path.with_suffix(".csv")
    crosswalk_df.to_csv(csv_path, index=False)
    print(f"Week crosswalk: {path}")
    print(f"Also: {csv_path}")
    return path


def run_all_weeks(
    exports_dir=None,
    by_slip_week=False,
    palletlines_source=PALLETLINES_SOURCE_PREFERRED,
    apply_manual_adjustments=True,
):
    """Compute traceability for all weeks in SQF_Traceability_Inputs.csv. Applies Reallocate_To_Prior_Week for 2025 (2-52) when set."""
    if exports_dir is None:
        exports_dir = EXPORTS_DIR
    exports_dir = Path(exports_dir)
    df = load_sqf_inputs()
    if df is None or df.empty or "Week" not in df.columns:
        return None
    results = []
    for _, row in df.iterrows():
        w = row.get("Week")
        if w is None or (isinstance(w, float) and pd.isna(w)):
            continue
        week_str = str(int(w)) if isinstance(w, (int, float)) else str(w).strip()
        if not week_str.isdigit():
            continue
        week_num = int(week_str)
        r = compute_week(week_num, row, exports_dir, by_slip_week=by_slip_week, palletlines_source=palletlines_source)
        results.append(r)
    _apply_reallocate_to_prior_week(results, df)
    if apply_manual_adjustments:
        _apply_manual_adjustments(results, load_traceability_adjustments())
    _apply_weekly_production_overrides(results, load_weekly_production_overrides())
    if apply_manual_adjustments and _failsafe_to_100_enabled():
        _apply_failsafe_topup_to_100(results)
    return results


def _expected_weeks_from_inputs():
    """Return sorted list of week numbers that appear in SQF_Traceability_Inputs.csv."""
    df = load_sqf_inputs()
    if df is None or df.empty or "Week" not in df.columns:
        return []
    weeks = []
    for v in df["Week"]:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        try:
            w = int(str(v).strip())
        except (TypeError, ValueError):
            continue
        weeks.append(w)
    return sorted(sorted(set(weeks)))


def validate_original_palletlines():
    """
    Ensure each expected week has at least one original PalletLines workbook.

    Originals are defined as the latest non-adjusted PalletLines workbooks found
    either in the top-level Traceability_Exports folder or the optional
    Traceability_Exports/Original subfolder.
    """
    expected_weeks = _expected_weeks_from_inputs()
    originals = _latest_palletlines_files_by_week(
        [EXPORTS_DIR, ORIGINAL_EXPORTS_DIR],
        source_mode=PALLETLINES_SOURCE_ORIGINAL_ONLY,
    )
    missing = [w for w in expected_weeks if w not in originals]
    if missing:
        raise RuntimeError(
            "Missing original PalletLines for weeks: "
            f"{missing}. Looked in Traceability_Exports and Traceability_Exports/Original."
        )


def validate_balanced_working():
    """Ensure BalancedWorking has an effective workbook per expected week."""
    expected_weeks = _expected_weeks_from_inputs()
    effective = build_balanced_working_view()
    missing = [w for w in expected_weeks if w not in effective]
    if missing:
        raise RuntimeError(f"Missing BalancedWorking PalletLines for weeks: {missing}. Run loading slip processing or balancing first.")


def run_all_weeks_original_to_csv(output_dir=None, csv_filename="Traceability_2025_original.csv"):
    """
    Convenience helper: run all weeks using ORIGINAL PalletLines only and export to CSV.
    """
    # Fail fast if any expected week is missing originals
    validate_original_palletlines()
    # Originals: use Original/ if it has any original PalletLines, else top-level for backward compat.
    orig_in_original = list(ORIGINAL_EXPORTS_DIR.glob("Week*_AllDays_PalletLines*.xlsx")) if ORIGINAL_EXPORTS_DIR.exists() else []
    orig_in_original = [p for p in orig_in_original if not _is_adjusted_palletlines_file(p)]
    source_dir = ORIGINAL_EXPORTS_DIR if orig_in_original else EXPORTS_DIR
    rows = run_all_weeks(source_dir, palletlines_source=PALLETLINES_SOURCE_ORIGINAL_ONLY)
    if not rows:
        print("No results for original traceability run; CSV will not be written.")
        return None
    out_root = Path(output_dir) if output_dir else ORIGINAL_EXPORTS_DIR
    out_root.mkdir(parents=True, exist_ok=True)
    path = out_root / csv_filename
    import csv

    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    print(f"Wrote original traceability CSV to: {path}")
    return path


def run_all_weeks_balanced_to_csv(output_dir=None, csv_filename="Traceability_2025_balanced.csv"):
    """
    Convenience helper: run all weeks using the BalancedWorking view (adjusted where present)
    and export to CSV, with lineage fields for each week.
    """
    # Ensure BalancedWorking has full coverage and get mapping {slip_week: path}
    effective_map = build_balanced_working_view()
    # Fail loudly if any expected week has no effective workbook
    validate_balanced_working()
    source_dir = BALANCED_WORKING_DIR

    rows = run_all_weeks(source_dir, palletlines_source=PALLETLINES_SOURCE_PREFERRED)
    if not rows:
        print("No results for balanced traceability run; CSV will not be written.")
        return None
    # Build quick lookup of whether a week is adjusted and which file drove it
    week_to_file = {}
    for slip_week, path in effective_map.items():
        week_to_file[int(slip_week)] = path

    # Enrich result rows with lineage for the balanced CSV
    for r in rows:
        week = r.get("Week")
        try:
            w_int = int(week) if week is not None else None
        except (TypeError, ValueError):
            w_int = None
        path = week_to_file.get(w_int)
        if path is None:
            source_mode = "Original"
            file_name = ""
        else:
            source_mode = "Adjusted" if _is_adjusted_palletlines_file(path) else "Original"
            file_name = path.name
        r["Slip_Week_ID"] = slip_week_id(r.get("Year", 0), week or 0)
        r["Source_Mode"] = source_mode
        r["Effective_PalletLines_File"] = file_name
        r["Week_Adjusted_Flag"] = source_mode == "Adjusted"

    out_root = Path(output_dir) if output_dir else BALANCED_WORKING_DIR
    out_root.mkdir(parents=True, exist_ok=True)
    path = out_root / csv_filename
    import csv

    try:
        with path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)
        print(f"Wrote balanced traceability CSV to: {path}")
        return path
    except PermissionError:
        # If the standard filename is locked (e.g. open in Excel), fall back to
        # a timestamped variant rather than failing the entire balanced run.
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt_path = out_root / f"{path.stem}_{ts}{path.suffix}"
        with alt_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)
        print(
            "Standard balanced CSV name was locked; wrote to alternate file instead:\n"
            f"  {alt_path}"
        )
        return alt_path


def _find_latest_balanced_csv(exports_dir=None):
    """Return path to the most recent Traceability_2025_balanced*.csv (BalancedWorking, then top-level)."""
    roots = [Path(exports_dir)] if exports_dir else [BALANCED_WORKING_DIR, EXPORTS_DIR]
    all_candidates = []
    for root in roots:
        if not root.exists():
            continue
        all_candidates.extend(root.glob("Traceability_2025_balanced*.csv"))
    if not all_candidates:
        return None
    return max(all_candidates, key=lambda p: p.stat().st_mtime)


def report_weeks_near_target(balanced_csv_path=None, band_low=99.5, band_high=100.5, exports_dir=None):
    """
    Read the balanced traceability CSV and report weeks within ±0.5% of 100%
    (i.e. band_low <= Traceability_Pct <= band_high, with Production > 0).
    Prints count, sorted list of weeks, and a short table.
    Returns (count, list_of_weeks, list_of_row_dicts).
    """
    import csv
    path = Path(balanced_csv_path) if balanced_csv_path else _find_latest_balanced_csv(exports_dir)
    if path is None or not path.exists():
        print("No balanced traceability CSV found in Traceability_Exports (Traceability_2025_balanced*.csv).")
        return 0, [], []

    with path.open(encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    def _num(val, default=0):
        if val is None or (isinstance(val, str) and val.strip() == ""):
            return default
        try:
            return float(val)
        except (TypeError, ValueError):
            return default

    near = []
    for r in rows:
        week_val = r.get("Week")
        if week_val is None or (isinstance(week_val, str) and not week_val.strip()):
            continue
        try:
            week = int(float(week_val))
        except (TypeError, ValueError):
            continue
        pct = _num(r.get("Traceability_Pct"), None)
        prod = _num(r.get("Production"), 0)
        if pct is None or prod <= 0:
            continue
        if band_low <= pct <= band_high:
            near.append({
                "Week": week,
                "Traceability_Pct": pct,
                "Accounted": _num(r.get("Accounted")),
                "Production": prod,
                "Shipped_Orders": _num(r.get("Shipped_Orders")),
                "Week_Adjusted_Flag": r.get("Week_Adjusted_Flag", ""),
                "Effective_PalletLines_File": r.get("Effective_PalletLines_File", ""),
            })

    near.sort(key=lambda x: x["Week"])
    weeks_list = [x["Week"] for x in near]
    count = len(weeks_list)

    print("\n=== Weeks within ±0.5% of 100% traceability (balanced CSV) ===")
    print(f"Source: {path.name}")
    print(f"Band: {band_low}% <= Traceability_Pct <= {band_high}% (Production > 0)")
    print(f"Count: {count} weeks")
    print(f"Weeks: {weeks_list}")
    if near:
        print("\nWeek | Trace%  | Accounted | Production | Shipped_Orders | Adjusted | Effective_PalletLines_File")
        print("-----+---------+-----------+------------+----------------+----------+---------------------------")
        for x in near:
            adj = "YES" if str(x["Week_Adjusted_Flag"]).lower() in ("true", "1") else "NO"
            fname = (x["Effective_PalletLines_File"] or "")[:40]
            print(
                f"{x['Week']:>4} | {x['Traceability_Pct']:6.2f} | {x['Accounted']:9.0f} | {x['Production']:10.0f} | "
                f"{x['Shipped_Orders']:14.0f} | {adj:^8} | {fname}"
            )
    return count, weeks_list, near


def reconcile_original_vs_balanced(original_csv, balanced_csv):
    """Print a concise reconciliation summary between original and balanced CSVs."""
    import csv
    from pathlib import Path

    o_path = Path(original_csv)
    b_path = Path(balanced_csv)
    if not o_path.exists() or not b_path.exists():
        print("Reconciliation skipped: original or balanced CSV missing.")
        return
    with o_path.open(encoding="utf-8") as f:
        orig_rows = list(csv.DictReader(f))
    with b_path.open(encoding="utf-8") as f:
        bal_rows = list(csv.DictReader(f))
    by_o = {int(r["Week"]): r for r in orig_rows if r.get("Week")}
    by_b = {int(r["Week"]): r for r in bal_rows if r.get("Week")}
    weeks_o = set(by_o)
    weeks_b = set(by_b)
    print("\n=== Traceability reconciliation (original vs balanced) ===")
    print(f"Weeks in original:  {sorted(weeks_o)}")
    print(f"Weeks in balanced:  {sorted(weeks_b)}")
    shared = sorted(weeks_o & weeks_b)
    if not shared:
        print("No overlapping weeks between original and balanced CSVs.")
        return

    def _num(val):
        try:
            return float(val or 0)
        except (TypeError, ValueError):
            return 0.0

    total_shipped_o = sum(_num(by_o[w].get("Shipped_Orders")) for w in shared)
    total_shipped_b = sum(_num(by_b[w].get("Shipped_Orders")) for w in shared)
    total_accounted_o = sum(_num(by_o[w].get("Accounted")) for w in shared)
    total_accounted_b = sum(_num(by_b[w].get("Accounted")) for w in shared)

    print(f"Total Shipped_Orders original: {total_shipped_o:.2f}")
    print(f"Total Shipped_Orders balanced: {total_shipped_b:.2f}")
    print(f"Total Accounted original:      {total_accounted_o:.2f}")
    print(f"Total Accounted balanced:      {total_accounted_b:.2f}")

    print("\nWeek | Shipped_Orig | Shipped_Bal | Trace%_Orig | Trace%_Bal | Adjusted")
    print("-----+--------------+-------------+------------+-----------+---------")
    adjusted_weeks = []
    for w in shared:
        ro = by_o[w]
        rb = by_b[w]
        so = _num(ro.get("Shipped_Orders"))
        sb = _num(rb.get("Shipped_Orders"))
        to = ro.get("Traceability_Pct")
        tb = rb.get("Traceability_Pct")
        try:
            to_f = float(to) if to not in (None, "") else 0.0
        except (TypeError, ValueError):
            to_f = 0.0
        try:
            tb_f = float(tb) if tb not in (None, "") else 0.0
        except (TypeError, ValueError):
            tb_f = 0.0
        adjusted_flag = str(rb.get("Week_Adjusted_Flag", "")).lower() in ("true", "1")
        if adjusted_flag:
            adjusted_weeks.append(w)
        print(
            f"{w:>4} | {so:12.0f} | {sb:11.0f} | {to_f:10.2f} | {tb_f:9.2f} | "
            f"{'YES' if adjusted_flag else 'NO'}"
        )

    print(f"\nAdjusted weeks (balanced): {adjusted_weeks}")


def print_report(results, week_filter=None, by_slip_week=False):
    """Print console report. week_filter = single week number or None for all."""
    if not results:
        print("No data. Add rows to Reference_Data/SQF_Traceability_Inputs.csv (Week, Production, Shipped_NestRun, OffGrades_GradeOut).")
        print("Shipped_Orders is auto from PalletLines.")
        return
    if week_filter is not None:
        results = [r for r in results if r["Week"] == week_filter]
        if not results:
            print(f"No data for week {week_filter}.")
            return
    print("=" * 70)
    print("SQF TRACEABILITY RECONCILIATION (Producer-Grader)")
    print("=" * 70)
    print(f"Target: {TARGET_PCT}%+ traceability (Accounted / Production)")
    print("Accounted = Shipped_Orders + Shipped_NestRun + OffGrades_GradeOut + Eggs_Stored")
    print("Data sources: Production=Hilly Acres Inputs | Orders=PalletLines | NestRun=CSV or Hilly Acres Pallet Information | OffGrades=gradeout CSV")
    print(f"When no off-grade data for a week: OffGrades = {DEFAULT_OFFGRADE_PCT}% of Production.")
    if by_slip_week:
        print("Shipped_Orders mode: slip-week (sum all boxes from Week N PalletLines file)")
    elif _get_order_attribution_mode() == "slip_day":
        print("Shipped_Orders mode: slip_day (ReportWeek from PalletLines: Wed/Thu/Fri of slip N + Mon/Tue of slip N+1 = Report N)")
        print("  Aligns with Hilly Acres production week. Set OrderAttributionMode in paths.json.")
    else:
        bb_days = _get_bb_to_pack_days()
        print(f"Shipped_Orders mode: bb_date (PackDate = BBDate - {bb_days} days -> production week)")
        print(f"  Set OrderAttributionMode=slip_day in paths.json for slip-day alignment.")
    print()
    for r in results:
        pct = r["Traceability_Pct"]
        status = "OK" if r["Meets_Target"] else "BELOW TARGET"
        if pct is None:
            pct_str = "N/A (set Production)"
        else:
            pct_str = f"{pct}%"
        print(f"{r['Year']} Week {r['Week']}:")
        print(f"  Production (Hilly Acres Inputs, TOTAL stacks): {r['Production']:,.0f}")
        print(f"  Shipped_Orders (PalletLines loading slips):    {r['Shipped_Orders']:,}")
        print(f"  Shipped_NestRun (CSV or Hilly Acres Pallet Info): {r['Shipped_NestRun']:,.0f}")
        print(f"  OffGrades_GradeOut (gradeout CSV + Hilly Acres):{r['OffGrades_GradeOut']:,.0f}")
        if r.get("Eggs_Stored"):
            print(f"  Eggs_Stored (flock transition):                {r['Eggs_Stored']:,.0f}")
        if r.get("Manual_Adjustments"):
            print(f"  Manual_Adjustments (evidence-backed):          {r['Manual_Adjustments']:,.0f}")
        if r.get("Failsafe_Topup_Cases"):
            print(f"  Failsafe_To100 (from authorized CSV):         {r['Failsafe_Topup_Cases']:,.0f}")
        print(f"  Accounted (sum):                               {r['Accounted']:,.0f}")
        print(f"  Variance (Production - Accounted):             {r['Variance']:,.0f}")
        print(f"  Traceability: {pct_str}  [{status}]")
        core_pct = r.get("Core_Traceability_Pct")
        if core_pct is not None:
            print(
                f"  Core (Orders+Nest only vs Production):         {core_pct}%  "
                f"(Core_Accounted {r.get('Core_Accounted', 0):,.0f}, Core_Variance {r.get('Core_Variance', 0):,.0f})"
            )
        print()
    below = [r["Week"] for r in results if not r["Meets_Target"] and r["Traceability_Pct"] is not None]
    na_weeks = [r["Week"] for r in results if r["Traceability_Pct"] is None and r["Production"] == 0]
    if below:
        print(f"Weeks below {TARGET_PCT}%: {below}")
    if na_weeks:
        print(f"Weeks with no Production entered (fill SQF_Traceability_Inputs.csv): {na_weeks[:10]}{'...' if len(na_weeks) > 10 else ''}")
    if not below and not na_weeks:
        print("All reported weeks meet target.")


def apply_inventory_flow_balance(results, max_carry_weeks=2, apply_to_traceability=False):
    """
    Apply FIFO carryover allocation across weeks.

    Models the reality that eggs produced in week N may be consumed in later weeks.
    Adds Flow_* fields to each result row. Optionally applies flow metrics to the
    primary traceability outputs for sandbox experimentation.
    """
    if not results:
        return results

    max_weeks = max(0, int(max_carry_weeks or 0))
    out = []
    buckets = []  # {"source_week": int, "age": int, "remaining": float}

    for r in results:
        row = dict(r)
        week = int(row.get("Week") or 0)
        production = float(row.get("Production") or 0)
        shipped_orders = float(row.get("Shipped_Orders") or 0)
        nest_run = float(row.get("Shipped_NestRun") or 0)
        offgrades = float(row.get("OffGrades_GradeOut") or 0)
        manual_adj = float(row.get("Manual_Adjustments") or 0)
        failsafe = float(row.get("Failsafe_Topup_Cases") or 0)

        for b in buckets:
            b["age"] = int(b.get("age", 0)) + 1

        expired_cases = 0.0
        kept = []
        for b in buckets:
            if int(b.get("age", 0)) > max_weeks:
                expired_cases += float(b.get("remaining", 0) or 0)
            else:
                kept.append(b)
        buckets = kept

        carry_in = float(sum(float(b.get("remaining", 0) or 0) for b in buckets))
        if production > 0:
            buckets.append({"source_week": week, "age": 0, "remaining": production})

        demand_core = shipped_orders + nest_run + offgrades
        available_now = float(sum(float(b.get("remaining", 0) or 0) for b in buckets))
        allocated_core = min(available_now, demand_core)

        consume = allocated_core
        for b in sorted(buckets, key=lambda x: (int(x.get("age", 0)), int(x.get("source_week", 0))), reverse=True):
            if consume <= 0:
                break
            rem = float(b.get("remaining", 0) or 0)
            if rem <= 0:
                continue
            take = rem if rem <= consume else consume
            b["remaining"] = rem - take
            consume -= take

        carry_out = float(sum(float(b.get("remaining", 0) or 0) for b in buckets))
        shortfall_core = max(0.0, demand_core - allocated_core)

        # Manual/failsafe remain evidence overlays, not inventory timing consumption.
        flow_accounted = allocated_core + manual_adj + failsafe
        flow_variance = production - flow_accounted
        flow_pct = round(flow_accounted / production * 100.0, 2) if production > 0 else None
        flow_meets = bool(flow_pct is not None and flow_pct >= TARGET_PCT)

        row["Flow_CarryIn"] = round(carry_in, 2)
        row["Flow_Demand_Core"] = round(demand_core, 2)
        row["Flow_Allocated_Core"] = round(allocated_core, 2)
        row["Flow_Core_Shortfall"] = round(shortfall_core, 2)
        row["Flow_CarryOut"] = round(carry_out, 2)
        row["Flow_Expired_Cases"] = round(expired_cases, 2)
        row["Flow_Accounted"] = round(flow_accounted, 2)
        row["Flow_Variance"] = round(flow_variance, 2)
        row["Flow_Traceability_Pct"] = flow_pct
        row["Flow_Meets_Target"] = flow_meets

        # Optional conservative apply mode:
        # only replace baseline traceability when flow is strictly closer to 100%
        # (smaller absolute variance to production).
        row["Flow_Applied_To_Traceability"] = False
        if apply_to_traceability:
            base_accounted = float(row.get("Accounted") or 0)
            base_gap = abs(float(production - base_accounted))
            flow_gap = abs(float(flow_variance))
            if flow_gap < base_gap:
                row["Accounted"] = row["Flow_Accounted"]
                row["Variance"] = row["Flow_Variance"]
                row["Traceability_Pct"] = row["Flow_Traceability_Pct"]
                row["Meets_Target"] = row["Flow_Meets_Target"]
                row["Flow_Applied_To_Traceability"] = True

        out.append(row)

    return out


def _fmt_reference_file(path_obj):
    """Return stable display metadata for a reference file/path."""
    p = Path(path_obj)
    exists = p.exists()
    mtime = ""
    if exists:
        try:
            mtime = datetime.fromtimestamp(p.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            mtime = ""
    return {
        "Reference_Name": p.name,
        "Reference_Path": _display_path(p),
        "Exists": "Yes" if exists else "No",
        "Last_Modified": mtime,
    }


def build_reference_files_inventory():
    """Build a file-level inventory of major reference inputs used by traceability."""
    rows = []
    base_files = [
        INPUT_CSV,
        ADJUSTMENTS_CSV,
        PRODUCTION_ADJUSTMENTS_CSV,
        PRODUCTION_WEEKLY_OVERRIDES_CSV,
        FAILSAFE_EVIDENCE_CSV,
        REF_DATA_DIR / "paths.json",
        GRADE_OUTS_DIR / GRADEOUT_CSV,
    ]
    for f in base_files:
        rows.append(_fmt_reference_file(f))

    # Include the effective pallet lines file for each available week.
    for slip_week, p in sorted(_latest_palletlines_files_by_week(EXPORTS_DIR, source_mode=PALLETLINES_SOURCE_PREFERRED).items()):
        rec = _fmt_reference_file(p)
        rec["Reference_Name"] = f"PalletLines Week {slip_week}"
        rows.append(rec)

    return pd.DataFrame(rows, columns=["Reference_Name", "Reference_Path", "Exists", "Last_Modified"])


def build_reference_usage_by_week(results, exports_dir=None):
    """Week-level view of exactly which reference sources were used."""
    exports_dir = Path(exports_dir) if exports_dir else EXPORTS_DIR
    source_rows = run_inputs_report(exports_dir=exports_dir, by_slip_week=False)
    src_df = pd.DataFrame(source_rows)
    if not src_df.empty and "Week" in src_df.columns:
        src_df["Week"] = pd.to_numeric(src_df["Week"], errors="coerce")

    out_rows = []
    for r in results:
        try:
            week = int(r.get("Week") or 0)
            year = int(r.get("Year") or 0)
        except Exception:
            continue
        src_row = {}
        if not src_df.empty and "Week" in src_df.columns:
            m = src_df[src_df["Week"] == week]
            if not m.empty:
                src_row = m.iloc[0].to_dict()
        out_rows.append({
            "Year": year,
            "Week": week,
            "Traceability_Pct": r.get("Traceability_Pct"),
            "Production": r.get("Production"),
            "Shipped_Orders": r.get("Shipped_Orders"),
            "Shipped_NestRun": r.get("Shipped_NestRun"),
            "OffGrades_GradeOut": r.get("OffGrades_GradeOut"),
            "Eggs_Stored": r.get("Eggs_Stored"),
            "Manual_Adjustments": r.get("Manual_Adjustments"),
            "Adjustment_Categories": r.get("Adjustment_Categories"),
            "Slip_Week_ID": r.get("Slip_Week_ID", ""),
            "Source_Mode": r.get("Source_Mode", ""),
            "Effective_PalletLines_File": r.get("Effective_PalletLines_File", ""),
            "Production_Source": src_row.get("Production_Source", ""),
            "PalletLines_Available": src_row.get("PalletLines", ""),
            "Gradeout_Source": src_row.get("Gradeout_Source", ""),
            "Missing_Inputs": src_row.get("Missing_Inputs", ""),
        })

    cols = [
        "Year", "Week", "Traceability_Pct",
        "Production", "Shipped_Orders", "Shipped_NestRun", "OffGrades_GradeOut", "Eggs_Stored",
        "Manual_Adjustments", "Adjustment_Categories",
        "Slip_Week_ID", "Source_Mode", "Effective_PalletLines_File",
        "Production_Source", "PalletLines_Available", "Gradeout_Source", "Missing_Inputs",
    ]
    return pd.DataFrame(out_rows, columns=cols)


def export_excel(results, path=None, exports_dir=None):
    """Write one Excel file for audit: traceability plus transparent 2025 reallocation detail."""
    if path is None:
        path = EXPORTS_DIR / f"SQF_Traceability_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    exports_dir = Path(exports_dir) if exports_dir else EXPORTS_DIR
    df = pd.DataFrame(results)
    if not df.empty and "Week" in df.columns:
        week_dates = df["Week"].apply(_report_week_to_dates)
        df["Week_Start"] = week_dates.apply(lambda x: x[0].strftime("%Y-%m-%d") if x and x[0] else "")
        df["Week_Ending"] = week_dates.apply(lambda x: x[1].strftime("%Y-%m-%d") if x and x[1] else "")
    col_order = ["Year", "Week", "Week_Start", "Week_Ending"] + [c for c in df.columns if c not in ("Year", "Week", "Week_Start", "Week_Ending")]
    df = df[[c for c in col_order if c in df.columns]]
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="Traceability", index=False)
        summary = [
            {"Metric": "Target traceability %", "Value": f"{TARGET_PCT}%"},
            {"Metric": "OrderAttributionMode", "Value": _get_order_attribution_mode()},
            {"Metric": "BBToPackDays (when bb_date mode)", "Value": _get_bb_to_pack_days()},
            {"Metric": "Weeks at or above target", "Value": sum(1 for r in results if r["Meets_Target"])},
            {"Metric": "Weeks below target", "Value": sum(1 for r in results if r["Traceability_Pct"] is not None and not r["Meets_Target"])},
            {"Metric": "Manual adjustments file", "Value": str(ADJUSTMENTS_CSV.name)},
        ]
        pd.DataFrame(summary).to_excel(w, sheet_name="Summary", index=False)
        inputs_df = load_sqf_inputs()
        adjustments_df = load_traceability_adjustments()
        gap_df = build_gap_analysis(results, exports_dir, inputs_df, adjustments_df)
        if not gap_df.empty:
            gap_df.to_excel(w, sheet_name="Gap_Analysis", index=False)
        balancing_adjustments_df, balancing_summary_df, balancing_candidates_df = load_balancing_logs_from_exports(exports_dir)
        if not balancing_adjustments_df.empty:
            balancing_adjustments_df.to_excel(w, sheet_name="Balancing_Adjustments", index=False)
        if not balancing_summary_df.empty:
            balancing_summary_df.to_excel(w, sheet_name="Balancing_By_Week", index=False)
        if not balancing_candidates_df.empty:
            balancing_candidates_df.to_excel(w, sheet_name="Balancing_Candidates", index=False)
        crosswalk_df = build_week_crosswalk(results, exports_dir, inputs_df)
        if not crosswalk_df.empty:
            crosswalk_df.to_excel(w, sheet_name="Week_Crosswalk", index=False)
        ref_files_df = build_reference_files_inventory()
        if not ref_files_df.empty:
            ref_files_df.to_excel(w, sheet_name="Reference_Files_Used", index=False)
        ref_usage_df = build_reference_usage_by_week(results, exports_dir)
        if not ref_usage_df.empty:
            ref_usage_df.to_excel(w, sheet_name="Reference_Usage_By_Week", index=False)
        flow_df = pd.DataFrame(results)
        if not flow_df.empty and "Flow_Allocated_Core" in flow_df.columns:
            flow_cols = [
                "Year", "Week", "Production",
                "Shipped_Orders", "Shipped_NestRun", "OffGrades_GradeOut", "Manual_Adjustments",
                "Flow_CarryIn", "Flow_Demand_Core", "Flow_Allocated_Core", "Flow_Core_Shortfall",
                "Flow_CarryOut", "Flow_Expired_Cases",
                "Flow_Accounted", "Flow_Variance", "Flow_Traceability_Pct", "Flow_Meets_Target",
            ]
            keep = [c for c in flow_cols if c in flow_df.columns]
            flow_df[keep].to_excel(w, sheet_name="Inventory_Flow_By_Week", index=False)
        adjustments_df.to_excel(w, sheet_name="Traceability_Adjustments", index=False)
        if PRODUCTION_ADJUSTMENTS_CSV.exists():
            try:
                pd.read_csv(PRODUCTION_ADJUSTMENTS_CSV).to_excel(w, sheet_name="Production_Adjustments", index=False)
            except Exception:
                pass
        # Eggs_Stored: weeks with stored eggs (e.g. Barn 6 flock transition 30-35). Explains extra production.
        eggs_stored_rows = [{"Year": r["Year"], "Week": r["Week"], "Eggs_Stored": r.get("Eggs_Stored") or 0} for r in results if (r.get("Eggs_Stored") or 0) > 0]
        if eggs_stored_rows:
            eggs_stored_df = pd.DataFrame(eggs_stored_rows)
            eggs_stored_df.to_excel(w, sheet_name="Eggs_Stored", index=False)
        suggested_eggs_stored_df = build_suggested_eggs_stored(results, gap_df)
        if not suggested_eggs_stored_df.empty:
            suggested_eggs_stored_df.to_excel(w, sheet_name="Suggested_Eggs_Stored", index=False)
        suggested_realloc_df = build_suggested_reallocations(results, gap_df)
        if not suggested_realloc_df.empty:
            suggested_realloc_df.to_excel(w, sheet_name="Suggested_Reallocation", index=False)
        suggested_adjust_df = build_suggested_adjustments(results, gap_df, suggested_realloc_df)
        if not suggested_adjust_df.empty:
            suggested_adjust_df.to_excel(w, sheet_name="Suggested_Adjustments", index=False)
        old_new_df = build_all_2025_old_new_by_day(exports_dir, inputs_df)
        if not old_new_df.empty:
            old_new_df.to_excel(w, sheet_name="OldDate_NewDate_ByDay", index=False)
            old_new_sku_df = build_all_2025_old_new_by_day_sku(exports_dir, inputs_df)
            reallocation_guide_df = build_2025_reallocation_guide(old_new_df, old_new_sku_df)
            if not reallocation_guide_df.empty:
                reallocation_guide_df.to_excel(w, sheet_name="Reallocation_Guide", index=False)
            if not old_new_sku_df.empty:
                old_new_sku_df.to_excel(w, sheet_name="OldDate_NewDate_BySKU", index=False)
        # Reconciliation_Steps: audit-ready guide for OD vs new-date when we don't have OD column data
        reconciliation_steps = pd.DataFrame([
            {"Step": 1, "Action": "Weeks OVER 100% (too much accounted)", "How": "Use Suggested_Reallocation sheet. Copy Suggested_Reallocate_To_Prior_Week to SQF_Traceability_Inputs.csv column Reallocate_To_Prior_Week for that From_Week row. Re-run traceability."},
            {"Step": 2, "Action": "Order attribution (consistent, aligns with Hilly Acres)", "How": "Default: slip_day. ReportWeek = Wed/Thu/Fri of slip N + Mon/Tue of slip N+1 = Report N. Set OrderAttributionMode in paths.json (slip_day or bb_date)."},
            {"Step": 3, "Action": "OD vs New Date justification (no OD column on slips)", "How": "OldDate_NewDate_ByDay uses BB date: Cases_OldDate = cases where BB date maps to prior production week. Audit accepts this as justification for reallocating to prior week."},
            {"Step": 4, "Action": "Weeks UNDER 99% (gap to target)", "How": "Use Traceability_Adjustments.csv: add rows (Week, Category, Cases, Evidence, Note) for evidence-backed buckets (NestRun, Internal Use, Writeoff). Or adjust NestRun/OffGrades in SQF_Traceability_Inputs.csv."},
            {"Step": 5, "Action": "No gradeout data for a week", "How": f"1.2% of Production is used as default OffGrades. Override by filling OffGrades_GradeOut in SQF_Traceability_Inputs.csv."},
            {"Step": 6, "Action": "Eggs stored during flock transitions", "How": "When a barn transitions flocks (e.g. Barn 6 weeks 30-35), eggs may be stored and rotated for grading later. Add Eggs_Stored (cases) in SQF_Traceability_Inputs.csv for those weeks. See Eggs_Stored and Suggested_Eggs_Stored sheets."},
            {"Step": 7, "Action": "Reallocation_Guide sheet", "How": "Shows by Day/SKU which cases have BB dates in prior week (Max_Move_To_Prior_Week). Use to justify specific SKU quantities when auditor asks."},
        ])
        reconciliation_steps.to_excel(w, sheet_name="Reconciliation_Steps", index=False)
        # Order_Detail_All: every order line with Week, StopName, SKU, QtyBoxes, BBDate, CaseEquiv. Filter to Week=N to see every box for that week.
        order_detail_df = build_order_detail_all(exports_dir, inputs_df)
        if not order_detail_df.empty:
            order_detail_df.to_excel(w, sheet_name="Order_Detail_All", index=False)
        # Week_Reconciliation: Production + NestRun + OffGrades + Orders (from detail) = Accounted. Orders_From_Detail must match Shipped_Orders.
        week_recon_df = build_week_reconciliation(results, order_detail_df, exports_dir, inputs_df)
        if not week_recon_df.empty:
            week_recon_df.to_excel(w, sheet_name="Week_Reconciliation", index=False)
        # Per-week detail sheets: "Week46_Detail" etc. Transparent audit - Production, Nest Run, Off Grades, every order with BB date, all add up to ~100%.
        by_week_result = {r["Week"]: r for r in results}
        detail_cols = ["DayName", "StopName", "SKU", "QtyBoxes", "BBDate", "CaseEquiv"]
        for week in sorted(set(order_detail_df["Week"].unique()) if not order_detail_df.empty else []):
            subset = order_detail_df[order_detail_df["Week"] == week]
            if subset.empty:
                continue
            r = by_week_result.get(week, {})
            prod = int(r.get("Production") or 0)
            prod_adjustment = int(r.get("Production_Adjustment") or 0)
            nest = int(r.get("Shipped_NestRun") or 0)
            off = int(round(r.get("OffGrades_GradeOut") or 0))
            eggs_stored = int(r.get("Eggs_Stored") or 0)
            orders_sum = int(round(subset["CaseEquiv"].sum()))
            accounted = int(round(r.get("Accounted") or 0))
            pct = r.get("Traceability_Pct")
            sheet_name = f"Week{week}_Detail"[:31]
            summary_rows = [
                ["Metric", "Value", "", "", "", ""],
                ["Week", week, "", "", "", ""],
                ["Production", prod, "", "", "", ""],
                ["Production_Adjustment", prod_adjustment, "", "", "", ""],
                ["Shipped_NestRun", nest, "", "", "", ""],
                ["OffGrades_GradeOut", off, "", "", "", ""],
            ]
            if eggs_stored > 0:
                summary_rows.append(["Eggs_Stored", eggs_stored, "", "", "", ""])
            summary_rows.extend([
                ["Shipped_Orders (sum below)", orders_sum, "", "", "", ""],
                ["Accounted", accounted, "", "", "", ""],
                ["Traceability_Pct", f"{pct}%" if pct is not None else "N/A", "", "", "", ""],
                ["", "", "", "", "", ""],
                ["Order Detail: every box, every stop, each BB date. Sum CaseEquiv = Shipped_Orders.", "", "", "", "", ""],
                ["", "", "", "", "", ""],
                detail_cols,
            ])
            summary_df = pd.DataFrame(summary_rows, columns=detail_cols)
            detail_df = subset[detail_cols].copy()
            combined = pd.concat([summary_df, detail_df], ignore_index=True)
            combined.to_excel(w, sheet_name=sheet_name, index=False)
    print(f"Exported: {path}")
    return path


def export_traceability_working(results, path=None):
    """
    Export week-by-week workbook for reaching ~100% traceability.
    Columns: Week, Production, Orders, OffGrade_Pct, OffGrades_Total, NestRun, Accounted, Variance, Traceability_Pct.
    Off grades are from gradeout percentages x production (spread by barn/day). Edit NestRun (and optionally
    OffGrade_Pct) to see impact; Accounted, Variance, Traceability_Pct use formulas.
    """
    if path is None:
        path = EXPORTS_DIR / "Traceability_By_Week_Working.xlsx"
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "By_Week"
    headers = [
        "Year", "Week", "Week_Start", "Week_Ending", "Production", "Orders", "OffGrade_Pct (editable)", "OffGrades_Total",
        "NestRun (editable)", "Accounted", "Variance", "Traceability_Pct"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    ws.row_dimensions[1].font = Font(bold=True)
    for idx, r in enumerate(results, 2):
        prod = r["Production"] or 0
        orders = r["Shipped_Orders"] or 0
        off_total = r["OffGrades_GradeOut"] or 0
        off_pct = (100.0 * off_total / prod) if prod and prod > 0 else 0
        nest = r["Shipped_NestRun"] or 0
        week_start, week_ending = _report_week_to_dates(r["Week"])
        ws.cell(row=idx, column=1, value=r["Year"])
        ws.cell(row=idx, column=2, value=r["Week"])
        ws.cell(row=idx, column=3, value=week_start.strftime("%Y-%m-%d") if week_start else "")
        ws.cell(row=idx, column=4, value=week_ending.strftime("%Y-%m-%d") if week_ending else "")
        ws.cell(row=idx, column=5, value=prod)
        ws.cell(row=idx, column=6, value=orders)
        ws.cell(row=idx, column=7, value=round(off_pct, 2))
        ws.cell(row=idx, column=8, value=f"=E{idx}*G{idx}/100")
        ws.cell(row=idx, column=9, value=nest)
        ws.cell(row=idx, column=10, value=f"=F{idx}+H{idx}+I{idx}")
        ws.cell(row=idx, column=11, value=f"=E{idx}-J{idx}")
        ws.cell(row=idx, column=12, value=f'=IF(E{idx}>0,ROUND(J{idx}/E{idx}*100,2),"")')
    for col in range(5, 10):
        for row in range(2, len(results) + 2):
            c = ws.cell(row=row, column=col)
            if col == 7:
                c.number_format = '0.00'
            else:
                c.number_format = '#,##0'
    for row in range(2, len(results) + 2):
        for col in (10, 11):
            ws.cell(row=row, column=col).number_format = '#,##0'
        ws.cell(row=row, column=12).number_format = '0.00'
    ws.cell(row=1, column=7).value = "OffGrade_Pct (edit to override)"
    instructions = wb.create_sheet("Instructions", 0)
    instructions["A1"] = "Traceability by week - working file"
    instructions["A2"] = "Week_Start and Week_Ending are included for date context. Edit NestRun (column I) to see how much nest run would close the gap to 100%."
    instructions["A3"] = "OffGrade_Pct is from gradeout percentages x production (spread by barn/day). You can override it to recalc OffGrades_Total and traceability."
    instructions["A4"] = "Accounted = Orders + OffGrades_Total + NestRun. Variance = Production - Accounted. Traceability % = Accounted/Production*100."
    instructions["A5"] = "Orders are case-equivalents: 30-pack Nova/Loblaws only = 8×30 eggs/box (4/3 case); not OC 30/Sobeys; Jumbo = 12 dozen/box (12/15 case); else 1 per box."
    instructions["A6"] = "To push NestRun back to CSV: update Reference_Data/SQF_Traceability_Inputs.csv column Shipped_NestRun, then re-run with --working to refresh."
    for r in range(1, 7):
        instructions.row_dimensions[r].font = Font(bold=(r == 1))
    wb.save(path)
    print(f"Working file (edit NestRun / OffGrade_Pct to see impact): {path}")
    return path


def export_comprehensive_report(results, path=None):
    """
    Export a comprehensive traceability report that explains every number and how far
    off each week is from 100%, so you can work adjustments (NestRun, Production, etc.).
    Sheets: Glossary, Summary, By_Week (with gap columns), Path_To_100 (only weeks below target).
    """
    if path is None:
        path = EXPORTS_DIR / f"Traceability_Comprehensive_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ----- Sheet 1: Glossary (what each number means) -----
    gloss = wb.active
    gloss.title = "Glossary"
    gloss["A1"] = "Traceability report — what the numbers mean"
    gloss["A1"].font = Font(bold=True, size=12)
    row = 3
    for title, lines in [
        ("Production", [
            "Total case-equivalent production for the week from Hilly Acres (egg production).",
            "Source: Hilly Acres workbooks (Inputs sheet, TOTAL stacks). Paths in Reference_Data/paths.json → HillyAcresPaths.",
            "Match to slip week: Friday ship date + 1 = Week Ending in Hilly Acres.",
        ]),
        ("Shipped_Orders", [
            "Case-equivalent boxes shipped via loading slips (customer orders).",
            "Source: PalletLines files in Traceability_Exports (generated by process_weekly_loading_slip.py from loading slips).",
            "30-pack Nova/Loblaws only = 8×30 eggs/box (4/3 case-equiv vs 180-egg case); not OC 30/Sobeys; Jumbo SKUs = 12 dozen/box (12/15 case-equiv); else 1 case-equiv per box.",
        ]),
        ("Shipped_NestRun", [
            "Surplus nest run (eggs shipped as nest run, not customer orders).",
            "Source: SQF_Traceability_Inputs.csv column Shipped_NestRun if set; otherwise Hilly Acres Pallet Information sheet (NR Total boxes by barn, summed).",
        ]),
        ("OffGrades_GradeOut", [
            "Grade-out / off-grades (eggs removed from table-egg stream).",
            "Source: gradeout_data_export.csv in grade outs folder, or SQF_Traceability_Inputs.csv. If missing, default 1.2% of Production is used.",
        ]),
        ("Eggs_Stored", [
            "Eggs stored during flock transitions (e.g. Barn 6 weeks 30-35). Stored and rotated for grading during transition.",
            "Source: SQF_Traceability_Inputs.csv column Eggs_Stored. Add case-equivalent for weeks when barns transition flocks.",
        ]),
        ("Accounted", [
            "Accounted = Shipped_Orders + Shipped_NestRun + OffGrades_GradeOut + Eggs_Stored",
            "Everything we can attribute to a destination (orders, nest run, grade-out, stored eggs).",
        ]),
        ("Variance", [
            "Variance = Production − Accounted",
            "The gap: production we have not yet accounted for. To reach 100%, Variance must go to zero (either add to Accounted or correct Production if wrong).",
        ]),
        ("Traceability %", [
            "Traceability % = Accounted / Production × 100",
            f"Target: {TARGET_PCT}%+ for SQF. 100% means every case of production is accounted for (orders + nest run + off-grades).",
        ]),
        ("Core_Accounted / Core_Traceability_Pct", [
            "Production vs the slip + nest run chain only: Core_Accounted = Shipped_Orders + Shipped_NestRun (excludes off-grades, eggs stored, manual adjustments).",
            "Core_Traceability_Pct = Core_Accounted / Production × 100. Core_Variance = Production − Core_Accounted.",
        ]),
    ]:
        gloss.cell(row=row, column=1, value=title)
        gloss.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        for line in lines:
            gloss.cell(row=row, column=1, value=line)
            gloss.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
            row += 1
        row += 1
    for col in [1]:
        gloss.column_dimensions[get_column_letter(col)].width = 80

    # ----- Sheet 2: Summary -----
    sum_sheet = wb.create_sheet("Summary", 1)
    sum_sheet["A1"] = "Overall summary"
    sum_sheet["A1"].font = Font(bold=True, size=12)
    total_prod = sum(r.get("Production") or 0 for r in results)
    total_acc = sum(r.get("Accounted") or 0 for r in results)
    overall_pct = (100.0 * total_acc / total_prod) if total_prod and total_prod > 0 else None
    total_var = sum(r.get("Variance") or 0 for r in results)
    at_target = [r["Week"] for r in results if r.get("Meets_Target")]
    below_target = [r["Week"] for r in results if r.get("Traceability_Pct") is not None and not r.get("Meets_Target")]
    no_prod = [r["Week"] for r in results if (r.get("Production") or 0) == 0]
    sum_rows = [
        ("Target traceability %", f"{TARGET_PCT}%"),
        ("Total Production (all weeks)", total_prod),
        ("Total Accounted (all weeks)", total_acc),
        ("Total Variance (all weeks)", total_var),
        ("Overall traceability %", f"{round(overall_pct, 2)}%" if overall_pct is not None else "N/A"),
        ("Weeks at or above target", len(at_target)),
        ("Weeks below target", len(below_target)),
        ("Weeks with no Production", len(no_prod)),
    ]
    for i, (label, val) in enumerate(sum_rows, 2):
        sum_sheet.cell(row=i, column=1, value=label)
        sum_sheet.cell(row=i, column=2, value=val)
    sum_sheet.cell(row=2, column=1).font = Font(bold=True)
    if below_target:
        sum_sheet.cell(row=12, column=1, value="Weeks below target (need attention):")
        sum_sheet.cell(row=12, column=1).font = Font(bold=True)
        sum_sheet.cell(row=13, column=1, value=", ".join(str(w) for w in below_target))

    # ----- Sheet 3: By_Week with gap analysis -----
    by_week = wb.create_sheet("By_Week", 2)
    headers = [
        "Year", "Week", "Week_Start", "Week_Ending", "Production", "Shipped_Orders", "Shipped_NestRun", "OffGrades_GradeOut", "Eggs_Stored",
        "Accounted", "Variance", "Traceability_Pct", "Core_Accounted", "Core_Variance", "Core_Traceability_Pct",
        "Gap_To_100_Pct", "Cases_Short_Of_100", "What_To_Reach_100"
    ]
    for col, h in enumerate(headers, 1):
        by_week.cell(row=1, column=col, value=h)
    by_week.row_dimensions[1].font = Font(bold=True)
    for idx, r in enumerate(results, 2):
        prod = r.get("Production") or 0
        acc = r.get("Accounted") or 0
        var = r.get("Variance") or 0
        pct = r.get("Traceability_Pct")
        gap_pct = (100.0 - pct) if pct is not None else None
        cases_short = int(var) if prod and var > 0 else 0  # cases we need to account for (or reduce prod)
        week_start, week_ending = _report_week_to_dates(r["Week"])
        if pct is None or prod == 0:
            what = "Set Production in SQF_Traceability_Inputs.csv"
        elif pct >= 100:
            what = "At or above 100%"
        else:
            what = f"Add {cases_short:,} to Accounted (e.g. Shipped_NestRun in CSV) OR reduce Production by {cases_short:,} if production was overstated"
        by_week.cell(row=idx, column=1, value=r.get("Year"))
        by_week.cell(row=idx, column=2, value=r["Week"])
        by_week.cell(row=idx, column=3, value=week_start.strftime("%Y-%m-%d") if week_start else "")
        by_week.cell(row=idx, column=4, value=week_ending.strftime("%Y-%m-%d") if week_ending else "")
        by_week.cell(row=idx, column=5, value=prod)
        by_week.cell(row=idx, column=6, value=r.get("Shipped_Orders") or 0)
        by_week.cell(row=idx, column=7, value=r.get("Shipped_NestRun") or 0)
        by_week.cell(row=idx, column=8, value=r.get("OffGrades_GradeOut") or 0)
        by_week.cell(row=idx, column=9, value=r.get("Eggs_Stored") or 0)
        by_week.cell(row=idx, column=10, value=acc)
        by_week.cell(row=idx, column=11, value=var)
        by_week.cell(row=idx, column=12, value=round(pct, 2) if pct is not None else "")
        by_week.cell(row=idx, column=13, value=r.get("Core_Accounted"))
        by_week.cell(row=idx, column=14, value=r.get("Core_Variance"))
        core_pct_row = r.get("Core_Traceability_Pct")
        by_week.cell(row=idx, column=15, value=round(core_pct_row, 2) if core_pct_row is not None else "")
        by_week.cell(row=idx, column=16, value=round(gap_pct, 2) if gap_pct is not None else "")
        by_week.cell(row=idx, column=17, value=cases_short if cases_short else "")
        by_week.cell(row=idx, column=18, value=what)
        by_week.cell(row=idx, column=18).alignment = Alignment(wrap_text=True)
    for c in range(5, 19):
        for row in range(2, len(results) + 2):
            by_week.cell(row=row, column=c).number_format = '#,##0' if c not in (12, 15, 16) else '0.00'
    by_week.column_dimensions["R"].width = 55

    # ----- Sheet 4: Path_To_100 (only weeks below target) -----
    path_sheet = wb.create_sheet("Path_To_100", 3)
    path_sheet["A1"] = "How far off from 100% — and what would fix it"
    path_sheet["A1"].font = Font(bold=True, size=12)
    path_sheet["A2"] = "Only weeks below target are listed. Use this to decide where to add NestRun, correct Production, or investigate missing orders/grade-out."
    path_sheet["A2"].alignment = Alignment(wrap_text=True)
    path_headers = ["Year", "Week", "Production", "Accounted", "Variance (gap)", "Traceability_%", "Gap_%", "Cases_short", "Options_to_reach_100%"]
    for col, h in enumerate(path_headers, 1):
        path_sheet.cell(row=4, column=col, value=h)
    path_sheet.row_dimensions[4].font = Font(bold=True)
    below_rows = [r for r in results if r.get("Traceability_Pct") is not None and not r.get("Meets_Target") and (r.get("Production") or 0) > 0]
    for i, r in enumerate(below_rows, 5):
        prod = r.get("Production") or 0
        acc = r.get("Accounted") or 0
        var = r.get("Variance") or 0
        pct = r.get("Traceability_Pct") or 0
        gap_pct = 100.0 - pct
        path_sheet.cell(row=i, column=1, value=r.get("Year"))
        path_sheet.cell(row=i, column=2, value=r["Week"])
        path_sheet.cell(row=i, column=3, value=prod)
        path_sheet.cell(row=i, column=4, value=acc)
        path_sheet.cell(row=i, column=5, value=var)
        path_sheet.cell(row=i, column=6, value=round(pct, 2))
        path_sheet.cell(row=i, column=7, value=round(gap_pct, 2))
        path_sheet.cell(row=i, column=8, value=int(var))
        path_sheet.cell(row=i, column=9, value=f"Add {int(var):,} to Accounted (NestRun/orders/off-grade) OR reduce Production by {int(var):,} if wrong")
        path_sheet.cell(row=i, column=9).alignment = Alignment(wrap_text=True)
    for c in range(3, 9):
        for row in range(5, 5 + len(below_rows)):
            path_sheet.cell(row=row, column=c).number_format = '#,##0' if c not in (6, 7) else '0.00'
    path_sheet.column_dimensions["I"].width = 52
    if not below_rows:
        path_sheet["A5"] = "No weeks below target — all at or above 100%."

    wb.save(path)
    print(f"Comprehensive report (glossary + gap analysis): {path}")
    return path


def _diagnose_shipped_orders(week_number, exports_dir=None):
    """
    Diagnostic: compare raw PalletLines total vs Option A filtered total.
    Shows why Shipped_Orders might be low (BB date parsing, week number mismatch).
    """
    if exports_dir is None:
        exports_dir = EXPORTS_DIR
    exports_dir = Path(exports_dir)
    if not exports_dir.exists():
        print(f"Exports dir not found: {exports_dir}")
        return
    print("=" * 60)
    print(f"SHIPPED_ORDERS DIAGNOSTIC - Week {week_number}")
    print("=" * 60)
    raw_total = 0
    by_production_week = {}  # (iso_year, iso_week) -> qty
    rows_with_bb = 0
    rows_no_bb = 0
    qty_no_bb = 0
    for slip_week in (week_number, week_number + 1):
        pattern = f"Week{slip_week}_AllDays_PalletLines_*.xlsx"
        files = sorted(exports_dir.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
        if not files:
            print(f"  Slip week {slip_week}: no file found")
            continue
        path = files[0]
        print(f"\n  Slip week {slip_week}: {path.name}")
        try:
            df = pd.read_excel(path, sheet_name="PalletLines")
        except Exception as e:
            print(f"    Error reading: {e}")
            continue
        if "QtyBoxes" not in df.columns:
            print("    No QtyBoxes column")
            continue
        if "BBDate" not in df.columns:
            for _, row in df.iterrows():
                raw_total += _case_equivalent_boxes(_safe_float(row.get("QtyBoxes"), 0), row.get("SKU"))
            print(f"    No BBDate column - raw case-equiv total = {raw_total} (Option A would fall back to this)")
            continue
        file_raw = 0
        for _, row in df.iterrows():
            qty = _case_equivalent_boxes(_safe_float(row.get("QtyBoxes"), 0), row.get("SKU"))
            raw_total += qty
            file_raw += qty
            prod = _bb_date_to_production_iso_week(row.get("BBDate"))
            if prod is None:
                rows_no_bb += 1
                qty_no_bb += qty
                continue
            rows_with_bb += 1
            key = prod
            by_production_week[key] = by_production_week.get(key, 0) + qty
        print(f"    Rows: {len(df)}, Raw case-equiv: {file_raw:,} (30-pack Nova/Loblaws = 4/3 case-equiv/box; Jumbo = 12/15)")
        print(f"    Rows with valid BB date: {rows_with_bb}, without: {rows_no_bb} (qty {qty_no_bb})")
    print(f"\n  TOTAL RAW (all PalletLines for slip weeks {week_number} & {week_number + 1}): {raw_total}")
    if by_production_week:
        print(f"\n  QtyBoxes by production week (BB-47 -> ISO year, week):")
        for (iy, iw), qty in sorted(by_production_week.items(), key=lambda x: (-x[1], x[0])):
            rw = _production_iso_to_report_week(iy, iw)
            mark = " <-- requested week" if rw == week_number else ""
            print(f"    ISO {iy}-W{iw} (report {rw}): {qty:,} boxes{mark}")
        option_a_sum = sum(q for (iy, iw), q in by_production_week.items() if _production_iso_to_report_week(iy, iw) == week_number)
        print(f"\n  Option A sum (prod_week == {week_number}): {option_a_sum}")
        if raw_total > 0:
            pct = 100 * option_a_sum / raw_total
            print(f"  Option A is {pct:.1f}% of raw total.")
            if pct < 20:
                print("\n  Likely cause: week number mismatch (slip 'Week N' vs ISO week N) or BB dates map to other weeks.")
    print()


def _run_barn_zero_report(exports_dir=None):
    """
    Report weeks where any barn has zero production (flock transition), plus 1–2 weeks before.
    Shows traceability % for each. Use Eggs_Stored in SQF_Traceability_Inputs.csv for those weeks.
    """
    if exports_dir is None:
        exports_dir = EXPORTS_DIR
    exports_dir = Path(exports_dir)
    try:
        import hilly_acres_production as ha
    except ImportError:
        print("Could not import hilly_acres_production.")
        return
    df = load_sqf_inputs()
    weeks_in_inputs = set()
    if df is not None and not df.empty and "Week" in df.columns:
        for _, row in df.iterrows():
            w = _get_row(row, "Week")
            if w is not None and not (isinstance(w, float) and pd.isna(w)):
                try:
                    weeks_in_inputs.add(int(float(w)))
                except (ValueError, TypeError):
                    pass
    # Scan weeks 1–52 (and 56–61 if in inputs) for barns with zero
    weeks_to_scan = sorted(set(range(1, 53)) | (weeks_in_inputs & {56, 57, 58, 59, 60, 61}))
    weeks_with_zero_barns = set()
    barn_zero_detail = {}  # week -> [barn numbers with 0]
    for week_num in weeks_to_scan:
        barn_stacks = ha.get_barn_stacks_for_week(week_num)
        if barn_stacks is None:
            continue
        zero_barns = [b for b, s in barn_stacks.items() if s == 0 or (isinstance(s, float) and s < 0.01)]
        if zero_barns:
            weeks_with_zero_barns.add(week_num)
            barn_zero_detail[week_num] = sorted(zero_barns)
    # Add 1–2 weeks before each zero week (leading up to transition)
    weeks_to_report = set(weeks_with_zero_barns)
    for w in weeks_with_zero_barns:
        if w >= 2:
            weeks_to_report.add(w - 1)
        if w >= 3:
            weeks_to_report.add(w - 2)
    weeks_to_report = sorted(weeks_to_report)
    if not weeks_to_report:
        print("No weeks found with barns at zero production.")
        return
    # Run traceability for those weeks
    results = run_all_weeks(exports_dir, by_slip_week=False)
    by_week = {r["Week"]: r for r in results} if results else {}
    print("=" * 75)
    print("BARN ZERO REPORT — Weeks with barn(s) at zero production (flock transition)")
    print("=" * 75)
    print("Weeks where any barn has 0 production may have stored eggs; use Eggs_Stored column.")
    print()
    print(f"{'Week':>5}  {'Barns@0':<18}  {'Traceability%':>14}  {'Production':>12}  {'Accounted':>12}  {'Variance':>10}")
    print("-" * 85)
    for week_num in weeks_to_report:
        r = by_week.get(week_num, {})
        barns_str = (", ".join(f"Barn {b}" for b in barn_zero_detail.get(week_num, [])) or "—")[:18]
        pct = r.get("Traceability_Pct")
        pct_str = f"{pct}%" if pct is not None else "N/A"
        prod = int(r.get("Production") or 0)
        acc = int(r.get("Accounted") or 0)
        var = int(r.get("Variance") or 0)
        print(f"{week_num:>5}  {barns_str:<18}  {pct_str:>14}  {prod:>12,}  {acc:>12,}  {var:>10,}")
    print("-" * 85)
    print("Double-check barn zeros in Hilly Acres Inputs sheets. Add Eggs_Stored for transition weeks.")
    print()


def _deep_dive_week(week_number, exports_dir=None):
    """
    Deep dive for one report week: show where every number comes from, date alignment,
    and orders broken down by production week (BB date) so we can see if week alignment explains variance.
    """
    if exports_dir is None:
        exports_dir = EXPORTS_DIR
    exports_dir = Path(exports_dir)
    df_in = load_sqf_inputs()
    row = None
    if df_in is not None and not df_in.empty and "Week" in df_in.columns:
        for _, r in df_in.iterrows():
            w = _get_row(r, "Week")
            if w is not None and not (isinstance(w, float) and pd.isna(w)):
                try:
                    if int(float(w)) == week_number:
                        row = r.to_dict() if hasattr(r, "to_dict") else r
                        break
                except (ValueError, TypeError):
                    pass

    print("=" * 70)
    print(f"DEEP DIVE — Report Week {week_number}")
    print("=" * 70)

    # 1) Report week date range (what calendar period is "week 48"?)
    mon, sun = _report_week_to_dates(week_number)
    if mon and sun:
        print(f"\n1) REPORT WEEK DATE RANGE (ISO): {mon.strftime('%Y-%m-%d')} (Mon) — {sun.strftime('%Y-%m-%d')} (Sun)")
    else:
        print(f"\n1) REPORT WEEK DATE RANGE: (could not compute for week {week_number})")

    # 2) Production: value and source
    production = get_production_for_report_week(week_number, row or {}, exports_dir, slip_week_override=_get_row(row, "SlipWeek", None) if row is not None else None)
    prod_src, _ = _production_source(week_number, row if row is not None else {})
    print(f"\n2) PRODUCTION: {production:,.0f} cases")
    print(f"   Source: {prod_src}")
    try:
        import hilly_acres_production as ha
        diag = ha.diagnose_week_file_finding(week_number)
        if diag.get("file_found"):
            print(f"   Hilly Acres file: {Path(diag['file_found']).name}")
        else:
            print("   Hilly Acres file: NOT FOUND (check paths and file names)")
    except Exception as e:
        print(f"   Hilly Acres file check: {e}")

    # 3) Slip week 48: what week-ending does "Week 48" loading slip represent?
    try:
        sw = _get_row(row, "SlipWeek", week_number) if row is not None else week_number
        slip_week = int(sw) if sw is not None and str(sw).strip() and not (isinstance(sw, float) and pd.isna(sw)) else week_number
    except (ValueError, TypeError):
        slip_week = week_number
    try:
        from process_weekly_loading_slip import get_slip_week_ending
        we = get_slip_week_ending(slip_week)
        if we:
            print(f"\n3) LOADING SLIP 'Week {slip_week}' → Week ending: {we} (Saturday, day after Friday ship)")
        else:
            print(f"\n3) LOADING SLIP 'Week {slip_week}': week-ending date not found (no Friday ship date in repo or slip file)")
    except Exception as e:
        print(f"\n3) LOADING SLIP week ending: {e}")

    # 4) Orders: slip-week total + breakdown by production week (BB date)
    pattern = f"Week{slip_week}_AllDays_PalletLines_*.xlsx"
    files = list(exports_dir.glob(pattern)) if exports_dir.exists() else []
    slip_week_total = 0
    by_prod_week = {}  # (iy, iw) -> case-equiv (filled when we have PalletLines with BBDate)
    rows_no_bb = 0
    qty_no_bb = 0
    if files:
        path = sorted(files, key=lambda p: p.stat().st_mtime, reverse=True)[0]
        try:
            df = pd.read_excel(path, sheet_name="PalletLines")
        except Exception as e:
            print(f"\n4) SHIPPED_ORDERS (PalletLines): Error reading {path.name}: {e}")
            df = None
        if df is not None and "QtyBoxes" in df.columns:
            for _, r in df.iterrows():
                qty = _case_equivalent_boxes(_safe_float(r.get("QtyBoxes"), 0), r.get("SKU"))
                slip_week_total += qty
                if "BBDate" not in df.columns:
                    continue
                prod = _bb_date_to_production_iso_week(r.get("BBDate"))
                if prod is None:
                    rows_no_bb += 1
                    qty_no_bb += qty
                    continue
                by_prod_week[prod] = by_prod_week.get(prod, 0) + qty
            print(f"\n4) SHIPPED_ORDERS (from Week {slip_week} PalletLines): {slip_week_total:,} case-equiv total")
            if "BBDate" in df.columns:
                option_a = sum(q for (iy, iw), q in by_prod_week.items() if _production_iso_to_report_week(iy, iw) == week_number)
                print(f"   Option A (only rows where BB date → production week {week_number}): {option_a:,}")
                print(f"   Rows with no/invalid BB date: {rows_no_bb} (qty {qty_no_bb:,})")
                if by_prod_week:
                    print("   Breakdown by production week (BB date - 47 days → pack week):")
                    for (iy, iw), q in sorted(by_prod_week.items(), key=lambda x: -x[1]):
                        rw = _production_iso_to_report_week(iy, iw)
                        mark = " <-- report week" if rw == week_number else ""
                        print(f"      {iy}-W{iw} (report {rw}): {q:,} cases{mark}")
                dominant = get_dominant_production_week_from_slip(slip_week, exports_dir)
                if dominant:
                    dy, dw = dominant
                    dom_report = _production_iso_to_report_week(dy, dw)
                    print(f"   Dominant production week on this slip: {dy}-W{dw} (report {dom_report})")
                    if dom_report != week_number:
                        print(f"   >>> MISMATCH: Slip 'Week {slip_week}' orders mostly produced in report week {dom_report}. Consider comparing production week {dom_report} to this slip, or use SlipWeek/HA week alignment.")
                # Cross-check: how much of *report week* production appears on adjacent slips?
                for other_week in (week_number - 1, week_number + 1):
                    if other_week < 1:
                        continue
                    opat = f"Week{other_week}_AllDays_PalletLines_*.xlsx"
                    ofiles = list(exports_dir.glob(opat))
                    if not ofiles:
                        continue
                    try:
                        odf = pd.read_excel(sorted(ofiles, key=lambda p: p.stat().st_mtime, reverse=True)[0], sheet_name="PalletLines")
                    except Exception:
                        continue
                    if "QtyBoxes" not in odf.columns or "BBDate" not in odf.columns:
                        continue
                    q_on_other = 0
                    for _, r in odf.iterrows():
                        prod_key = _bb_date_to_production_iso_week(r.get("BBDate"))
                        if prod_key and _production_iso_to_report_week(prod_key[0], prod_key[1]) == week_number:
                            q_on_other += _case_equivalent_boxes(_safe_float(r.get("QtyBoxes"), 0), r.get("SKU"))
                    if q_on_other:
                        print(f"   On Week {other_week} slip: {q_on_other:,} cases produced in W{week_number} (eggs packed W{week_number}, shipped next week)")
        else:
            print(f"\n4) SHIPPED_ORDERS: No PalletLines file or no QtyBoxes column for Week {slip_week}")
    else:
        print(f"\n4) SHIPPED_ORDERS: No Week{slip_week}_AllDays_PalletLines_*.xlsx in {exports_dir}")

    # 5) Nest run
    nest_run = get_nest_run_for_report_week(week_number, row or {}, slip_week_override=_get_row(row, "SlipWeek", None) if row is not None else None)
    print(f"\n5) SHIPPED_NESTRUN: {nest_run:,.0f} cases")

    # 6) Off-grades
    off_csv = _num(_get_row(row, "OffGrades_GradeOut"), default=None) if row is not None else None
    off_grades = get_off_grades_from_gradeout(week_number) if (off_csv is None or off_csv == 0) else off_csv
    if off_grades is None:
        off_grades = off_csv if off_csv is not None else 0
    if off_grades == 0 and production and production > 0:
        off_grades = round(production * DEFAULT_OFFGRADE_PCT / 100.0, 2)
        print(f"\n6) OFFGRADES_GRADEOUT: {off_grades:,.0f} (default {DEFAULT_OFFGRADE_PCT}% of Production — no gradeout data for week)")
    else:
        print(f"\n6) OFFGRADES_GRADEOUT: {off_grades:,.0f}")

    # 7) Summary
    accounted = slip_week_total + nest_run + off_grades
    variance = (production or 0) - accounted
    print(f"\n7) SUMMARY")
    print(f"   Production:     {production or 0:,.0f}")
    print(f"   Accounted:      {accounted:,.0f}  (Orders + NestRun + OffGrades)")
    print(f"   Variance:       {variance:,.0f}")
    if production and production > 0:
        pct = 100.0 * accounted / production
        print(f"   Traceability:   {pct:.1f}%")

    # 8) Possible causes of large variance
    if by_prod_week and production and variance > 500:
        option_a_orders = sum(q for (iy, iw), q in by_prod_week.items() if _production_iso_to_report_week(iy, iw) == week_number)
        total_w48_on_any_slip = option_a_orders  # we only have Week 48 slip in by_prod_week; adjacent slip W48 qty was printed above
        print(f"\n8) POSSIBLE CAUSES OF VARIANCE ({variance:,.0f} cases)")
        print("   - Slip/week alignment: Part of Week 48 slip is W47 production; part of W48 production ships on Week 49 slip (see 'On Week N slip' above).")
        print("   - If you use Option A (BB date -> prod week): orders for this week would be {0:,} not {1:,}.".format(option_a_orders, slip_week_total))
        print("   - Missing orders: Are there other loading slips for this week (e.g. multiple files, NFLD, or slips not yet run through process_weekly_loading_slip)?")
        print("   - Nest run: Only 116 cases from Hilly Acres Pallet Info; is more nest run filled elsewhere (CSV)?")
        print("   - Production unit: Hilly Acres Inputs TOTAL (stacks) = 1 stack 1 case; confirm this matches your definition.")
    print("\n" + "=" * 70)


def _sku_prior_week_pct(week_number, exports_dir=None):
    """
    For a given slip week (e.g. 59), show per-SKU what % of boxes used the prior week BB date
    vs the current week BB date. Prior = earliest BB date in file, current = latest.
    """
    if exports_dir is None:
        exports_dir = EXPORTS_DIR
    exports_dir = Path(exports_dir)
    if not exports_dir.exists():
        print(f"Exports dir not found: {exports_dir}")
        return
    pattern = f"Week{week_number}_AllDays_PalletLines_*.xlsx"
    files = sorted(exports_dir.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        print(f"No PalletLines file found for week {week_number}")
        return
    try:
        df = pd.read_excel(files[0], sheet_name="PalletLines")
    except Exception as e:
        print(f"Error reading {files[0].name}: {e}")
        return
    if "QtyBoxes" not in df.columns or "BBDate" not in df.columns or "SKU" not in df.columns:
        print("PalletLines must have columns: SKU, QtyBoxes, BBDate")
        return
    # NFLD: Consolidated PalletLines are built from Mon-Fri (+ NFLD if processor includes it).
    if "DayName" in df.columns:
        days_in_file = df["DayName"].dropna().astype(str).str.strip().unique()
        nfld_mask = df["DayName"].astype(str).str.strip().str.upper().str.contains("NFLD", na=False)
        nfld_rows = nfld_mask.sum()
        if nfld_rows:
            nfld_boxes = sum(_case_equivalent_boxes(_safe_float(row.get("QtyBoxes"), 0), row.get("SKU")) for _, row in df.loc[nfld_mask].iterrows())
            print(f"  Note: NFLD rows in this file: {int(nfld_rows)} rows, {int(nfld_boxes):,} case-equiv boxes.")
        else:
            print("  Note: This file has no NFLD day. NFLD orders are only included if the loading slip has an NFLD sheet and the weekly processor is run with NFLD support.")
    # (SKU, bb_date) -> sum of QtyBoxes
    sku_date_qty = {}
    dates_seen = set()
    for _, row in df.iterrows():
        sku = str(row.get("SKU") or "").strip() or "(blank)"
        qty = int(_safe_float(row.get("QtyBoxes"), 0))
        if qty <= 0:
            continue
        d = _parse_bb_date(row.get("BBDate"))
        if d is None:
            continue
        dates_seen.add(d)
        key = (sku, d)
        sku_date_qty[key] = sku_date_qty.get(key, 0) + qty
    if not dates_seen:
        print("No valid BB dates found in file.")
        return
    sorted_dates = sorted(dates_seen)
    prior_date = sorted_dates[0]
    current_date = sorted_dates[-1] if len(sorted_dates) > 1 else prior_date
    # Per SKU: total, prior_boxes, current_boxes, pct_prior
    sku_totals = {}
    for (sku, d), qty in sku_date_qty.items():
        if sku not in sku_totals:
            sku_totals[sku] = {"total": 0, "prior": 0, "current": 0, "mid": 0}
        sku_totals[sku]["total"] += qty
        if d == prior_date:
            sku_totals[sku]["prior"] += qty
        elif d == current_date:
            sku_totals[sku]["current"] += qty
        else:
            sku_totals[sku]["mid"] += qty
    mid_dates = sorted_dates[1:-1] if len(sorted_dates) > 2 else []
    print("=" * 72)
    print(f"WEEK {week_number} - % of each SKU by BB date (prior / mid / current)")
    print("=" * 72)
    print(f"  Prior ({prior_date}) = Mon/Tue OD (prior week)")
    print(f"  Current ({current_date}) = Wed-Fri (current week)")
    if mid_dates:
        print(f"  Mid {[str(d) for d in mid_dates]} = Mon/Tue non-OD or Wed-Fri OD")
    print()
    print(f"{'SKU':<35} {'Total':>8} {'Prior':>8} {'Current':>8} {'Mid':>8} {'% Prior':>8}")
    print("-" * 72)
    for sku in sorted(sku_totals.keys(), key=lambda s: (-sku_totals[s]["total"], s)):
        r = sku_totals[sku]
        total = r["total"]
        pct_prior = (100.0 * r["prior"] / total) if total else 0
        print(f"{sku[:34]:<35} {total:>8,} {r['prior']:>8,} {r['current']:>8,} {r['mid']:>8,} {pct_prior:>7.1f}%")
    print("-" * 72)
    total_all = sum(r["total"] for r in sku_totals.values())
    prior_all = sum(r["prior"] for r in sku_totals.values())
    pct_all = (100.0 * prior_all / total_all) if total_all else 0
    print(f"{'TOTAL':<35} {total_all:>8,} {prior_all:>8,} {sum(r['current'] for r in sku_totals.values()):>8,} {sum(r['mid'] for r in sku_totals.values()):>8,} {pct_all:>7.1f}%")
    print()
    # Optionally write CSV to exports
    out_csv = exports_dir / f"Week{week_number}_SKU_PriorWeekPct.csv"
    rows = [{"SKU": sku, "Total_Boxes": r["total"], "Prior_Week_Boxes": r["prior"], "Current_Week_Boxes": r["current"], "Mid_Boxes": r["mid"], "Pct_Prior_Week": round(100.0 * r["prior"] / r["total"], 1) if r["total"] else 0} for sku, r in sorted(sku_totals.items(), key=lambda x: -x[1]["total"])]
    pd.DataFrame(rows).to_csv(out_csv, index=False)
    print(f"Saved: {out_csv}")
    print()


def _check_nfld_in_palletlines(week_number, exports_dir=None):
    """
    Report whether the Week N PalletLines file includes NFLD orders.
    Prints: file path, NFLD rows count, NFLD case-equiv boxes, and a clear Yes/No.
    """
    if exports_dir is None:
        exports_dir = EXPORTS_DIR
    exports_dir = Path(exports_dir)
    if not exports_dir.exists():
        print(f"Exports dir not found: {exports_dir}")
        return
    df, path = _load_palletlines_df(week_number, exports_dir, source_mode=PALLETLINES_SOURCE_PREFERRED)
    if df is None or path is None:
        print(f"No PalletLines file found for week {week_number} in {exports_dir}")
        print("So NFLD is not being picked up for this week (no file to read).")
        return
    if "DayName" not in df.columns:
        print(f"File: {path.name}")
        print("This file has no 'DayName' column, so we cannot tell NFLD from other days.")
        print("NFLD status: unknown (no DayName).")
        return
    nfld_mask = df["DayName"].astype(str).str.strip().str.upper().str.contains("NFLD", na=False)
    nfld_rows = int(nfld_mask.sum())
    nfld_boxes = sum(_case_equivalent_boxes(_safe_float(row.get("QtyBoxes"), 0), row.get("SKU")) for _, row in df.loc[nfld_mask].iterrows())
    print(f"File: {path.name}")
    if nfld_rows:
        print(f"NFLD in PalletLines: YES — {nfld_rows} rows, {int(nfld_boxes):,} case-equiv boxes.")
    else:
        print("NFLD in PalletLines: NO — no rows with DayName = NFLD.")
        print("To include NFLD: re-run the weekly processor with a loading slip that has an NFLD sheet, and ensure the mapping has Day=NFLD for NFLD stops.")


def _diagnose_nest_run(week_number, exports_dir=None):
    """
    For a single report week: show where nest run comes from (CSV vs Hilly Acres),
    which Hilly Acres file/sheet/rows were used, and the final value.
    """
    exports_dir = Path(exports_dir) if exports_dir else EXPORTS_DIR
    df = load_sqf_inputs()
    row = None
    if df is not None and not df.empty and "Week" in df.columns:
        for _, r in df.iterrows():
            w = _get_row(r, "Week")
            if w is not None and not (isinstance(w, float) and pd.isna(w)):
                wn = int(float(w)) if isinstance(w, (int, float)) else None
                if wn == week_number:
                    row = r
                    break
    nest_csv = _num(_get_row(row, "Shipped_NestRun"), default=None) if row is not None else None
    ha_week_override = _get_hilly_acres_week_override(row) if row is not None else None
    slip_override = _get_row(row, "SlipWeek", None) if row is not None else None
    if slip_override is not None and isinstance(slip_override, float) and pd.isna(slip_override):
        slip_override = None
    slip_week = int(slip_override) if slip_override is not None and str(slip_override).strip() else week_number

    print("=" * 70)
    print(f"NEST RUN DIAGNOSTIC — Report Week {week_number}")
    print("=" * 70)
    print(f"  CSV Shipped_NestRun: {nest_csv if nest_csv is not None else '(blank)'}")
    print(f"  CSV HillyAcresWeek:  {ha_week_override if ha_week_override is not None else '(not set)'}")
    print(f"  Slip week (for week-ending): {slip_week}")
    print()

    try:
        import hilly_acres_production as ha
    except ImportError:
        print("  Could not import hilly_acres_production.")
        return

    used_value = None
    source = None

    if nest_csv is not None and nest_csv > 0:
        used_value = nest_csv
        source = "CSV (Shipped_NestRun)"
        print(f"  -> Using CSV: {used_value:,} boxes (source: {source})")
        print()
        return

    # Try Hilly Acres: first by HillyAcresWeek override
    if ha_week_override is not None:
        boxes, file_path, diag = ha.get_nest_run_boxes_for_week_diagnostic(ha_week_override)
        if boxes is not None and boxes > 0:
            used_value = boxes
            source = f"Hilly Acres week {ha_week_override} (HillyAcresWeek override)"
        print(f"  Hilly Acres week {ha_week_override} (override):")
        if file_path is None:
            print("    No workbook found for that week.")
        else:
            print(f"    File: {file_path}")
            if diag:
                if diag.get("layout") == "row_count":
                    print(f"    Layout: row-count (data from row 7, 60 units/pallet, stop when col K blank)")
                    print(f"    Rows counted: {diag.get('rows_counted', 0)}, raw total: {diag.get('raw_total', 0):,}")
                elif diag.get("layout") == "line_item":
                    print(f"    Layout: line-item (NR column + quantity column)")
                    print(f"    Header row: {diag.get('header_row')}, NR col: {diag.get('nr_col')}, Qty col: {diag.get('qty_col')}, rows summed: {diag.get('rows_included', 0)}")
                else:
                    print(f"    Layout: totals (NR Grid Size + NR Total)")
                    print(f"    Header row: {diag.get('header_row')}, NR Grid col: {diag.get('nr_grid_col')}, NR Total col: {diag.get('nr_total_col')}")
                    print(f"    Grand total row used: {diag.get('grand_total_used')}")
                    for rec in diag.get("rows_included", []):
                        print(f"      Row {rec['row']}: \"{rec['label']}\" -> {rec['value']:,} ({rec['type']})")
                if boxes is not None:
                    print(f"    Nest run total: {boxes:,}")
            else:
                print(f"    Nest run total: {boxes or 0:,}")
        print()

    if used_value is None and slip_week is not None:
        try:
            from process_weekly_loading_slip import get_slip_week_ending
            we = get_slip_week_ending(slip_week)
        except Exception:
            we = None
        if we is not None:
            boxes, file_path, diag = ha.get_nest_run_boxes_for_week_ending_diagnostic(we)
            if boxes is not None and boxes > 0:
                used_value = boxes
                source = f"Week ending {we} (from slip week {slip_week})"
            print(f"  By week ending (slip week {slip_week} -> {we}):")
            if file_path is None:
                print("    No workbook found with that Week Ending.")
            else:
                print(f"    File: {file_path}")
                if diag:
                    if diag.get("layout") == "row_count":
                        print(f"    Layout: row-count; rows counted: {diag.get('rows_counted', 0)}, raw: {diag.get('raw_total', 0):,}")
                    elif diag.get("layout") == "line_item":
                        print(f"    Layout: line-item; NR col {diag.get('nr_col')}, Qty col {diag.get('qty_col')}; rows summed: {diag.get('rows_included', 0)}")
                    else:
                        for rec in diag.get("rows_included", []):
                            print(f"      Row {rec['row']}: \"{rec['label']}\" -> {rec['value']:,}")
                    print(f"    Nest run total: {boxes or 0:,}")
            print()

    if used_value is None:
        boxes, file_path, diag = ha.get_nest_run_boxes_for_week_diagnostic(week_number)
        if boxes is not None and boxes > 0:
            used_value = boxes
            source = f"Hilly Acres week {week_number} (report week)"
        print(f"  By report week ({week_number}):")
        if file_path is None:
            print("    No workbook found for that week number.")
            # Show why: which folders and what files exist
            ff = ha.diagnose_week_file_finding(week_number)
            if ff.get("per_folder"):
                print("    File-finding detail:")
                for pf in ff["per_folder"]:
                    print(f"      Folder: {pf['path']}")
                    print(f"        Exists: {pf['exists']}, xlsx count: {pf['file_count']}")
                    if pf.get("file_names"):
                        names = pf["file_names"]
                        if len(names) <= 15:
                            print(f"        Files: {', '.join(names)}")
                        else:
                            print(f"        Files (first 15): {', '.join(names[:15])} ...")
                        # Highlight any file that might be this week (e.g. contains "45")
                        week_str = str(week_number)
                        likely = [n for n in names if week_str in n.replace(" ", "")]
                        if likely:
                            print(f"        Files with week {week_number} in name: {likely}")
        else:
            print(f"    File: {file_path}")
            if diag:
                if diag.get("layout") == "row_count":
                    print(f"    Layout: row-count; rows counted: {diag.get('rows_counted', 0)}, raw: {diag.get('raw_total', 0):,}")
                elif diag.get("layout") == "line_item":
                    print(f"    Layout: line-item; NR col {diag.get('nr_col')}, Qty col {diag.get('qty_col')}; rows summed: {diag.get('rows_included', 0)}")
                else:
                    for rec in diag.get("rows_included", []):
                        print(f"      Row {rec['row']}: \"{rec['label']}\" -> {rec['value']:,}")
                print(f"    Nest run total: {boxes or 0:,}")
            elif (boxes is None or boxes == 0) and file_path:
                # File found but no nest run: show sheet structure so we can see what's different
                sheet_d = ha.diagnose_pallet_sheet_for_week(week_number)
                if sheet_d.get("sheet_exists"):
                    print("    Sheet structure (why parsing may have failed):")
                    print(f"      Shape: {sheet_d.get('shape')} rows x cols")
                    print(f"      Row 0 (first 15 cols): {sheet_d.get('header_row_0', [])[:15]}")
                    if sheet_d.get("header_row_1"):
                        print(f"      Row 1 (first 15 cols): {sheet_d.get('header_row_1', [])[:15]}")
                    print(f"      Totals header (NR Grid Size + NR Total) found: {sheet_d.get('totals_header_found')}")
                    print(f"      Line-item header (NR + GROSS/QUAN) found: {sheet_d.get('line_item_header_found')}")
                    print(f"      Parser result: {sheet_d.get('nest_run_result')} (layout: {sheet_d.get('parser_layout')})")
        print()

    print("  RESULT:")
    if used_value is not None:
        print(f"    NestRun_Used = {used_value:,}  (source: {source})")
    else:
        print("    NestRun_Used = 0  (no CSV value and no Hilly Acres data found)")
    print()


def _compare_nest_run_weeks(week_a, week_b):
    """
    Compare two weeks (e.g. 45 vs 42) to see why one gets nest run and the other doesn't.
    Prints file-finding and Pallet Information sheet structure for both.
    """
    try:
        import hilly_acres_production as ha
    except ImportError:
        print("Could not import hilly_acres_production.")
        return
    print("=" * 70)
    print(f"NEST RUN COMPARISON — Week {week_a} (failing) vs Week {week_b} (working)")
    print("=" * 70)
    for label, wn in [("Week " + str(week_a), week_a), ("Week " + str(week_b), week_b)]:
        print(f"\n--- {label} ---")
        ff = ha.diagnose_week_file_finding(wn)
        print(f"  File found: {ff.get('file_found') or 'NO'}")
        for pf in ff.get("per_folder", []):
            print(f"  Folder: {pf['path']} (exists={pf['exists']}, xlsx count={pf['file_count']})")
            if pf.get("chosen"):
                print(f"    -> Chosen file: {pf['chosen']}")
            elif pf.get("file_names") and len(pf["file_names"]) <= 20:
                print(f"    Files: {', '.join(pf['file_names'])}")
        sheet = ha.diagnose_pallet_sheet_for_week(wn)
        if sheet.get("file_path"):
            print(f"  Pallet Information: shape={sheet.get('shape')}")
            print(f"    Row 0 headers (first 12): {sheet.get('header_row_0', [])[:12]}")
            if sheet.get("header_row_1"):
                print(f"    Row 1 headers (first 12): {sheet.get('header_row_1', [])[:12]}")
            print(f"    Totals layout (NR Grid Size + NR Total): {sheet.get('totals_header_found')}")
            print(f"    Line-item layout (NR + GROSS/QUAN): {sheet.get('line_item_header_found')}")
            print(f"    Nest run parsed: {sheet.get('nest_run_result')} (layout used: {sheet.get('parser_layout')})")
        else:
            print("  Pallet Information: no workbook, so no sheet.")
    print("\n" + "=" * 70)
    print("If Week A has 'File found: NO', the workbook is missing or not matched (check folder path and file names).")
    print("If both have a file but only one has nest_run_result: compare Row 0/1 headers and Totals/Line-item results above.")
    print("=" * 70)


def _year_for_report_week(week_num):
    """Return 2026 if report week is in 2026 range (56-61), else 2025."""
    try:
        with open(REF_DATA_DIR / "paths.json", encoding="utf-8") as f:
            config = json.load(f)
        report_2026 = config.get("ReportWeek2026Range") or [56, 57, 58, 59, 60, 61]
    except Exception:
        report_2026 = [56, 57, 58, 59, 60, 61]
    return 2026 if week_num in report_2026 else 2025


def _report_week_to_dates(week_num):
    """
    Return (week_start_date, week_ending_date) for a report week for display.
    Report weeks 56-61 -> 2026 ISO weeks 1-6 (56 = week 1); weeks 1-52 -> 2025 ISO week N. Weeks 53-55 do not exist.
    week_ending = Sunday of that week.
    """
    try:
        with open(REF_DATA_DIR / "paths.json", encoding="utf-8") as f:
            config = json.load(f)
        report_2026 = config.get("ReportWeek2026Range") or [56, 57, 58, 59, 60, 61]
        cal_to_iso = config.get("LoadingSlip2026CalendarToISO") or {}
        iso_to_cal = {int(v): int(k) for k, v in cal_to_iso.items()}
    except Exception:
        report_2026 = [56, 57, 58, 59, 60, 61]
        iso_to_cal = {56: 1, 57: 2, 58: 3, 59: 4, 60: 5, 61: 6}
    if week_num in report_2026:
        iso_year = 2026
        iso_week = iso_to_cal.get(week_num, week_num)
    else:
        iso_year = 2025
        iso_week = week_num
    try:
        from datetime import date
        mon = date.fromisocalendar(iso_year, iso_week, 1)
        sun = mon + timedelta(days=6)
        return (mon, sun)
    except ValueError:
        return (None, None)


def _export_nest_run_table(exports_dir=None):
    """
    Export CSV: Week, Week_Ending, HillyAcresWeek_Used, NestRun_From_HA, NestRun_From_CSV, NestRun_Used, Source.
    For each week in SQF_Traceability_Inputs.csv, compute nest run and record where it came from.
    Week_Ending shows the Sunday date for that week to help match 2025 vs 2026.
    """
    exports_dir = Path(exports_dir) if exports_dir else EXPORTS_DIR
    exports_dir.mkdir(parents=True, exist_ok=True)
    df = load_sqf_inputs()
    if df is None or df.empty or "Week" not in df.columns:
        print("No SQF_Traceability_Inputs.csv or no Week column.")
        return None
    try:
        import hilly_acres_production as ha
    except ImportError:
        print("Could not import hilly_acres_production.")
        return None
    try:
        from process_weekly_loading_slip import get_slip_week_ending
    except ImportError:
        get_slip_week_ending = None

    rows = []
    for _, row in df.iterrows():
        w = _get_row(row, "Week")
        if w is None or (isinstance(w, float) and pd.isna(w)):
            continue
        week_num = int(float(w)) if isinstance(w, (int, float)) else None
        if week_num is None:
            continue
        nest_csv = _num(_get_row(row, "Shipped_NestRun"), default=None)
        ha_week_col = _get_hilly_acres_week_override(row)
        slip_override = _get_row(row, "SlipWeek", None)
        if slip_override is not None and isinstance(slip_override, float) and pd.isna(slip_override):
            slip_override = None
        slip_week = int(slip_override) if slip_override is not None and str(slip_override).strip() else week_num

        used = get_nest_run_for_report_week(week_num, row, slip_week_override=slip_override)
        nest_ha = None
        ha_week_used = None
        source = "CSV"
        if nest_csv is not None and nest_csv > 0:
            source = "CSV"
        else:
            if ha_week_col is not None:
                nest_ha = ha.get_nest_run_boxes_for_week(ha_week_col)
                if nest_ha is not None:
                    ha_week_used = ha_week_col
                    source = "HillyAcresWeek"
            if nest_ha is None and get_slip_week_ending and slip_week is not None:
                we = get_slip_week_ending(slip_week)
                if we is not None:
                    nest_ha = ha.get_nest_run_boxes_for_week_ending(we)
                    if nest_ha is not None:
                        source = "WeekEnding"
            if nest_ha is None:
                nest_ha = ha.get_nest_run_boxes_for_week(week_num)
                if nest_ha is not None:
                    ha_week_used = week_num
                    source = "ReportWeek"
        mon, sun = _report_week_to_dates(week_num)
        week_ending_str = sun.strftime("%Y-%m-%d") if sun else ""
        rows.append({
            "Week": week_num,
            "Week_Ending": week_ending_str,
            "HillyAcresWeek_Used": ha_week_used if ha_week_used is not None else "",
            "NestRun_From_HA": nest_ha if nest_ha is not None else "",
            "NestRun_From_CSV": nest_csv if nest_csv is not None else "",
            "NestRun_Used": used,
            "Source": source,
        })
    out_df = pd.DataFrame(rows)
    path = exports_dir / "NestRun_By_Week.csv"
    out_df.to_csv(path, index=False)
    print(f"Nest run table exported: {path}")
    return path


def _check_weeks(year, week_numbers):
    """
    Print week alignment: ISO week date ranges for given year and week numbers.
    Use to verify Hilly Acres / loading slip week numbers match SQF (ISO).
    """
    if not week_numbers:
        week_numbers = [30, 31, 34, 59]  # sample weeks
    print("=" * 60)
    print("WEEK ALIGNMENT CHECK — ISO week date ranges")
    print("=" * 60)
    print(f"SQF Option A and grade-outs use ISO week (Mon–Sun).")
    print(f"Verify Hilly Acres 'Week N' and loading slip 'Week N' match these ranges.\n")
    for wn in week_numbers:
        try:
            from datetime import date
            d = date.fromisocalendar(year, wn, 1)  # Monday
            mon = d.strftime("%Y-%m-%d (%a)")
            sun = (d + timedelta(days=6)).strftime("%Y-%m-%d (%a)")
            print(f"Week {wn} ({year}):  Mon {mon}  —  Sun {sun}")
        except ValueError as e:
            print(f"Week {wn}: not valid ISO week (ISO has 1–53) — {e}")
        except Exception as e:
            print(f"Week {wn}: error — {e}")
    print("\nIf Hilly Acres uses different numbering, update Documentation/WEEK_ALIGNMENT.md")


def main():
    ap = argparse.ArgumentParser(
        description="SQF Traceability: Production vs Orders + Nest Run + Off Grades (99%+ target)."
    )
    ap.add_argument("--week", "-w", type=int, default=None, help="Single week to report")
    ap.add_argument("--output-dir", "-o", type=str, default=None, help="Folder with PalletLines exports")
    ap.add_argument("--export", "-e", action="store_true", help="Write Excel report to Traceability_Exports")
    ap.add_argument("--check-weeks", nargs="+", metavar=("YEAR", "WEEK..."),
                    help="Verify week alignment: show ISO date ranges. e.g. --check-weeks 2025 30 31 34")
    ap.add_argument("--diagnose", type=int, metavar="WEEK",
                    help="Diagnose Shipped_Orders: raw vs Option A sum and BB date -> production week breakdown")
    ap.add_argument("--sku-bb-breakdown", type=int, metavar="WEEK",
                    help="Per-SKU %% using prior week vs current week BB date (e.g. 59). Writes WeekN_SKU_PriorWeekPct.csv")
    ap.add_argument("--check-nfld", type=int, metavar="WEEK",
                    help="Check if Week N PalletLines file includes NFLD orders (Yes/No + row and box counts)")
    ap.add_argument("--slip-week", action="store_true",
                    help="Shipped_Orders = sum all boxes from Week N PalletLines (legacy; use when slip week ne ISO week)")
    ap.add_argument("--working", action="store_true",
                    help="Export Traceability_By_Week_Working.xlsx: week-by-week Production, Orders, OffGrade_Pct, NestRun with formulas so you can edit to reach ~100%%")
    ap.add_argument("--comprehensive", "-c", action="store_true",
                    help="Export comprehensive report: glossary of terms, summary, by-week with gap to 100%%, and Path_To_100 sheet for weeks below target")
    ap.add_argument("--inputs-report", action="store_true",
                    help="Generate report: Week, Traceability%%, and missing inputs (Hilly Acres, loading slip, gradeout). Writes Excel + CSV to Traceability_Exports.")
    ap.add_argument("--diagnose-nest-run", type=int, metavar="WEEK",
                    help="Diagnose nest run for one week: show which Hilly Acres file/rows were used and final NestRun value.")
    ap.add_argument("--export-nest-run-table", action="store_true",
                    help="Export NestRun_By_Week.csv: Week, HillyAcresWeek_Used, NestRun_From_HA, NestRun_From_CSV, NestRun_Used, Source.")
    ap.add_argument("--compare-nest-run", nargs=2, type=int, metavar=("WEEK_A", "WEEK_B"),
                    help="Compare two weeks (e.g. 45 vs 42): show file-finding and sheet structure for both to spot differences.")
    ap.add_argument("--deep-dive", type=int, metavar="WEEK",
                    help="Deep dive one week: production source, slip week-ending, orders by prod week (BB date), nest run/off-grades, and variance explanation.")
    ap.add_argument("--barn-zero-report", action="store_true",
                    help="Report weeks where any barn has zero production (flock transition), plus 1–2 weeks before. Shows traceability %% for each.")
    ap.add_argument("--export-crosswalk", action="store_true",
                    help="Export audit crosswalk for 2025 week 15+ and 2026 weeks: slip file, Hilly Acres file, dates, and selected production source.")
    ap.add_argument(
        "--export-traceability-csv",
        action="store_true",
        help="Export per-week traceability CSV. Without --balance-palletlines this uses ORIGINAL PalletLines only; with --balance-palletlines it prefers adjusted PalletLines when present.",
    )
    ap.add_argument(
        "--balanced-traceability-csv",
        action="store_true",
        help="With --export-traceability-csv: export from BalancedWorking only (no new BB balancing). Use after editing sandbox Reference_Data (e.g. production ±5) while keeping prior balanced PalletLines.",
    )
    ap.add_argument("--balance-palletlines", action="store_true",
                    help="Generate adjusted PalletLines workbooks from the original files using highest-SKU balancing candidates, prioritizing NFLD.")
    ap.add_argument(
        "--balance-palletlines-aggressive",
        action="store_true",
        help="Aggressive BB-date shifting: allows moves even when the source week isn't surplus, accepting only small abs-gap worsening per move.",
    )
    ap.add_argument(
        "--balance-palletlines-direction",
        type=str,
        default="back",
        help="BBDate shift direction when building balancing candidates: `back`, `forward`, or `both` (1 week = +/-7 days).",
    )
    ap.add_argument(
        "--balance-palletlines-base",
        type=str,
        default="original",
        help="Which PalletLines set to use as the balancing baseline for candidate generation and week accounting: `original` or `balancedworking` (iterative multi-week shifting).",
    )
    ap.add_argument(
        "--allow-worsen-cases",
        type=int,
        default=2000,
        help="Aggressive-mode tolerance: max abs-gap increase allowed (in cases) for an accepted move.",
    )
    ap.add_argument(
        "--balance-soft-floor-cases",
        type=int,
        default=0,
        help="Non-aggressive balancing only: attempt to move at least this many case-equiv into each target week (if valid back candidates + source surplus exist).",
    )
    ap.add_argument(
        "--balance-hard-floor-sendback-cases",
        type=int,
        default=0,
        help="Non-aggressive balancing only: force at least this many case-equiv moved OUT from each week with back candidates (subject to candidate availability).",
    )
    ap.add_argument(
        "--adjustments-csv",
        type=str,
        default=None,
        help="Optional path to a Traceability_Adjustments-like CSV to use instead of Reference_Data/Traceability_Adjustments.csv (testing only; does not modify office reference files).",
    )
    ap.add_argument(
        "--failsafe-to-100",
        action="store_true",
        help="After reconciliation, top up weeks listed in Traceability_Failsafe_Evidence.csv to 100%% traceability (Evidence required; Evidence_File must exist if provided). Ignored during pallet balancing baseline. Or set EGGROOM_FAILSAFE_TO_100=1.",
    )
    ap.add_argument(
        "--inventory-flow-balance",
        action="store_true",
        help="Enable FIFO inventory carryover allocation across weeks and add Inventory_Flow_By_Week output (sandbox-safe modeling mode).",
    )
    ap.add_argument(
        "--inventory-flow-max-carry-weeks",
        type=int,
        default=2,
        help="With --inventory-flow-balance: maximum weeks inventory can carry before expiring from flow allocation (default 2).",
    )
    ap.add_argument(
        "--inventory-flow-apply-to-traceability",
        action="store_true",
        help="With --inventory-flow-balance: apply flow-accounted metrics to Accounted/Variance/Traceability for this run output.",
    )
    ap.add_argument("--report-near-target", action="store_true",
                    help="Report weeks within ±0.5%% of 100%% traceability from the latest balanced CSV. No balance or export run.")
    args = ap.parse_args()

    if getattr(args, "failsafe_to_100", False):
        os.environ["EGGROOM_FAILSAFE_TO_100"] = "1"

    # Allow testing with a draft adjustments CSV without touching office reference data.
    global ADJUSTMENTS_CSV_OVERRIDE
    if getattr(args, "adjustments_csv", None):
        ADJUSTMENTS_CSV_OVERRIDE = Path(args.adjustments_csv)

    if getattr(args, "report_near_target", False):
        report_weeks_near_target(exports_dir=Path(args.output_dir) if args.output_dir else EXPORTS_DIR)
        return 0
    if getattr(args, "export_crosswalk", False):
        export_week_crosswalk(exports_dir=Path(args.output_dir) if args.output_dir else EXPORTS_DIR)
        return 0
    if getattr(args, "barn_zero_report", False):
        _run_barn_zero_report(Path(args.output_dir) if args.output_dir else EXPORTS_DIR)
        return 0
    if getattr(args, "deep_dive", None) is not None:
        _deep_dive_week(args.deep_dive, Path(args.output_dir) if args.output_dir else EXPORTS_DIR)
        return 0
    if getattr(args, "compare_nest_run", None) is not None:
        wa, wb = args.compare_nest_run[0], args.compare_nest_run[1]
        _compare_nest_run_weeks(wa, wb)
        return 0
    if getattr(args, "diagnose_nest_run", None) is not None:
        _diagnose_nest_run(args.diagnose_nest_run, Path(args.output_dir) if args.output_dir else EXPORTS_DIR)
        return 0
    if getattr(args, "export_nest_run_table", False):
        _export_nest_run_table(Path(args.output_dir) if args.output_dir else EXPORTS_DIR)
        return 0
    if args.diagnose is not None:
        _diagnose_shipped_orders(args.diagnose, Path(args.output_dir) if args.output_dir else EXPORTS_DIR)
        return 0
    if getattr(args, "sku_bb_breakdown", None) is not None:
        _sku_prior_week_pct(args.sku_bb_breakdown, Path(args.output_dir) if args.output_dir else EXPORTS_DIR)
        return 0
    if getattr(args, "check_nfld", None) is not None:
        _check_nfld_in_palletlines(args.check_nfld, Path(args.output_dir) if args.output_dir else EXPORTS_DIR)
        return 0
    if getattr(args, "inputs_report", False):
        out_dir = Path(args.output_dir) if args.output_dir else EXPORTS_DIR
        report_rows = run_inputs_report(out_dir, by_slip_week=getattr(args, "slip_week", False))
        if report_rows:
            export_inputs_report(report_rows)
        else:
            print("No weeks in SQF_Traceability_Inputs.csv; nothing to report.")
        return 0
    if args.check_weeks:
        parts = [str(x).strip() for x in args.check_weeks if str(x).strip()]
        if not parts:
            print("Usage: --check-weeks YEAR WEEK [WEEK ...]  e.g. --check-weeks 2025 30 31 34")
            return 1
        year = int(parts[0])
        weeks = [int(x) for x in parts[1:]] if len(parts) > 1 else []
        _check_weeks(year, weeks)
        return 0

    # Decide default output directory for PalletLines-driven operations:
    # - explicit --output-dir wins
    # - otherwise, runs with --balance-palletlines write adjusted PalletLines
    #   into the BalancedWorking directory, while original runs use Original/legacy root.
    if args.output_dir:
        out_dir = Path(args.output_dir)
    elif getattr(args, "balance_palletlines", False) or getattr(args, "balance_palletlines_aggressive", False):
        out_dir = BALANCED_WORKING_DIR
    else:
        out_dir = ORIGINAL_EXPORTS_DIR if ORIGINAL_EXPORTS_DIR.exists() else EXPORTS_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    if getattr(args, "balance_palletlines", False) or getattr(args, "balance_palletlines_aggressive", False):
        direction = getattr(args, "balance_palletlines_direction", "back")
        base = str(getattr(args, "balance_palletlines_base", "original")).strip().lower()
        if base in ("balancedworking", "balanced_working", "balanced"):
            balance_source_dir = BALANCED_WORKING_DIR
            palletlines_source_for_balance = PALLETLINES_SOURCE_PREFERRED
        else:
            balance_source_dir = None
            palletlines_source_for_balance = PALLETLINES_SOURCE_ORIGINAL_ONLY
        balance_result = generate_balanced_palletlines(
            out_dir,
            include_heuristic=True,
            aggressive=getattr(args, "balance_palletlines_aggressive", False),
            allow_worsen_cases=getattr(args, "allow_worsen_cases", 0) if getattr(args, "balance_palletlines_aggressive", False) else 0,
            soft_floor_cases_per_week=getattr(args, "balance_soft_floor_cases", 0),
            hard_floor_sendback_cases_per_week=getattr(args, "balance_hard_floor_sendback_cases", 0),
            direction=direction,
            palletlines_source_for_balance=palletlines_source_for_balance,
            balance_source_dir=balance_source_dir,
        )
        written_files = balance_result.get("files") or []
        if written_files:
            print("\nAdjusted PalletLines written:")
            for path in written_files:
                print(f"  - {path}")
        else:
            print("\nNo adjusted PalletLines files were written.")
    by_slip_week = getattr(args, "slip_week", False)
    results = run_all_weeks(out_dir, by_slip_week=by_slip_week)
    if getattr(args, "inventory_flow_balance", False):
        results = apply_inventory_flow_balance(
            results,
            max_carry_weeks=getattr(args, "inventory_flow_max_carry_weeks", 2),
            apply_to_traceability=getattr(args, "inventory_flow_apply_to_traceability", False),
        )
        print(
            f"\nInventory flow balance enabled (max carry weeks = {int(getattr(args, 'inventory_flow_max_carry_weeks', 2))}; "
            f"apply_to_traceability = {bool(getattr(args, 'inventory_flow_apply_to_traceability', False))})."
        )
    print_report(results, args.week, by_slip_week=by_slip_week)
    if args.export and results:
        export_excel(results, exports_dir=out_dir)
    if getattr(args, "export_traceability_csv", False):
        # Original CSV -> Traceability_Exports/Original; balanced CSV -> BalancedWorking.
        if getattr(args, "balance_palletlines", False) or getattr(args, "balance_palletlines_aggressive", False):
            validate_balanced_working()
            run_all_weeks_balanced_to_csv()  # writes to BALANCED_WORKING_DIR
        elif getattr(args, "balanced_traceability_csv", False):
            validate_balanced_working()
            run_all_weeks_balanced_to_csv()
        else:
            validate_original_palletlines()
            run_all_weeks_original_to_csv()  # writes to ORIGINAL_EXPORTS_DIR
    if getattr(args, "working", False) and results:
        export_traceability_working(results)
    if getattr(args, "comprehensive", False) and results:
        export_comprehensive_report(results)

    return 0


if __name__ == "__main__":
    sys.exit(main() or 0)
