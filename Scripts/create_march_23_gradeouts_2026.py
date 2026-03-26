from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl


def round_half_up(value: float, ndigits: int) -> float:
    q = Decimal("1").scaleb(-ndigits)
    return float(Decimal(str(value)).quantize(q, rounding=ROUND_HALF_UP))


def pct(n: int, total: int) -> float:
    if total <= 0:
        return 0.0
    return float(Decimal(n) * Decimal(100) / Decimal(total))


@dataclass(frozen=True)
class BarnCounts:
    jumbo: int
    xlarge: int
    large: int
    medium: int
    small: int
    peewee: int
    liquid: int
    blood: int
    crack: int
    dirt: int
    avg_wt_g: float
    prod_doz: float

    @property
    def jxl(self) -> int:
        return self.jumbo + self.xlarge

    @property
    def undergrade(self) -> int:
        return self.liquid + self.blood + self.crack + self.dirt

    @property
    def total_eggs(self) -> int:
        # Total from the gradeout table convention you shared
        return self.jxl + self.large + self.medium + self.small + self.peewee + self.undergrade


def main() -> None:
    out = Path("grade outs") / "March_23_grade_outs_2026.xlsx"

    # Match Feb header order.
    barns = [10, 11, 6]

    # Counts transcribed by you for 2026-03-23.
    by_barn: dict[int, BarnCounts] = {
        10: BarnCounts(
            jumbo=1773,
            xlarge=14733,
            large=31274,
            medium=3667,
            small=36,
            peewee=3,
            liquid=147,
            blood=95,
            crack=166,
            dirt=266,
            avg_wt_g=62.037,
            prod_doz=4346.75,
        ),
        6: BarnCounts(
            jumbo=2520,
            xlarge=24244,
            large=62188,
            medium=8843,
            small=83,
            peewee=8,  # Peewee + WeePee (7+1) as you noted
            liquid=178,
            blood=103,
            crack=90,
            dirt=691,
            avg_wt_g=61.491,
            prod_doz=8245.83,
        ),
        11: BarnCounts(
            jumbo=1077,
            xlarge=10642,
            large=26350,
            medium=3299,
            small=29,
            peewee=1,
            liquid=77,
            blood=59,
            crack=46,
            dirt=520,
            avg_wt_g=61.645,
            prod_doz=3508.42,
        ),
    }

    categories = [
        "J/XL",
        "Jumbo",
        "Xlarge",
        "Large",
        "Medium",
        "Small",
        "Peewee",
        "Undergrade",
        "Avg. Wt. (g)",
        "Prod. (doz.)",
        "Feed (kg)",
        "Liquid",
        "Blood",
        "Crack",
        "Dirt",
        "Total %",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = ["Category"] + [f"Barn {b}" for b in barns]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    col_map = {barn: 2 + i for i, barn in enumerate(barns)}

    percent_rows = {
        "J/XL",
        "Jumbo",
        "Xlarge",
        "Large",
        "Medium",
        "Small",
        "Peewee",
        "Undergrade",
        "Liquid",
        "Blood",
        "Crack",
        "Dirt",
        "Total %",
    }

    for row_idx, category in enumerate(categories, start=2):
        ws.cell(row=row_idx, column=1, value=category)
        for barn in barns:
            if barn not in by_barn:
                continue
            bc = by_barn[barn]
            total = bc.total_eggs

            v: float | None
            if category == "J/XL":
                v = pct(bc.jxl, total)
            elif category == "Jumbo":
                v = pct(bc.jumbo, total)
            elif category == "Xlarge":
                v = pct(bc.xlarge, total)
            elif category == "Large":
                v = pct(bc.large, total)
            elif category == "Medium":
                v = pct(bc.medium, total)
            elif category == "Small":
                v = pct(bc.small, total)
            elif category == "Peewee":
                v = pct(bc.peewee, total)
            elif category == "Undergrade":
                v = pct(bc.undergrade, total)
            elif category == "Avg. Wt. (g)":
                v = bc.avg_wt_g
            elif category == "Prod. (doz.)":
                v = bc.prod_doz
            elif category == "Feed (kg)":
                v = None
            elif category == "Liquid":
                v = pct(bc.liquid, total)
            elif category == "Blood":
                v = pct(bc.blood, total)
            elif category == "Crack":
                v = pct(bc.crack, total)
            elif category == "Dirt":
                v = pct(bc.dirt, total)
            elif category == "Total %":
                v = 100.0
            else:
                v = None

            if v is None:
                continue

            if category in percent_rows:
                v = round_half_up(float(v), 1)
            elif category in {"Avg. Wt. (g)", "Prod. (doz.)"}:
                v = round_half_up(float(v), 1)

            ws.cell(row=row_idx, column=col_map[barn], value=v)

    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()

