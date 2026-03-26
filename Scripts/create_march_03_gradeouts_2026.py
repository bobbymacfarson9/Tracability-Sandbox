from __future__ import annotations

from pathlib import Path

import openpyxl


def main() -> None:
    out = Path("grade outs") / "March_03_grade_outs_2026.xlsx"

    # Columns follow your March 3 sequence: Barn 7 -> Barn 6 -> Barn 11
    barns = [7, 6, 11]

    # Percentages come from your March 3 YAML.
    # Note: Excel doesn't have a separate "SJumbo" row, so we add SJumbo% into the "Jumbo" row.
    # Note: YAML doesn't show "Undergrade" explicitly; we compute it so Liquid+Blood+Crack+Dirt+Undergrade
    # equals the YAML offgrades_summary percentage (rounding tolerance).
    data = {
        "J/XL": {7: 30.1, 6: 29.7, 11: 32.8},
        "Jumbo": {7: 3.1, 6: 2.4, 11: 3.5},
        "Xlarge": {7: 27.0, 6: 27.3, 11: 29.3},
        "Large": {7: 61.7, 6: 62.3, 11: 60.6},
        "Medium": {7: 7.0, 6: 6.2, 11: 5.4},
        "Small": {7: 0.1, 6: 0.1, 11: 0.1},
        "Peewee": {7: 0.0, 6: 0.0, 11: 0.0},
        "Undergrade": {7: 0.1, 6: 0.1, 11: 0.0},
        "Avg. Wt. (g)": {7: 61.912, 6: 61.949, 11: 62.295},
        "Prod. (doz.)": {7: 26620 / 12, 6: 72643 / 12, 11: 27349 / 12},
        "Feed (kg)": {7: None, 6: None, 11: None},
        "Liquid": {7: 0.2, 6: 0.4, 11: 0.2},
        "Blood": {7: 0.0, 6: 0.1, 11: 0.1},
        "Crack": {7: 0.3, 6: 0.5, 11: 0.1},
        "Dirt": {7: 0.5, 6: 0.7, 11: 0.8},
        "Total %": {7: 100.0, 6: 100.0, 11: 100.0},
    }

    rows = [
        "J/XL",
        "Jumbo",
        "Xlarge",
        "Large",
        "Medium",
        "Small",
        "Peewee",
        "Undergrade",
        "Avg. Wt. (g)",
        "Prod. (doz.)",
        "Feed (kg)",
        "Liquid",
        "Blood",
        "Crack",
        "Dirt",
        "Total %",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = ["Category"] + [f"Barn {b}" for b in barns]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Excel columns: A=Category, B=Barn 7, C=Barn 6, D=Barn 11
    col_map = {7: 2, 6: 3, 11: 4}

    for row_idx, category in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=category)
        for barn in barns:
            v = data[category][barn]
            if category == "Prod. (doz.)" and v is not None:
                v = round(float(v), 1)
            ws.cell(row=row_idx, column=col_map[barn], value=v)

    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()

