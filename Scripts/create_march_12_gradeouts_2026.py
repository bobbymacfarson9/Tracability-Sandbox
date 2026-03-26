from __future__ import annotations

from pathlib import Path

import openpyxl
from decimal import Decimal, ROUND_HALF_UP


def round_half_up(value: float, ndigits: int) -> float:
    q = Decimal("1").scaleb(-ndigits)  # e.g. ndigits=1 => 0.1
    return float(Decimal(str(value)).quantize(q, rounding=ROUND_HALF_UP))


def main() -> None:
    out = Path("grade outs") / "March_12_grade_outs_2026.xlsx"

    # Match your existing multi-barn Excel convention (see other months):
    # Barn order: 10, 6, 7, 11
    barns = [10, 6, 7, 11]

    # Source from your March 12 transcription.
    # "J/XL" is Jumbo% + XLarge%.
    categories = [
        "J/XL",
        "Jumbo",
        "Xlarge",
        "Large",
        "Medium",
        "Small",
        "Peewee",
        "Undergrade",
        "Avg. Wt. (g)",
        "Prod. (doz.)",
        "Feed (kg)",
        "Liquid",
        "Blood",
        "Crack",
        "Dirt",
        "Total %",
    ]

    data = {
        "J/XL": {10: 36.9, 6: 35.8, 7: 12.6, 11: 32.6},
        "Jumbo": {10: 4.6, 6: 4.4, 7: 1.7, 11: 3.6},
        "Xlarge": {10: 32.3, 6: 31.4, 7: 10.9, 11: 29.0},
        "Large": {10: 57.2, 6: 57.1, 7: 15.6, 11: 59.9},
        "Medium": {10: 4.4, 6: 6.0, 7: 36.5, 11: 6.0},
        "Small": {10: 0.1, 6: 0.1, 7: 16.0, 11: 0.0},
        "Peewee": {10: 0.0, 6: 0.0, 7: 1.4, 11: 0.0},
        "Undergrade": {10: 1.6, 6: 1.0, 7: 17.9, 11: 1.5},
        "Avg. Wt. (g)": {10: 62.764, 6: 62.508, 7: 54.567, 11: 62.250},
        # Counted/12
        "Prod. (doz.)": {10: 29010 / 12, 6: 49285 / 12, 7: 1215 / 12, 11: 25910 / 12},
        "Feed (kg)": {10: None, 6: None, 7: None, 11: None},
        "Liquid": {10: 0.3, 6: 0.2, 7: 16.2, 11: 0.2},
        "Blood": {10: 0.1, 6: 0.1, 7: 1.3, 11: 0.2},
        "Crack": {10: 0.4, 6: 0.1, 7: 0.2, 11: 0.2},
        "Dirt": {10: 0.8, 6: 0.6, 7: 0.2, 11: 0.9},
        "Total %": {10: 100.0, 6: 100.0, 7: 100.0, 11: 100.0},
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = ["Category"] + [f"Barn {b}" for b in barns]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    col_map = {barn: 2 + i for i, barn in enumerate(barns)}

    for row_idx, category in enumerate(categories, start=2):
        ws.cell(row=row_idx, column=1, value=category)
        for barn in barns:
            v = data[category][barn]
            if v is None:
                continue
            if category in {"J/XL", "Jumbo", "Xlarge", "Large", "Medium", "Small", "Peewee", "Undergrade", "Liquid", "Blood", "Crack", "Dirt", "Total %"}:
                # Store 1-decimal floats for consistent Excel display.
                v = round_half_up(float(v), 1)
            elif category == "Avg. Wt. (g)":
                v = round_half_up(float(v), 1)
            elif category == "Prod. (doz.)":
                v = round_half_up(float(v), 1)
            ws.cell(row=row_idx, column=col_map[barn], value=v)

    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()

