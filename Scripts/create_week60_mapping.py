"""
Create Week 60 SKU Mapping Table
Compares Week 56 (old) vs Week 60 (new with OD column) loading slips
and generates a new mapping table with correct cell addresses.
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
import sys
from datetime import datetime
import io

# Fix Windows console encoding for special characters
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# Paths - automatically detect script location
SCRIPT_DIR = Path(__file__).parent
BASE_DIR = SCRIPT_DIR.parent  # Go up one level from Scripts folder
REF_DATA_DIR = BASE_DIR / "Reference_Data"

# File names (slip files can be named slightly differently; we search flexibly)
OLD_SLIP = "Week 56 Loading Slipp 2026.xlsx"   # original expected name
NEW_SLIP = "Week 60 loading slip 2026.xlsx"    # original expected name
OLD_MAPPING = "Week_42_Stop_SKU_Final_POLISHED.xlsx"
NEW_MAPPING = "Week_60_Stop_SKU_Final_POLISHED.xlsx"


def _find_slip_like(pattern_candidates):
    """
    Find a slip file in REF_DATA_DIR matching any of the pattern candidates.
    pattern_candidates: list of (exact_name, case_insensitive_contains) tuples.
    Returns Path or None.
    """
    # Prefer exact file name if present
    for exact, _ in pattern_candidates:
        if not exact:
            continue
        p = REF_DATA_DIR / exact
        if p.exists():
            return p
    # Fallback: search by contains/substring, case-insensitive
    all_files = list(REF_DATA_DIR.glob("*.xlsx"))
    for _, substr in pattern_candidates:
        if not substr:
            continue
        substr_low = substr.lower()
        for f in all_files:
            name_low = f.name.lower()
            if substr_low in name_low:
                return f
    return None


def find_header_row(ws):
    """Find the header row (usually row 4, but search for common headers)"""
    for row_num in range(1, 11):
        cell_value = str(ws.cell(row=row_num, column=1).value or "").upper()
        if any(keyword in cell_value for keyword in ["OC", "OD", "SKU", "QTY", "EGG"]):
            return row_num
    return 4  # Default to row 4


def get_column_structure(ws, header_row):
    """Extract column headers and their positions"""
    headers = {}
    max_col = ws.max_column
    
    for col_num in range(1, max_col + 1):
        header = str(ws.cell(row=header_row, column=col_num).value or "").strip()
        if header:
            col_letter = get_column_letter(col_num)
            headers[col_letter] = {
                'header': header,
                'col_num': col_num,
                'col_letter': col_letter
            }
    
    return headers, header_row


def compare_structures(old_headers, new_headers):
    """Compare old vs new column structures"""
    print("\n=== COLUMN STRUCTURE COMPARISON ===\n")
    
    print("WEEK 56 (OLD) COLUMNS:")
    for col_letter in sorted(old_headers.keys()):
        print(f"  {col_letter}: {old_headers[col_letter]['header']}")
    
    print("\nWEEK 60 (NEW) COLUMNS:")
    for col_letter in sorted(new_headers.keys()):
        print(f"  {col_letter}: {new_headers[col_letter]['header']}")
    
    # Find new columns
    print("\nNEW COLUMNS IN WEEK 60:")
    new_cols = []
    for col_letter, info in new_headers.items():
        if col_letter not in old_headers:
            print(f"  {col_letter}: {info['header']}")
            new_cols.append(col_letter)
    
    # Find column shifts
    print("\nCOLUMN SHIFTS (same header, different position):")
    shifts = {}
    for old_col, old_info in old_headers.items():
        header = old_info['header'].upper()
        # Find matching header in new structure
        for new_col, new_info in new_headers.items():
            if new_info['header'].upper() == header and old_col != new_col:
                old_num = old_info['col_num']
                new_num = new_info['col_num']
                shift = new_num - old_num
                if shift != 0:
                    print(f"  '{old_info['header']}': {old_col} → {new_col} (shift: {shift:+d})")
                    shifts[old_col] = {'new_col': new_col, 'shift': shift}
                break
    
    return shifts, new_cols


def read_mapping_table(mapping_file):
    """Read the existing Week 42 mapping table"""
    print(f"\nReading mapping table: {mapping_file}")
    
    try:
        df = pd.read_excel(mapping_file, sheet_name=0)
        print(f"  Found {len(df)} rows")
        print(f"  Columns: {list(df.columns)}")
        return df
    except Exception as e:
        print(f"  Error reading mapping table: {e}")
        return None


def shift_cell_address(cell_addr, shift):
    """Shift a cell address by N columns (e.g., 'B5' + 1 = 'C5')"""
    if pd.isna(cell_addr) or not cell_addr:
        return cell_addr
    
    cell_addr = str(cell_addr).strip()
    
    # Extract column letter(s) and row number
    col_part = ""
    row_part = ""
    
    for char in cell_addr:
        if char.isalpha():
            col_part += char
        elif char.isdigit():
            row_part += char
    
    if not col_part or not row_part:
        return cell_addr  # Invalid format, return as-is
    
    # Convert column to number, shift, convert back
    col_num = 0
    for i, char in enumerate(reversed(col_part)):
        col_num += (ord(char.upper()) - 64) * (26 ** i)
    
    new_col_num = col_num + shift
    
    if new_col_num < 1:
        return cell_addr  # Can't shift left past column A
    
    # Convert back to column letter(s)
    new_col_letter = ""
    while new_col_num > 0:
        new_col_num -= 1
        new_col_letter = chr(65 + (new_col_num % 26)) + new_col_letter
        new_col_num //= 26
    
    return new_col_letter + row_part


def create_week60_mapping(old_mapping_df, shifts):
    """Create new Week 60 mapping table by shifting cell addresses"""
    print("\n=== CREATING WEEK 60 MAPPING TABLE ===\n")
    
    # Copy the old mapping
    new_mapping = old_mapping_df.copy()
    
    # Find QtyCellAddr and SKUCellAddr columns (handle various naming conventions)
    qty_col = None
    sku_col = None
    
    for col in new_mapping.columns:
        col_upper = str(col).upper().replace(" ", "")
        # Check for quantity cell column (various names)
        if any(term in col_upper for term in ["QTYCELLADDR", "QTYCELL", "QUANTITYCELL", "QUANTITYCELLADDR", "QTYCELLADDRESS"]):
            qty_col = col
        # Check for SKU cell column (various names)
        if any(term in col_upper for term in ["SKUCELLADDR", "SKUCELL", "SKUCELLADDRESS"]):
            sku_col = col
    
    # Also check for "Quantity Cell" and "SKU Cell" (with spaces)
    if not qty_col:
        for col in new_mapping.columns:
            if "quantity" in str(col).lower() and "cell" in str(col).lower():
                qty_col = col
                break
    
    if not sku_col:
        for col in new_mapping.columns:
            if "sku" in str(col).lower() and "cell" in str(col).lower():
                sku_col = col
                break
    
    if not qty_col or not sku_col:
        print("ERROR: Could not find QtyCellAddr or SKUCellAddr columns")
        print(f"Available columns: {list(new_mapping.columns)}")
        return None
    
    print(f"Found QtyCellAddr column: {qty_col}")
    print(f"Found SKUCellAddr column: {sku_col}")
    
    # Calculate shift amount (typically +1 for OD column insertion)
    # We'll use the average shift from comparison, or default to +1
    avg_shift = shifts.get('B', {}).get('shift', 1) if shifts else 1
    print(f"Using column shift: +{avg_shift}")
    
    # Add OD Cell Address column if it doesn't exist
    od_col = None
    for col in new_mapping.columns:
        col_upper = str(col).upper().replace(" ", "")
        if "ODCELL" in col_upper or ("OD" in col_upper and "CELL" in col_upper):
            od_col = col
            break
    
    if od_col is None:
        # Insert OD Cell column after Quantity Cell column
        qty_col_idx = list(new_mapping.columns).index(qty_col)
        new_mapping.insert(qty_col_idx + 1, 'OD Cell', '')
        od_col = 'OD Cell'
        print(f"Added new column: {od_col}")
    else:
        print(f"Found existing OD Cell column: {od_col}")
    
    # Shift all cell addresses and calculate OD cell addresses
    updated_qty = 0
    updated_sku = 0
    updated_od = 0
    
    for idx in new_mapping.index:
        # Shift QtyCellAddr
        old_qty_addr = new_mapping.at[idx, qty_col]
        if pd.notna(old_qty_addr) and str(old_qty_addr).strip():
            new_qty_addr = shift_cell_address(old_qty_addr, avg_shift)
            if new_qty_addr != old_qty_addr:
                new_mapping.at[idx, qty_col] = new_qty_addr
                updated_qty += 1
                
                # Calculate OD Cell address (one column to the left of Quantity Cell)
                od_addr = shift_cell_address(new_qty_addr, -1)
                if od_addr and od_addr != new_qty_addr:
                    new_mapping.at[idx, od_col] = od_addr
                    updated_od += 1
        
        # Shift SKUCellAddr
        old_sku_addr = new_mapping.at[idx, sku_col]
        if pd.notna(old_sku_addr) and str(old_sku_addr).strip():
            new_sku_addr = shift_cell_address(old_sku_addr, avg_shift)
            if new_sku_addr != old_sku_addr:
                new_mapping.at[idx, sku_col] = new_sku_addr
                updated_sku += 1
    
    print(f"\nUpdated {updated_qty} QtyCellAddr addresses")
    print(f"Updated {updated_sku} SKUCellAddr addresses")
    print(f"Updated {updated_od} OD Cell addresses")
    
    return new_mapping


def save_mapping_table(df, output_file):
    """Save the new mapping table to Excel"""
    print(f"\nSaving new mapping table: {output_file}")
    
    try:
        # Create Excel writer
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Sheet1', index=False)
            
            # Format the worksheet
            worksheet = writer.sheets['Sheet1']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Format header row
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        print(f"  [OK] Successfully saved {len(df)} rows")
        return True
        
    except Exception as e:
        print(f"  [ERROR] Error saving file: {e}")
        return False


def main():
    """Main execution"""
    print("=" * 60)
    print("WEEK 60 SKU MAPPING TABLE GENERATOR")
    print("=" * 60)
    
    # Check files exist
    old_slip_path = REF_DATA_DIR / OLD_SLIP
    new_slip_path = REF_DATA_DIR / NEW_SLIP
    old_mapping_path = REF_DATA_DIR / OLD_MAPPING
    new_mapping_path = REF_DATA_DIR / NEW_MAPPING
    
    print(f"\nChecking files...")
    print(f"  Reference Data folder: {REF_DATA_DIR}")
    
    missing_files = []
    if not REF_DATA_DIR.exists():
        print(f"  ✗ Reference_Data folder not found: {REF_DATA_DIR}")
        missing_files.append(f"Folder: {REF_DATA_DIR}")
    else:
        if not old_slip_path.exists():
            missing_files.append(f"Week 56 slip: {old_slip_path}")
            print(f"  [X] {OLD_SLIP} not found")
        else:
            print(f"  [OK] {OLD_SLIP} found")
        
        if not new_slip_path.exists():
            missing_files.append(f"Week 60 slip: {new_slip_path}")
            print(f"  [X] {NEW_SLIP} not found")
        else:
            print(f"  [OK] {NEW_SLIP} found")
        
        if not old_mapping_path.exists():
            missing_files.append(f"Mapping table: {old_mapping_path}")
            print(f"  [X] {OLD_MAPPING} not found")
        else:
            print(f"  [OK] {OLD_MAPPING} found")
    
    if missing_files:
        print("\nERROR: Missing required files:")
        for f in missing_files:
            print(f"  - {f}")
        print(f"\nPlease ensure all files are in: {REF_DATA_DIR}")
        return
    
    # Load workbooks
    print("\nLoading files...")
    old_wb = openpyxl.load_workbook(old_slip_path, data_only=True)
    new_wb = openpyxl.load_workbook(new_slip_path, data_only=True)
    
    old_ws = old_wb.active
    new_ws = new_wb.active
    
    # Find header rows
    old_header_row = find_header_row(old_ws)
    new_header_row = find_header_row(new_ws)
    
    print(f"  Week 56 header row: {old_header_row}")
    print(f"  Week 60 header row: {new_header_row}")
    
    # Get column structures
    old_headers, _ = get_column_structure(old_ws, old_header_row)
    new_headers, _ = get_column_structure(new_ws, new_header_row)
    
    # Compare structures
    shifts, new_cols = compare_structures(old_headers, new_headers)
    
    # Read old mapping table
    old_mapping_df = read_mapping_table(old_mapping_path)
    if old_mapping_df is None:
        return
    
    # Create new mapping table
    new_mapping_df = create_week60_mapping(old_mapping_df, shifts)
    if new_mapping_df is None:
        return
    
    # Save new mapping table
    if save_mapping_table(new_mapping_df, new_mapping_path):
        print("\n" + "=" * 60)
        print("SUCCESS!")
        print("=" * 60)
        print(f"\nNew mapping table created: {new_mapping_path}")
        print(f"Total rows: {len(new_mapping_df)}")
        print("\nNext steps:")
        print("  1. Review the new mapping table")
        print("  2. Test with Week 60 loading slip")
        print("  3. Update your system to use Week_60_Stop_SKU_Final_POLISHED.xlsx")
    else:
        print("\nFailed to save mapping table")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
