"""
Weekly Loading Slip Processor
Processes an entire week's loading slip, extracts all SKUs, quantities, and OD quantities,
and generates a complete traceability report.
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
import sys
from datetime import datetime, timedelta
import io
import argparse

# Fix Windows console encoding
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# Paths
SCRIPT_DIR = Path(__file__).parent
BASE_DIR = SCRIPT_DIR.parent
REF_DATA_DIR = BASE_DIR / "Reference_Data"
EXPORTS_DIR = BASE_DIR / "Traceability_Exports"


def _get_loading_slip_dirs():
    """Return list of directories to search for loading slips: Reference_Data plus paths from paths.json LoadingSlipsPaths."""
    dirs = [REF_DATA_DIR]
    try:
        import json
        paths_file = REF_DATA_DIR / "paths.json"
        if paths_file.exists():
            with open(paths_file, encoding="utf-8") as f:
                config = json.load(f)
            for path in config.get("LoadingSlipsPaths") or []:
                p = Path(path)
                if not p.is_absolute():
                    p = BASE_DIR / p
                if p.exists() and p.is_dir() and p not in dirs:
                    dirs.append(p)
    except Exception:
        pass
    return dirs


def _get_2026_slip_week_mapping():
    """
    Load 2026 calendar week (1-6) -> ISO slip week (56-61) from paths.json.
    Returns dict: 1->56, 2->57, ... and inverse 56->1, 57->2, ...
    """
    try:
        import json
        paths_file = REF_DATA_DIR / "paths.json"
        if paths_file.exists():
            with open(paths_file, encoding="utf-8") as f:
                config = json.load(f)
            cal_to_iso = config.get("LoadingSlip2026CalendarToISO") or {}
            cal_to_iso = {int(k): int(v) for k, v in cal_to_iso.items()}
            iso_to_cal = {v: k for k, v in cal_to_iso.items()}
            return cal_to_iso, iso_to_cal
    except Exception:
        pass
    return {}, {}


def _is_2026_loading_slip_folder(path):
    """True if this directory is the 2026 loading slip folder (calendar-week naming)."""
    name = path.name if hasattr(path, "name") else str(path)
    return "2026" in name and ("loading" in name.lower() or "slip" in name.lower())


def find_header_row(ws):
    """Find the header row (usually row 3-4)"""
    for row_num in range(1, 11):
        cell_value = str(ws.cell(row=row_num, column=1).value or "").upper()
        if any(keyword in cell_value for keyword in ["OC", "OD", "SKU", "QTY", "EGG"]):
            return row_num
    return 4


def find_stop_headers(ws, header_row):
    """Find all stop headers and their column positions, also detect OD column positions
    Stops can appear in multiple rows, not just the header row"""
    stops = {}
    max_col = ws.max_column
    max_row = min(ws.max_row, 100)  # Scan up to row 100
    seen_stops = set()  # Track stops we've already found by name
    
    # Scan all rows for stops (they can appear anywhere)
    for row_num in range(1, max_row + 1):
        for col_num in range(1, max_col + 1):
            cell_value = str(ws.cell(row=row_num, column=col_num).value or "").strip()
            
            # Look for stop names (usually contain numbers and store names)
            if cell_value and any(char.isdigit() for char in cell_value):
                # Check if it looks like a stop header (e.g., "3. Sobeys - Prince Street")
                if "." in cell_value or any(word in cell_value.upper() for word in ["SOBEYS", "SUPERSTORE", "WALMART", "FOODLAND", "SHOPPERS", "PHARMASAVE", "HARBOUR", "NORTHSIDE", "MINERS", "SEXTON", "GENERAL"]):
                    # Skip if we already found this exact stop name
                    if cell_value in seen_stops:
                        continue
                    
                    seen_stops.add(cell_value)
                    col_letter = get_column_letter(col_num)
                    
                    # Find OD column - look left from stop header for "OD" header
                    od_col = None
                    for check_col in range(col_num - 1, max(1, col_num - 3), -1):  # Check up to 3 columns left
                        check_value = str(ws.cell(row=row_num, column=check_col).value or "").strip().upper()
                        if check_value == "OD":
                            od_col = check_col
                            break
                    
                    # Use unique key: column letter + row number to handle multiple stops in same column
                    stop_key = f"{col_letter}{row_num}"
                    stops[stop_key] = {
                        'name': cell_value,
                        'col_num': col_num,
                        'col_letter': col_letter,
                        'od_col': od_col,  # Column number where OD header is found
                        'header_row': row_num  # Row where this stop header was found
                    }
    
    return stops


def _looks_like_stop_header_text(value):
    """True when a cell looks like a stop header such as '8. Midtown Pharmasave'."""
    if value is None:
        return False
    text = str(value).strip()
    if not text:
        return False
    upper = text.upper()
    if "TOTAL" in upper:
        return False
    return any(ch.isdigit() for ch in text) and "." in text


def load_mapping_table(week_number):
    """Load the mapping table for the specified week.
    Every week before 59 (weeks 1–58, no OD date columns): use Week_42 mapping only.
    Week 59 and later (slips with OD columns): use week-specific or Week_60 mapping.
    NFLD sheets use the same table; mapping must have Day=NFLD rows for NFLD stops.
    """
    week_str = str(week_number).strip()
    try:
        w = int(week_number)
    except (TypeError, ValueError):
        w = 59
    if w < 59:
        # Every sheet before 59: Week_42 SKU mapping only
        mapping_files = ["Week_42_Stop_SKU_Final_POLISHED.xlsx"]
    else:
        # OD columns: prefer week-specific then Week_60 then Week_42
        mapping_files = [
            f"Week_{week_str}_Stop_SKU_Final_POLISHED.xlsx",
            "Week_60_Stop_SKU_Final_POLISHED.xlsx",
            "Week_42_Stop_SKU_Final_POLISHED.xlsx"
        ]
    for mapping_file in mapping_files:
        mapping_path = REF_DATA_DIR / mapping_file
        if mapping_path.exists():
            print(f"  Loading mapping table: {mapping_file}")
            try:
                df = pd.read_excel(mapping_path, sheet_name=0)
                return df
            except Exception as e:
                print(f"  Error loading {mapping_file}: {e}")
                continue
    
    print("  WARNING: No mapping table found, will scan columns manually")
    return None


def _mapping_stops_for_day(mapping_df, day_name):
    """
    Build stop entries directly from the mapping table for a given day.
    This is especially useful for NFLD sheets, where customer headers do not look like
    numbered store stops and the generic stop-header scanner misses them.
    """
    if mapping_df is None or day_name is None:
        return {}

    day_col = None
    stop_col = None
    qty_cell_col = None
    sku_cell_col = None
    for col in mapping_df.columns:
        col_upper = str(col).upper().replace(" ", "")
        col_orig = str(col)
        if day_col is None and "DAY" in col_upper:
            day_col = col
        if stop_col is None and (("STOP" in col_upper and "NAME" in col_upper) or col_upper == "STOP"):
            stop_col = col
        if qty_cell_col is None and ("QTYCELL" in col_upper or "QUANTITYCELL" in col_upper or "QTYCELLADDR" in col_upper or col_orig == "Quantity Cell"):
            qty_cell_col = col
        if sku_cell_col is None and ("SKUCELL" in col_upper or "SKUCELLADDR" in col_upper or col_orig == "SKU Cell"):
            sku_cell_col = col

    if day_col is None or stop_col is None:
        return {}

    day_norm = str(day_name).strip().upper()[:3]
    day_rows = mapping_df[mapping_df[day_col].astype(str).str.strip().str.upper().str[:3] == day_norm]
    if day_rows.empty:
        return {}

    stops = {}
    seen = set()
    for _, row in day_rows.iterrows():
        stop_name = str(row.get(stop_col) or "").strip()
        if not stop_name or stop_name in seen:
            continue
        seen.add(stop_name)
        qty_cell = str(row.get(qty_cell_col) or "").strip() if qty_cell_col else ""
        sku_cell = str(row.get(sku_cell_col) or "").strip() if sku_cell_col else ""
        cell_ref = qty_cell or sku_cell
        col_num = None
        col_letter = ""
        header_row = 0
        if cell_ref:
            try:
                from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
                col_letter, header_row = coordinate_from_string(cell_ref)
                col_num = column_index_from_string(col_letter)
            except Exception:
                col_num = None
                col_letter = ""
                header_row = 0
        stop_key = f"{col_letter or 'MAP'}_{len(stops)+1}"
        stops[stop_key] = {
            "name": stop_name,
            "col_num": col_num,
            "col_letter": col_letter,
            "od_col": None,
            "header_row": header_row,
        }
    return stops


def extract_sku_data_using_mapping(ws, stop_name, day_name, mapping_df, stop_col_info=None):
    """Extract SKU data using mapping table cell addresses
    
    Uses EXACT cell addresses from mapping table for the exact stop match.
    If exact match not found, falls back to column structure from store type match.
    
    Args:
        ws: Worksheet to read from
        stop_name: Name of the stop to find
        day_name: Day name (Monday, Tuesday, etc.)
        mapping_df: Mapping table DataFrame
        stop_col_info: Dict with 'col_num' and 'header_row' for where stop appears
    """
    skus = []
    
    if mapping_df is None:
        return skus
    
    # Debug: Print what we're looking for
    print(f"    Looking for: Day='{day_name}', Stop='{stop_name}'")
    
    # Find columns in mapping table
    day_col = None
    stop_col = None
    qty_cell_col = None
    sku_cell_col = None
    od_cell_col = None
    
    for col in mapping_df.columns:
        col_upper = str(col).upper().replace(" ", "")
        col_orig = str(col)
        
        # Day column
        if "DAY" in col_upper and day_col is None:
            day_col = col
        
        # Stop column
        if ("STOP" in col_upper and "NAME" in col_upper) or (col_upper == "STOP"):
            stop_col = col
        
        # Quantity Cell column - handle various naming conventions
        if qty_cell_col is None:
            if ("QTYCELL" in col_upper or "QUANTITYCELL" in col_upper or 
                "QTYCELLADDR" in col_upper or col_orig == "Quantity Cell"):
                qty_cell_col = col
        
        # SKU Cell column - handle various naming conventions
        if sku_cell_col is None:
            if ("SKUCELL" in col_upper or "SKUCELLADDR" in col_upper or 
                col_orig == "SKU Cell"):
                sku_cell_col = col
        
        # OD Cell column (may not exist yet)
        if od_cell_col is None:
            if ("ODCELL" in col_upper or ("OD" in col_upper and "CELL" in col_upper) or
                col_orig == "OD Cell" or col_orig == "OD Cell Address"):
                od_cell_col = col
    
    if day_col is None or stop_col is None or qty_cell_col is None or sku_cell_col is None:
        print(f"    WARNING: Missing required columns in mapping table")
        return skus
    
    # Normalize day name for matching
    day_normalized = day_name.upper()[:3]  # Mon, Tue, Wed, etc.
    day_map = {
        'MON': 'Monday', 'TUE': 'Tuesday', 'WED': 'Wednesday',
        'THU': 'Thursday', 'FRI': 'Friday', 'SAT': 'Saturday', 'SUN': 'Sunday'
    }
    day_full = day_map.get(day_normalized, day_name)
    
    # Find EXACT stop match first, then fall back to store type match
    day_matches = 0
    exact_stop_matches = []  # Store exact matches with cell addresses
    store_type_matches = []  # Store store type matches for fallback
    stop_normalized = stop_name.upper().strip()  # Normalize stop name once
    
    # Ensure we have column names, not indices
    if not isinstance(day_col, str) or not isinstance(stop_col, str):
        print(f"    ERROR: Column names are not strings - day_col={day_col}, stop_col={stop_col}")
        return skus
    
    # First pass: find EXACT stop matches
    for idx in range(len(mapping_df)):
        try:
            row_data = mapping_df.iloc[idx]
            mapping_day_val = row_data[day_col] if day_col in row_data.index else None
            mapping_stop_val = row_data[stop_col] if stop_col in row_data.index else None
            
            mapping_day = str(mapping_day_val).strip().upper()[:3] if pd.notna(mapping_day_val) else ""
            mapping_stop = str(mapping_stop_val).strip() if pd.notna(mapping_stop_val) else ""
        except Exception as e:
            print(f"    Error reading row {idx}: {e}")
            continue
        
        # Check if day matches
        if mapping_day != day_normalized:
            continue
        
        day_matches += 1
        
        # Check if stop matches - need EXACT match
        mapping_stop_normalized = mapping_stop.upper().strip()
        
        # Try exact match first
        exact_match = (stop_normalized == mapping_stop_normalized)
        
        # If not exact, try normalized (remove extra spaces, dots, dashes)
        if not exact_match:
            stop_clean = stop_normalized.replace(".", "").replace("-", "").replace("  ", " ").strip()
            map_clean = mapping_stop_normalized.replace(".", "").replace("-", "").replace("  ", " ").strip()
            exact_match = (stop_clean == map_clean)
        
        # Only process if this is an exact match for THIS stop
        if not exact_match:
            # Check for store type match for fallback (but don't add to exact_stop_matches)
            import re
            store_keywords = ['SOBEYS', 'WALMART', 'SUPERSTORE', 'FOODLAND', 'SHOPPERS', 'PHARMASAVE']
            stop_keyword = [kw for kw in store_keywords if kw in stop_normalized]
            map_keyword = [kw for kw in store_keywords if kw in mapping_stop_normalized]
            
            if stop_keyword and map_keyword and stop_keyword[0] == map_keyword[0]:
                if not store_type_matches:  # Only store first one for column structure
                    try:
                        row_data = mapping_df.iloc[idx]
                        qty_cell = str(row_data[qty_cell_col]).strip() if pd.notna(row_data[qty_cell_col]) else ""
                        sku_cell = str(row_data[sku_cell_col]).strip() if pd.notna(row_data[sku_cell_col]) else ""
                        od_cell = ""
                        if od_cell_col and pd.notna(row_data[od_cell_col]):
                            od_cell = str(row_data[od_cell_col]).strip()
                        
                        if qty_cell and sku_cell:
                            store_type_matches.append({
                                'qty_cell': qty_cell,
                                'sku_cell': sku_cell,
                                'od_cell': od_cell,
                                'stop_name': mapping_stop
                            })
                    except:
                        pass
            continue  # Skip this row - not an exact match
        
        # Get cell addresses from this row (only if exact match)
        try:
            row_data = mapping_df.iloc[idx]
            qty_cell = str(row_data[qty_cell_col]).strip() if pd.notna(row_data[qty_cell_col]) else ""
            sku_cell = str(row_data[sku_cell_col]).strip() if pd.notna(row_data[sku_cell_col]) else ""
            od_cell = ""
            if od_cell_col and pd.notna(row_data[od_cell_col]):
                od_cell = str(row_data[od_cell_col]).strip()
            
            if qty_cell and sku_cell:
                # Store exact match with cell addresses
                exact_stop_matches.append({
                    'qty_cell': qty_cell,
                    'sku_cell': sku_cell,
                    'od_cell': od_cell,
                    'stop_name': mapping_stop
                })
        except Exception as e:
            continue
    
    # Use exact matches if available, otherwise fall back to store type matches
    matches_to_use = exact_stop_matches if exact_stop_matches else store_type_matches
    
    if not matches_to_use:
        print(f"    No matching rows found in mapping table")
        return skus
    
    if exact_stop_matches:
        print(f"      Found {len(exact_stop_matches)} exact stop match(es) for '{stop_name}'")
        # Filter to only matches that actually match our stop name
        filtered_matches = []
        for match in exact_stop_matches:
            match_stop_normalized = match['stop_name'].upper().strip()
            if match_stop_normalized == stop_normalized:
                filtered_matches.append(match)
            else:
                # Try normalized comparison
                match_clean = match_stop_normalized.replace(".", "").replace("-", "").replace("  ", " ").strip()
                stop_clean = stop_normalized.replace(".", "").replace("-", "").replace("  ", " ").strip()
                if match_clean == stop_clean:
                    filtered_matches.append(match)
        
        if filtered_matches:
            exact_stop_matches = filtered_matches
            print(f"      Filtered to {len(exact_stop_matches)} matches for exact stop '{stop_name}'")
        else:
            print(f"      WARNING: No exact matches after filtering - using first match anyway")
    elif store_type_matches:
        print(f"      Using store type match from '{store_type_matches[0]['stop_name']}'")
    
    # Get column structure from first match
    first_match = matches_to_use[0]
    # Convert cell addresses to column/row numbers
    def cell_to_col_row(cell_addr):
        """Convert cell address like 'C5' to (column_num, row_num)"""
        import re
        match = re.match(r'([A-Z]+)(\d+)', str(cell_addr).upper())
        if not match:
            return None, None
        col_letters = match.group(1)
        row_num = int(match.group(2))
        # Convert column letters to number
        col_num = 0
        for char in col_letters:
            col_num = col_num * 26 + (ord(char) - ord('A') + 1)
        return col_num, row_num
    
    qty_col_map, qty_row_map = cell_to_col_row(first_match['qty_cell'])
    sku_col_map, sku_row_map = cell_to_col_row(first_match['sku_cell'])
    od_col_map = None
    if first_match['od_cell']:
        od_col_map, _ = cell_to_col_row(first_match['od_cell'])
    
    if not qty_col_map or not sku_col_map:
        print(f"    Invalid cell addresses in mapping")
        return skus
    
    # Use EXACT cell addresses from mapping table - 1:1 mapping
    # The mapping table has exact cell addresses like K5, L5, I5 for each SKU
    # Use those addresses directly - no offset calculation needed!
    
    if exact_stop_matches:
        print(f"      Using EXACT cell addresses from mapping (1:1): {len(exact_stop_matches)} SKU rows")
        
        # Read directly from the exact cell addresses in the mapping
        for match in exact_stop_matches:
            try:
                # Parse cell addresses
                qty_col, qty_row = cell_to_col_row(match['qty_cell'])
                sku_col, sku_row = cell_to_col_row(match['sku_cell'])
                od_col = None
                if match['od_cell']:
                    od_col, _ = cell_to_col_row(match['od_cell'])
                
                if not qty_col or not sku_col:
                    continue
                
                # Read directly from these exact cells
                # Note: data_only=True reads calculated values, but if file has formulas
                # that haven't been calculated/saved, they'll show as None or 0
                qty_value = ws.cell(row=qty_row, column=qty_col).value
                sku_value = ws.cell(row=sku_row, column=sku_col).value
                od_value = ws.cell(row=qty_row, column=od_col).value if od_col else None
                
                # Parse quantities - handle None, empty strings, and numeric values
                qty = 0
                od_qty = 0
                
                try:
                    if qty_value is not None:
                        # Handle different types: int, float, string
                        if isinstance(qty_value, (int, float)):
                            qty = float(qty_value)
                        else:
                            qty_str = str(qty_value).strip()
                            if qty_str and qty_str.lower() not in ['none', '', 'nan']:
                                qty = float(qty_str)
                except (ValueError, TypeError):
                    # If parsing fails, qty remains 0
                    pass
                
                try:
                    if od_value is not None:
                        if isinstance(od_value, (int, float)):
                            od_qty = float(od_value)
                        else:
                            od_str = str(od_value).strip()
                            if od_str and od_str.lower() not in ['none', '', 'nan']:
                                od_qty = float(od_str)
                except (ValueError, TypeError):
                    pass
                
                # Get SKU name
                sku = str(sku_value).strip() if sku_value is not None else ""
                
                # Add if valid SKU (with or without quantity - user wants to see all SKUs)
                if sku and sku.upper() not in ["BOX TOTAL", "TOTAL", ""]:
                    skus.append({
                        'stop': stop_name,
                        'sku': sku,
                        'qty': int(qty) if qty > 0 else 0,
                        'od_qty': int(od_qty) if od_qty > 0 else 0,
                        'qty_cell': match['qty_cell'],
                        'sku_cell': match['sku_cell'],
                        'od_cell': match['od_cell'] if match['od_cell'] else ""
                    })
            except Exception as e:
                continue
    elif store_type_matches and stop_col_info:
        # Fallback: use column structure for store type matches
        stop_col = stop_col_info.get('col_num')
        header_row = stop_col_info.get('header_row', 3)
        data_start_row = header_row + 1
        
        mapped_stop_col = qty_col_map
        col_offset = stop_col - mapped_stop_col
        actual_qty_col = qty_col_map + col_offset
        actual_sku_col = sku_col_map + col_offset
        
        od_col_from_stop = stop_col_info.get('od_col')
        if od_col_from_stop:
            actual_od_col = od_col_from_stop
        elif od_col_map:
            actual_od_col = od_col_map + col_offset
        else:
            actual_od_col = qty_col_map - 1 + col_offset
        
        print(f"      Using column structure: Qty={get_column_letter(actual_qty_col)}, SKU={get_column_letter(actual_sku_col)}, OD={get_column_letter(actual_od_col) if actual_od_col else 'N/A'}")
        
        # Scan all data rows
        for row_num in range(data_start_row, data_start_row + 100):
            try:
                qty_value = ws.cell(row=row_num, column=actual_qty_col).value
                sku_value = ws.cell(row=row_num, column=actual_sku_col).value
                od_value = ws.cell(row=row_num, column=actual_od_col).value if actual_od_col else None
                stop_col_value = ws.cell(row=row_num, column=stop_col).value if stop_col else None
                
                # Check if we've hit the end (next stop header or totals)
                if sku_value:
                    sku_str = str(sku_value).strip().upper()
                    if "TOTAL" in sku_str or "BOX" in sku_str:
                        break
                if row_num > data_start_row and (
                    _looks_like_stop_header_text(stop_col_value)
                    or _looks_like_stop_header_text(qty_value)
                ):
                    break
                if qty_value is not None:
                    qty_text = str(qty_value).strip().upper()
                    if "TOTAL" in qty_text and "PHARMA" in qty_text:
                        break
                
                # Parse quantities
                qty = 0
                od_qty = 0
                
                try:
                    if qty_value is not None:
                        qty = float(str(qty_value))
                except:
                    pass
                
                try:
                    if od_value is not None and str(od_value).strip():
                        od_qty = float(str(od_value))
                except:
                    pass
                
                # Get SKU name
                sku = str(sku_value).strip() if sku_value is not None else ""
                
                # Add if valid SKU with quantity
                if sku and qty > 0 and sku.upper() not in ["BOX TOTAL", "TOTAL", ""]:
                    skus.append({
                        'stop': stop_name,
                        'sku': sku,
                        'qty': int(qty),
                        'od_qty': int(od_qty) if od_qty > 0 else 0,
                        'qty_cell': get_column_letter(actual_qty_col) + str(row_num),
                        'sku_cell': get_column_letter(actual_sku_col) + str(row_num),
                        'od_cell': get_column_letter(actual_od_col) + str(row_num) if actual_od_col else ""
                    })
            except Exception as e:
                continue
    
    if day_matches == 0:
        print(f"    No rows found for day '{day_name}' in mapping table")
        print(f"    Mapping table has days: {mapping_df[day_col].unique() if day_col else 'N/A'}")
    elif not exact_stop_matches and not store_type_matches:
        print(f"    Found {day_matches} rows for day '{day_name}' but none match stop '{stop_name}'")
        print(f"    Available stops for this day: {mapping_df[mapping_df[day_col].str.upper().str[:3] == day_normalized][stop_col].unique()[:5] if day_col and stop_col else 'N/A'}")
    elif len(skus) == 0:
        num_matches = len(exact_stop_matches) if exact_stop_matches else (len(store_type_matches) if store_type_matches else 0)
        print(f"    Found {num_matches} matching stop row(s) but no valid SKU/quantity data")
        print(f"    This may mean the cell addresses don't match the loading slip layout")
    
    return skus


def extract_sku_data_from_stop(ws, stop_col, header_row, mapping_df=None, day_name=""):
    """Extract all SKU/quantity pairs from a stop - uses mapping table if available"""
    stop_name = stop_col['name']
    
    # Use the row where the stop header was found, or fall back to header_row
    stop_header_row = stop_col.get('header_row', header_row)
    
    # Use mapping table when available (Week_42 for week<=58, Week_60 style for 59+)
    if mapping_df is not None and day_name:
        stop_col_info = {
            'col_num': stop_col.get('col_num'),
            'header_row': stop_header_row,
            'od_col': stop_col.get('od_col')
        }
        skus = extract_sku_data_using_mapping(ws, stop_name, day_name, mapping_df, stop_col_info)
        if skus:
            return skus
        # Mapping table exists but had no match for this stop/day (e.g. NFLD not in table yet)
        print(f"    No mapping match for this stop/day; add Day={day_name}, Stop to mapping table")
        return []
    
    # Fallback to manual scanning only when no mapping table was found at all
    print(f"    Using manual column scanning (mapping table not found)")
    skus = []
    col_num = stop_col['col_num']
    start_row = stop_header_row + 1 if stop_header_row else header_row + 1
    max_row = min(ws.max_row, start_row + 80)

    # Try multiple column layouts: [OD, Qty, SKU] blocks at various offsets from stop column
    layouts = [
        (col_num, col_num + 1, col_num + 2),       # stop | OD | Qty | SKU
        (col_num + 1, col_num + 2, col_num + 3),   # stop, OD | Qty | SKU
        (col_num - 1, col_num, col_num + 1),       # OD | Qty | SKU (stop left)
        (col_num + 2, col_num + 3, col_num + 4),   # stop, ?, OD | Qty | SKU
        (col_num + 3, col_num + 4, col_num + 5),   # wide block
        (col_num, col_num + 2, col_num + 3),       # OD | ? | Qty | SKU
    ]
    seen_row_sku = set()

    for row_num in range(start_row, max_row + 1):
        check_cell = str(ws.cell(row=row_num, column=col_num).value or "").strip().upper()
        if check_cell and "TOTAL" in check_cell and "BOX" in check_cell:
            break
        if check_cell and any(c.isdigit() for c in check_cell) and "." in check_cell and row_num > start_row + 8:
            # Likely next stop header (e.g. "7. Superstore")
            break

        for od_col, qty_col, sku_col in layouts:
            if od_col < 1 or qty_col < 1 or sku_col < 1 or sku_col > ws.max_column:
                continue
            od_value = ws.cell(row=row_num, column=od_col).value
            qty_value = ws.cell(row=row_num, column=qty_col).value
            sku_value = ws.cell(row=row_num, column=sku_col).value

            qty = 0
            od_qty = 0
            try:
                if qty_value is not None:
                    qty_str = str(qty_value).strip().replace(",", "")
                    if qty_str.replace(".", "").isdigit():
                        qty = float(qty_str)
            except (ValueError, TypeError):
                pass
            try:
                if od_value is not None:
                    od_str = str(od_value).strip().replace(",", "")
                    if od_str.replace(".", "").isdigit():
                        od_qty = float(od_str)
            except (ValueError, TypeError):
                pass

            sku = str(sku_value).strip() if sku_value is not None else ""
            if not sku or sku.upper() in ("TOTAL", "BOX TOTAL", "BOX", ""):
                continue
            if qty <= 0:
                continue
            key = (row_num, sku)
            if key in seen_row_sku:
                continue
            seen_row_sku.add(key)
            skus.append({
                'stop': stop_name,
                'sku': sku,
                'qty': int(qty),
                'od_qty': int(od_qty) if od_qty > 0 else 0,
                'row': row_num,
                'qty_col': get_column_letter(qty_col),
                'sku_col': get_column_letter(sku_col),
                'od_col': get_column_letter(od_col)
            })
            break  # one layout worked for this row
    return skus


def extract_all_stops_data(ws, header_row, week_number, day_name):
    """Extract data from all stops in the loading slip using mapping table"""
    print("\n=== EXTRACTING STOP DATA ===\n")
    
    # Load mapping table
    print("Loading mapping table...")
    mapping_df = load_mapping_table(week_number)

    # NFLD sheets use plain customer headers (e.g. Atlantic Grocery) instead of numbered stop names.
    # Build those stop entries from the mapping table so they don't get dropped.
    if str(day_name).strip().upper() == "NFLD" and mapping_df is not None:
        stops = _mapping_stops_for_day(mapping_df, day_name)
        if not stops:
            stops = find_stop_headers(ws, header_row)
    else:
        stops = find_stop_headers(ws, header_row)
    print(f"Found {len(stops)} stops:")
    for stop_key, stop_info in stops.items():
        print(f"  {stop_info['col_letter']}{stop_info['header_row']}: {stop_info['name']}")
    
    all_skus = []
    
    for stop_key, stop_info in stops.items():
        print(f"\nProcessing {stop_info['name']}...")
        stop_skus = extract_sku_data_from_stop(ws, stop_info, header_row, mapping_df, day_name)
        print(f"  Found {len(stop_skus)} SKUs")
        all_skus.extend(stop_skus)
    
    return all_skus


def determine_day_from_filename(filename):
    """Try to determine day from filename"""
    filename_upper = filename.upper()
    
    day_map = {
        'MON': 'Monday',
        'TUE': 'Tuesday',
        'WED': 'Wednesday',
        'THU': 'Thursday',
        'FRI': 'Friday',
        'SAT': 'Saturday',
        'SUN': 'Sunday',
        'MONDAY': 'Monday',
        'TUESDAY': 'Tuesday',
        'WEDNESDAY': 'Wednesday',
        'THURSDAY': 'Thursday',
        'FRIDAY': 'Friday',
        'SATURDAY': 'Saturday',
        'SUNDAY': 'Sunday'
    }
    
    for key, day in day_map.items():
        if key in filename_upper:
            return day
    
    # Default to Monday if not found
    return 'Monday'


def determine_day_from_sheet(ws):
    """Try to determine day from worksheet name"""
    sheet_name = ws.title.upper()
    
    day_map = {
        'MON': 'Monday',
        'TUE': 'Tuesday',
        'TUES': 'Tuesday',
        'WED': 'Wednesday',
        'THU': 'Thursday',
        'THUR': 'Thursday',
        'THURS': 'Thursday',
        'FRI': 'Friday',
        'SAT': 'Saturday',
        'SUN': 'Sunday',
        'MONDAY': 'Monday',
        'TUESDAY': 'Tuesday',
        'WEDNESDAY': 'Wednesday',
        'THURSDAY': 'Thursday',
        'FRIDAY': 'Friday',
        'SATURDAY': 'Saturday',
        'SUNDAY': 'Sunday'
    }
    
    for key, day in day_map.items():
        if key in sheet_name:
            return day
    
    return None


def extract_week_number(filename):
    """Extract week number from filename"""
    import re
    match = re.search(r'[Ww]eek\s*(\d+)', filename)
    if match:
        return match.group(1)
    return None


def _expected_year_for_slip_week(week_number):
    """Report weeks 1-52 belong to 2025; report weeks 56-61 belong to 2026."""
    try:
        wn = int(str(week_number).strip())
    except (TypeError, ValueError):
        return None
    if 1 <= wn <= 52:
        return 2025
    if 56 <= wn <= 61:
        return 2026
    return None


def _filename_matches_week(file_path, target_week, alt_week=None):
    """
    True when a loading slip filename is an exact week match.
    Avoids false matches like Week 3 -> Week 34 / Week 35.
    alt_week is only used for 2026 folder matching (calendar-week naming).
    """
    file_week = extract_week_number(file_path.name)
    if file_week is None:
        return False
    try:
        file_week_num = int(str(file_week).strip())
        target_week_num = int(str(target_week).strip())
    except (TypeError, ValueError):
        return False
    if file_week_num == target_week_num:
        return True
    if alt_week is not None:
        try:
            alt_week_num = int(str(alt_week).strip())
        except (TypeError, ValueError):
            return False
        return file_week_num == alt_week_num and "2026" in file_path.name.lower()
    return False


def _loading_slip_candidate_score(file_path, target_week):
    """
    Rank loading slip candidates so we prefer the correct year/folder and avoid auto-recovered copies.
    Higher score is better.
    """
    expected_year = _expected_year_for_slip_week(target_week)
    path_str = str(file_path).lower()
    name_lower = file_path.name.lower()
    score = 0
    if expected_year is not None and str(expected_year) in name_lower:
        score += 100
    if expected_year == 2025 and "2025 reports" in path_str:
        score += 80
    if expected_year == 2026 and "2026" in path_str:
        score += 80
    if "loading slip" in name_lower or "loading slipp" in name_lower:
        score += 20
    if "autorecovered" in name_lower:
        score -= 200
    if "copy" in name_lower:
        score -= 100
    # Prefer shallower/more canonical names over noisy variants.
    score -= len(file_path.name) / 1000.0
    return score


def _sheet_to_day_pairs():
    """Known worksheet names for each day."""
    return [
        ('Mon', 'Monday'), ('Monday', 'Monday'), ('MON', 'Monday'),
        ('Tues', 'Tuesday'), ('Tuesday', 'Tuesday'), ('TUES', 'Tuesday'), ('Tue', 'Tuesday'),
        ('Wed', 'Wednesday'), ('Wednesday', 'Wednesday'), ('WED', 'Wednesday'),
        ('Thurs', 'Thursday'), ('Thursday', 'Thursday'), ('THURS', 'Thursday'), ('Thur', 'Thursday'),
        ('Fri', 'Friday'), ('Friday', 'Friday'), ('FRI', 'Friday'),
        ('Sat', 'Saturday'), ('Saturday', 'Saturday'), ('SAT', 'Saturday'),
        ('Sun', 'Sunday'), ('Sunday', 'Sunday'), ('SUN', 'Sunday'),
        ('NFLD', 'NFLD'), ('Nfld', 'NFLD'), ('nfld', 'NFLD'),
    ]


def _get_sheets_by_day(file_path):
    """Return {day_name: sheet_name} for recognized sheets in a workbook."""
    try:
        wb_temp = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        sheetnames = wb_temp.sheetnames
        wb_temp.close()
    except Exception:
        return {}
    sheets_by_day = {}
    for sh in sheetnames:
        for sheet_key, day_name in _sheet_to_day_pairs():
            if sh == sheet_key and day_name not in sheets_by_day:
                sheets_by_day[day_name] = sh
                break
    return sheets_by_day


def _find_best_full_week_slip(week_number, search_dirs):
    """
    Find the best full-week loading slip candidate for this week.
    Prefer the correct year and a workbook with multiple recognized day sheets.
    Returns (file_path or None, sheets_by_day dict).
    """
    exclude_patterns = ['palletlines', 'traceability', 'master', 'inventory']
    _, iso_to_cal = _get_2026_slip_week_mapping()
    try:
        wn = int(week_number)
    except (TypeError, ValueError):
        wn = None
    candidates = []
    for data_dir in search_dirs:
        week_patterns = [str(week_number)]
        if wn is not None and wn in iso_to_cal and _is_2026_loading_slip_folder(data_dir):
            week_patterns.append(str(iso_to_cal[wn]))
        for wpat in week_patterns:
            for file_path in data_dir.glob("*.xlsx"):
                filename_lower = file_path.name.lower()
                if any(pattern in filename_lower for pattern in exclude_patterns):
                    continue
                if 'loading' not in filename_lower and 'slip' not in filename_lower:
                    continue
                if not _filename_matches_week(file_path, week_number, wpat if wpat != str(week_number) else None):
                    continue
                sheets_by_day = _get_sheets_by_day(file_path)
                score = _loading_slip_candidate_score(file_path, week_number)
                if len(sheets_by_day) >= 2:
                    score += 200
                candidates.append((score, file_path.stat().st_mtime, file_path, sheets_by_day))
    if not candidates:
        return None, {}
    candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
    best = candidates[0]
    return best[2], best[3]


def load_bb_date_from_repository(week_number, day_name=None):
    """
    Load Ship Date and/or BB Date from the week date repository (CSV).
    Supports per-day dates: Week, Day, ShipDate, BBDate (BBD can switch e.g. Wed).
    Use this when loading slips have formulas - no dependency on Excel calculated values.
    Returns (bb_date_str or None, ship_date_str or None).
    """
    repo_path = REF_DATA_DIR / "Week_ShipDate_BBDate.csv"
    if not repo_path.exists():
        return None, None
    try:
        df = pd.read_csv(repo_path)
        if df.empty or 'Week' not in df.columns:
            return None, None
        week_str = str(week_number).strip()
        week_match = df[df['Week'].astype(str).str.strip() == week_str]
        if week_match.empty:
            return None, None
        # Prefer (Week, Day) match when Day column and day_name present
        row = None
        if day_name and 'Day' in df.columns:
            day_norm = day_name.strip()
            day_match = week_match[week_match['Day'].astype(str).str.strip().str.lower() == day_norm.lower()]
            if not day_match.empty:
                row = day_match.iloc[0]
        if row is None:
            row = week_match.iloc[0]
        ship_date = None
        bb_date = None
        if 'ShipDate' in row.index and pd.notna(row.get('ShipDate')):
            ship_date = str(row['ShipDate']).strip()
            if ship_date and len(ship_date) >= 8:
                try:
                    datetime.strptime(ship_date, "%Y-%m-%d")
                except ValueError:
                    ship_date = None
        if 'BBDate' in row.index and pd.notna(row.get('BBDate')):
            bb_date = str(row['BBDate']).strip()
            if bb_date and len(bb_date) >= 8:
                try:
                    datetime.strptime(bb_date, "%Y-%m-%d")
                except ValueError:
                    bb_date = None
        return bb_date, ship_date
    except Exception as e:
        print(f"  Note: Could not read date repository: {e}")
        return None, None


def _find_slip_file_for_week(slip_week):
    """Return path to the loading slip file for the given week (multi-sheet preferred).
    Supports 2026 folder: when slip_week is 56-61, also looks for 'Week 1 Loading Slip 2026.xlsx' etc. (calendar-week naming).
    """
    exclude = {'palletlines', 'traceability', 'master', 'inventory', 'pallet', 'sku_final', 'mapping', 'polished'}
    week_str = str(slip_week)
    _, iso_to_cal = _get_2026_slip_week_mapping()
    best = None
    for data_dir in _get_loading_slip_dirs():
        for path in data_dir.rglob("*.xlsx"):
            name_lower = path.name.lower()
            if any(x in name_lower for x in exclude) or 'copy of' in name_lower or name_lower.startswith('copy '):
                continue
            if 'loading' not in name_lower and 'slip' not in name_lower:
                continue
            rest = name_lower.replace("week", " ", 1)
            parts = rest.split()
            if not parts or not parts[0].strip(" .").isdigit():
                continue
            file_week = int(parts[0].strip(" ."))
            # Match by slip week (e.g. 57) or by 2026 calendar week in 2026 folder (e.g. "Week 2 ... 2026" -> 57)
            if file_week == slip_week:
                pass
            elif _is_2026_loading_slip_folder(path.parent) and slip_week in iso_to_cal and file_week == iso_to_cal[slip_week]:
                if "2026" in name_lower:
                    pass  # match
                else:
                    continue
            else:
                continue
            has_day = any(d in name_lower for d in ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday', 'mon', 'tues', 'wed', 'thurs', 'fri'])
            if best is None or (not has_day and ('day' in str(best).lower() or best is None)):
                best = path
            elif not has_day:
                best = path
    return best


def get_slip_week_ending(slip_week):
    """
    Return the 'week ending' date for this loading slip week: Friday ship date + 1.
    Used to match slips to Hilly Acres (Week Ending in Inputs row 3 col C).
    Tries Week_ShipDate_BBDate.csv (Day=Friday) first, then reads Friday sheet from slip file (F2).
    Returns datetime.date or None.
    """
    # 1) Repository: Week N, Day Friday -> ShipDate; week ending = ShipDate + 1
    repo_bb, repo_ship = load_bb_date_from_repository(slip_week, "Friday")
    if repo_ship and len(repo_ship) >= 8:
        try:
            ship_dt = datetime.strptime(repo_ship.strip()[:10], "%Y-%m-%d")
            return (ship_dt + timedelta(days=1)).date()
        except ValueError:
            pass
    # 2) Open slip file, get Friday sheet, read ship date from F2/E2/G2
    path = _find_slip_file_for_week(slip_week)
    if not path or not path.exists():
        return None
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheet_name in ("Friday", "Fri", "FRIDAY"):
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for cell_ref in ("F2", "E2", "G2"):
                    try:
                        v = ws[cell_ref].value
                        d = _parse_bb_or_ship_date(v)
                        if d:
                            wb.close()
                            return (d + timedelta(days=1)).date()
                    except Exception:
                        continue
        wb.close()
    except Exception:
        pass
    return None


def _parse_bb_or_ship_date(cell_value, default_year=None):
    """Parse a cell value as date; supports 'Mar 22', 'Mar 22 2026', YYYY-MM-DD, M/D/YYYY, Excel serial."""
    if cell_value is None:
        return None
    year = default_year or datetime.now().year
    bb_date = None
    if isinstance(cell_value, datetime):
        bb_date = cell_value
    elif isinstance(cell_value, (int, float)):
        if 40000 < cell_value < 50000:
            bb_date = datetime(1900, 1, 1) + timedelta(days=int(cell_value) - 2)
    elif isinstance(cell_value, str):
        s = cell_value.strip()
        if not s:
            return None
        for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%b %d %Y", "%b %d", "%B %d, %Y"]:
            try:
                bb_date = datetime.strptime(s, fmt)
                # For formats without year (%b %d) or 2-digit year, use default year
                if fmt == "%b %d" or (fmt == "%m/%d/%y" and bb_date.year < 2000):
                    bb_date = bb_date.replace(year=year)
                break
            except ValueError:
                continue
    if bb_date and 2020 <= bb_date.year <= 2030:
        return bb_date
    return None


def extract_bb_date_from_slip(ws):
    """Extract BB date from loading slip. BBD often in I2/J2 (e.g. 'Mar 22'); Ship Date in F2 for year."""
    # First try to get year from Ship Date (e.g. F2 = "February 9, 2026")
    ship_year = None
    for cell_ref in ['F2', 'E2', 'G2']:
        try:
            v = ws[cell_ref].value
            if v is None:
                continue
            if isinstance(v, datetime):
                ship_year = v.year
                break
            if isinstance(v, str) and len(v) >= 4:
                # "February 9, 2026" or "Feb 9 2026"
                for fmt in ["%B %d, %Y", "%b %d, %Y", "%b %d %Y", "%Y-%m-%d"]:
                    try:
                        d = datetime.strptime(v.strip()[:50], fmt)
                        ship_year = d.year
                        break
                    except ValueError:
                        continue
                if ship_year:
                    break
        except Exception:
            continue

    if ship_year is None:
        ship_year = datetime.now().year

    # Cells where BBD often appears (slip shows "Mar 22" in I2)
    for cell_ref in ['I2', 'J2', 'I3', 'J3', 'I1', 'H1', 'G1']:
        try:
            cell_value = ws[cell_ref].value
            bb_date = _parse_bb_or_ship_date(cell_value, default_year=ship_year)
            if bb_date:
                print(f"  Found BB date in {cell_ref}: {bb_date.strftime('%Y-%m-%d')}")
                return bb_date.strftime("%Y-%m-%d")
        except Exception:
            continue

    # Scan row 2 for "Mar 22" / "BBD" style
    for col in range(1, min(ws.max_column + 1, 20)):
        try:
            cell_value = ws.cell(row=2, column=col).value
            bb_date = _parse_bb_or_ship_date(cell_value, default_year=ship_year)
            if bb_date:
                col_letter = get_column_letter(col)
                print(f"  Found BB date in {col_letter}2: {bb_date.strftime('%Y-%m-%d')}")
                return bb_date.strftime("%Y-%m-%d")
        except Exception:
            continue

    return None


def create_pallet_lines_format(sku_data, day_name, week_number, bb_date, barn_code=""):
    """Create PalletLines table format - matches tblPalletLines structure exactly"""
    print("\n=== CREATING PALLET LINES FORMAT ===\n")
    
    # Parse BB date
    try:
        current_bb_date = datetime.strptime(bb_date, "%Y-%m-%d")
    except:
        current_bb_date = datetime.now() + timedelta(days=47)
    
    # Prior week BB date (for OD quantities)
    prior_week_bb_date = (current_bb_date - timedelta(days=7)).strftime("%Y-%m-%d")
    current_bb_date_str = current_bb_date.strftime("%Y-%m-%d")
    
    print(f"Current week BB date: {current_bb_date_str}")
    print(f"Prior week BB date (for OD): {prior_week_bb_date}")
    
    # Create list of records matching PalletLines structure
    pallet_lines = []
    
    # Generate sequential IDs
    line_id = 1
    pallet_id = 1
    current_pallet_id = pallet_id
    
    for idx, record in enumerate(sku_data):
        # Each SKU/Stop/Qty combination gets a line
        # If OD quantity > 0, create separate line for OD items
        regular_qty = record['qty'] - record['od_qty']
        od_qty = record['od_qty']
        
        # Regular quantity line (if > 0) - uses current week BB date
        if regular_qty > 0:
            pallet_lines.append({
                'LineID': line_id,
                'PalletID': current_pallet_id,
                'SKU': record['sku'],
                'StopName': record['stop'],
                'QtyBoxes': regular_qty,
                'BBDate': current_bb_date_str,
                'IsPrevWeek': 0,  # False
                'BarnCode': barn_code,
                'WeekNumber': str(week_number),
                'DayName': day_name,
                'ODQty': 0
            })
            line_id += 1
        
        # OD quantity line (if > 0) - uses PRIOR WEEK BB date automatically
        if od_qty > 0:
            pallet_lines.append({
                'LineID': line_id,
                'PalletID': current_pallet_id,
                'SKU': record['sku'],
                'StopName': record['stop'],
                'QtyBoxes': od_qty,
                'BBDate': prior_week_bb_date,  # OD = prior week BB date automatically
                'IsPrevWeek': 1,  # True (OD is from prior week)
                'BarnCode': barn_code,
                'WeekNumber': str(week_number),
                'DayName': day_name,
                'ODQty': od_qty  # Mark as OD quantity
            })
            line_id += 1
        
        # Increment pallet ID every few items (or use same pallet for same stop)
        # For now, using same pallet ID for items from same stop
        if idx < len(sku_data) - 1:
            if sku_data[idx + 1]['stop'] != record['stop']:
                current_pallet_id += 1
    
    # Create DataFrame with exact PalletLines column order
    columns_order = [
        'LineID',
        'PalletID',
        'SKU',
        'StopName',
        'QtyBoxes',
        'BBDate',
        'IsPrevWeek',
        'BarnCode',
        'WeekNumber',
        'DayName',
        'ODQty'
    ]
    
    df = pd.DataFrame(pallet_lines)
    
    # Ensure all columns exist and in correct order
    for col in columns_order:
        if col not in df.columns:
            df[col] = 0 if col in ['LineID', 'PalletID', 'QtyBoxes', 'IsPrevWeek', 'ODQty'] else ""
    
    df = df[columns_order]
    
    return df


def generate_summary_stats_pallet_lines(df):
    """Generate summary statistics for PalletLines format"""
    print("\n=== SUMMARY STATISTICS ===\n")
    
    total_stops = df['StopName'].nunique()
    total_skus = df['SKU'].nunique()
    total_boxes = df['QtyBoxes'].sum()
    total_od_boxes = df[df['ODQty'] > 0]['QtyBoxes'].sum()
    total_regular_boxes = df[df['ODQty'] == 0]['QtyBoxes'].sum()
    total_lines = len(df)
    
    print(f"Total Stops: {total_stops}")
    print(f"Total Unique SKUs: {total_skus}")
    print(f"Total Boxes: {total_boxes:,}")
    print(f"  - Regular: {total_regular_boxes:,}")
    print(f"  - OD (Old Date): {total_od_boxes:,}")
    print(f"Total Pallet Lines: {total_lines}")
    
    # Per-stop summary
    print("\n=== PER-STOP SUMMARY ===")
    stop_summary = df.groupby('StopName').agg({
        'QtyBoxes': 'sum',
        'ODQty': 'sum',
        'SKU': 'count'
    }).rename(columns={'SKU': 'LineCount'})
    stop_summary['RegularQty'] = stop_summary['QtyBoxes'] - stop_summary['ODQty']
    stop_summary = stop_summary.sort_values('QtyBoxes', ascending=False)
    print(stop_summary.to_string())
    
    # Per-SKU summary
    print("\n=== PER-SKU SUMMARY ===")
    sku_summary = df.groupby('SKU').agg({
        'QtyBoxes': 'sum',
        'ODQty': 'sum',
        'StopName': 'count'
    }).rename(columns={'StopName': 'LineCount'})
    sku_summary['RegularQty'] = sku_summary['QtyBoxes'] - sku_summary['ODQty']
    sku_summary = sku_summary.sort_values('QtyBoxes', ascending=False)
    print(sku_summary.head(20).to_string())  # Top 20 SKUs
    
    return {
        'total_stops': total_stops,
        'total_skus': total_skus,
        'total_boxes': total_boxes,
        'total_od_boxes': total_od_boxes,
        'total_regular_boxes': total_regular_boxes,
        'total_lines': total_lines,
        'stop_summary': stop_summary,
        'sku_summary': sku_summary
    }


def save_pallet_lines_report(df, week_number, day_name, stats, output_dir=None):
    """Save PalletLines format report to Excel"""
    print("\n=== SAVING PALLET LINES REPORT ===\n")
    
    # Use specified output directory or default to exports
    if output_dir is None:
        output_dir = EXPORTS_DIR
        output_dir.mkdir(exist_ok=True)
    else:
        output_dir = Path(output_dir)
        output_dir.mkdir(exist_ok=True)
    
    # Generate filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Week{week_number}_{day_name}_PalletLines_{timestamp}.xlsx"
    filepath = output_dir / filename
    
    print(f"Saving to: {filepath}")
    
    # Create Excel writer
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Main PalletLines data sheet (matches tblPalletLines structure)
        df.to_excel(writer, sheet_name='PalletLines', index=False)
        
        # Summary sheet
        summary_data = {
            'Metric': ['Total Stops', 'Total SKUs', 'Total Lines', 'Total Boxes', 'Regular Boxes', 'OD Boxes'],
            'Value': [
                stats['total_stops'],
                stats['total_skus'],
                stats['total_lines'],
                stats['total_boxes'],
                stats['total_regular_boxes'],
                stats['total_od_boxes']
            ]
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
        
        # Per-stop summary
        stats['stop_summary'].to_excel(writer, sheet_name='Stop Summary')
        
        # Per-SKU summary
        stats['sku_summary'].to_excel(writer, sheet_name='SKU Summary')
        
        # Format worksheets
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Format header row
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    print(f"[OK] Report saved successfully!")
    return filepath


def find_week_files(week_number, data_dir):
    """Find all loading slip files for a given week. In 2026 folder, also matches calendar week (e.g. week 57 -> Week 2 ... 2026)."""
    files = {}
    candidates_by_day = {}
    day_patterns = {
        'Monday': ['mon', 'monday'],
        'Tuesday': ['tue', 'tuesday', 'tues'],
        'Wednesday': ['wed', 'wednesday'],
        'Thursday': ['thu', 'thursday', 'thurs'],
        'Friday': ['fri', 'friday'],
        'Saturday': ['sat', 'saturday'],
        'Sunday': ['sun', 'sunday'],
        'NFLD': ['nfld'],
    }
    exclude_patterns = ['palletlines', 'traceability', 'master', 'inventory']
    _, iso_to_cal = _get_2026_slip_week_mapping()
    try:
        wn = int(week_number)
    except (TypeError, ValueError):
        wn = None
    week_patterns = [str(week_number)]
    if wn is not None and wn in iso_to_cal and _is_2026_loading_slip_folder(data_dir):
        week_patterns.append(str(iso_to_cal[wn]))

    for wpat in week_patterns:
        for file_path in data_dir.glob("*.xlsx"):
            filename_lower = file_path.name.lower()
            if not _filename_matches_week(file_path, week_number, wpat if wpat != str(week_number) else None):
                continue
            if any(pattern in filename_lower for pattern in exclude_patterns):
                continue
            for day, patterns in day_patterns.items():
                if any(pattern in filename_lower for pattern in patterns):
                    candidates_by_day.setdefault(day, []).append(file_path)
                    break

        for file_path in data_dir.glob("*.xlsx"):
            filename_lower = file_path.name.lower()
            if not _filename_matches_week(file_path, week_number, wpat if wpat != str(week_number) else None):
                continue
            if any(pattern in filename_lower for pattern in exclude_patterns):
                continue
            if 'loading' in filename_lower or 'slip' in filename_lower:
                pass  # collected below if multi-sheet
    for day, day_candidates in candidates_by_day.items():
        if not day_candidates:
            continue
        best = sorted(day_candidates, key=lambda p: (_loading_slip_candidate_score(p, week_number), p.stat().st_mtime), reverse=True)[0]
        files[day] = best
    return files


def process_single_file(file_path, week_number, day_name, output_dir=None, save_day_report=False):
    """Process a single loading slip file. Set save_day_report=True to save per-day Excel files."""
    print(f"\n{'='*70}")
    print(f"PROCESSING: {file_path.name}")
    print(f"{'='*70}")
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # If day_name is specified and file has multiple sheets, try to find matching sheet
        if day_name:
            day_sheet_map = {
                'Monday': ['Mon', 'Monday', 'MON'],
                'Tuesday': ['Tues', 'Tuesday', 'TUES', 'Tues'],
                'Wednesday': ['Wed', 'Wednesday', 'WED'],
                'Thursday': ['Thurs', 'Thursday', 'THURS', 'Thur'],
                'Friday': ['Fri', 'Friday', 'FRI'],
                'Saturday': ['Sat', 'Saturday', 'SAT'],
                'Sunday': ['Sun', 'Sunday', 'SUN'],
                'NFLD': ['NFLD', 'Nfld', 'nfld'],
            }
            
            # Try to find matching sheet
            target_sheets = day_sheet_map.get(day_name, [])
            ws = None
            for sheet_name in target_sheets:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    print(f"  Using sheet: {sheet_name}")
                    break
            
            # Fallback to active sheet if no match found
            if ws is None:
                ws = wb.active
                print(f"  Using active sheet: {ws.title}")
        else:
            ws = wb.active
    except PermissionError:
        print(f"  ERROR: Cannot open file - it may be open in Excel")
        return None
    except Exception as e:
        print(f"  ERROR: Failed to load workbook: {e}")
        return None
    
    # Find header row
    header_row = find_header_row(ws)
    print(f"  Header row: {header_row}")
    
    # BB date: 1) date repository (no formula dependency), 2) extract from slip, 3) fallback
    bb_date = None
    repo_bb, repo_ship = load_bb_date_from_repository(week_number, day_name)
    if repo_bb:
        bb_date = repo_bb
        print(f"  BB date from repository: {bb_date}")
    elif repo_ship:
        try:
            ship_dt = datetime.strptime(repo_ship, "%Y-%m-%d")
            bb_date = (ship_dt + timedelta(days=47)).strftime("%Y-%m-%d")
            print(f"  BB date from repository (Ship {repo_ship} + 47 days): {bb_date}")
        except Exception:
            bb_date = None
    if not bb_date:
        bb_date = extract_bb_date_from_slip(ws)
    if not bb_date:
        print("  WARNING: BB date not found, deriving from Ship Date or week number")
        ship_date = None
        for cell_ref in ['F2', 'E2', 'G2']:
            try:
                v = ws[cell_ref].value
                if v is None:
                    continue
                if isinstance(v, datetime):
                    ship_date = v
                    break
                if isinstance(v, str) and len(v) >= 4:
                    for fmt in ["%B %d, %Y", "%b %d, %Y", "%b %d %Y", "%Y-%m-%d", "%m/%d/%Y"]:
                        try:
                            ship_date = datetime.strptime(v.strip()[:50], fmt)
                            break
                        except ValueError:
                            continue
                if ship_date:
                    break
            except Exception:
                continue
        if ship_date:
            bb_date = (ship_date + timedelta(days=47)).strftime("%Y-%m-%d")
            print(f"  Using Ship Date {ship_date.strftime('%Y-%m-%d')} + 47 days = {bb_date}")
        else:
            current_year = datetime.now().year
            try:
                week_num = int(week_number)
                jan1 = datetime(current_year, 1, 1)
                week_start = jan1 + timedelta(weeks=week_num - 1, days=-jan1.weekday())
                day_offset = {'Monday': 0, 'Tuesday': 1, 'Wednesday': 2, 'Thursday': 3,
                              'Friday': 4, 'Saturday': 5, 'Sunday': 6}
                ship_date_calc = week_start + timedelta(days=day_offset.get(day_name, 0))
                bb_date = (ship_date_calc + timedelta(days=47)).strftime("%Y-%m-%d")
            except Exception:
                bb_date = (datetime.now() + timedelta(days=47)).strftime("%Y-%m-%d")
    
    print(f"  BB Date: {bb_date}")
    
    # Extract data using mapping table
    sku_data = extract_all_stops_data(ws, header_row, week_number, day_name)
    if not sku_data:
        print(f"  ERROR: No SKU data found!")
        return None
    
    print(f"  Extracted {len(sku_data)} SKU records")
    
    # Create PalletLines format
    df = create_pallet_lines_format(sku_data, day_name, week_number, bb_date, "")
    
    # Generate statistics
    stats = generate_summary_stats_pallet_lines(df)
    
    # Save individual day report only if requested (default: weekly + master only)
    if save_day_report and output_dir:
        filepath = save_pallet_lines_report(df, week_number, day_name, stats, output_dir)
        print(f"  Saved: {filepath.name}")
    
    return df


def consolidate_weekly_data(all_day_dataframes):
    """Consolidate all days into one weekly report"""
    if not all_day_dataframes:
        return None
    
    # Combine all dataframes
    consolidated = pd.concat(all_day_dataframes, ignore_index=True)
    
    # Re-number LineIDs and PalletIDs sequentially
    consolidated['LineID'] = range(1, len(consolidated) + 1)
    
    # Group PalletIDs by stop (each stop gets its own pallet ID range)
    pallet_id = 1
    current_stop = None
    for idx in consolidated.index:
        if consolidated.at[idx, 'StopName'] != current_stop:
            current_stop = consolidated.at[idx, 'StopName']
            pallet_id += 1
        consolidated.at[idx, 'PalletID'] = pallet_id
    
    return consolidated


def append_to_master_inventory(df, master_file_path):
    """Append data to master inventory file (creates if doesn't exist)"""
    master_file = Path(master_file_path)
    
    if master_file.exists():
        # Read existing data
        try:
            existing_df = pd.read_excel(master_file, sheet_name='PalletLines')
            # Combine
            combined_df = pd.concat([existing_df, df], ignore_index=True)
            # Re-number LineIDs
            combined_df['LineID'] = range(1, len(combined_df) + 1)
        except:
            combined_df = df
    else:
        combined_df = df
    
    # Save master inventory
    with pd.ExcelWriter(master_file, engine='openpyxl') as writer:
        combined_df.to_excel(writer, sheet_name='PalletLines', index=False)
        
        # Format
        worksheet = writer.sheets['PalletLines']
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    return len(combined_df)


def main():
    """Main execution"""
    parser = argparse.ArgumentParser(description='Process weekly loading slips and generate PalletLines reports')
    parser.add_argument('--week', '-w', type=str, required=True,
                        help='Week number (e.g., 59)')
    parser.add_argument('--file', '-f', type=str, default=None,
                        help='Single file to process (if not provided, processes all days of week)')
    parser.add_argument('--day', '-d', type=str, default=None,
                        help='Day name (only used with --file)')
    parser.add_argument('--output-dir', '-o', type=str, default=None,
                        help='Output directory (default: Traceability_Exports, or same as input file if --same-folder)')
    parser.add_argument('--same-folder', action='store_true',
                        help='Save output in same folder as input file')
    parser.add_argument('--master-inventory', '-m', type=str, default=None,
                        help='Path to master inventory file (creates/updates master inventory for all weeks)')
    parser.add_argument('--daily', action='store_true',
                        help='Also save per-day PalletLines files (default: only weekly consolidated + master)')
    
    args = parser.parse_args()
    
    print("=" * 70)
    print("WEEKLY LOADING SLIP PROCESSOR")
    print("=" * 70)
    print(f"\nWeek: {args.week}")
    
    # Determine output directory (default: Traceability_Exports = "somewhere else" from Reference_Data)
    if args.same_folder:
        output_dir = REF_DATA_DIR
    elif args.output_dir:
        output_dir = Path(args.output_dir)
    else:
        output_dir = EXPORTS_DIR  # Traceability_Exports
        output_dir.mkdir(exist_ok=True)
    print(f"Output: {output_dir} (weekly report + master only; use --daily for per-day files)")
    
    all_day_dataframes = []
    all_stats = []
    
    # Process single file or all week files
    if args.file:
        # Single file mode: resolve path (run_all_weeks passes path relative to repo BASE_DIR)
        arg_path = Path(args.file)
        if arg_path.is_absolute() and arg_path.exists():
            file_path = arg_path
        else:
            file_path = BASE_DIR / arg_path
            if not file_path.exists():
                file_path = REF_DATA_DIR / arg_path
        if not file_path.exists():
            print(f"\nERROR: File not found: {file_path}")
            return
        
        # Sheet name to day name mapping (order preserved for sorting)
        sheet_to_day = [
            ('Mon', 'Monday'), ('Monday', 'Monday'), ('MON', 'Monday'),
            ('Tues', 'Tuesday'), ('Tuesday', 'Tuesday'), ('TUES', 'Tuesday'), ('Tues', 'Tuesday'),
            ('Wed', 'Wednesday'), ('Wednesday', 'Wednesday'), ('WED', 'Wednesday'),
            ('Thurs', 'Thursday'), ('Thursday', 'Thursday'), ('THURS', 'Thursday'), ('Thur', 'Thursday'),
            ('Fri', 'Friday'), ('Friday', 'Friday'), ('FRI', 'Friday'),
            ('Sat', 'Saturday'), ('Saturday', 'Saturday'),
            ('Sun', 'Sunday'), ('Sunday', 'Sunday'),
            ('NFLD', 'NFLD'), ('Nfld', 'NFLD'), ('nfld', 'NFLD'),
        ]
        day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'NFLD', 'Saturday', 'Sunday']
        
        if args.day:
            # Single day: process only that day's sheet
            day_name = args.day
            df = process_single_file(file_path, args.week, day_name, output_dir, save_day_report=args.daily)
            if df is not None:
                all_day_dataframes.append(df)
                stats = generate_summary_stats_pallet_lines(df)
                all_stats.append({'day': day_name, 'stats': stats})
        else:
            # No --day: process all day sheets in the file (full week from one file)
            try:
                wb_temp = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
                sheetnames = wb_temp.sheetnames
                wb_temp.close()
            except Exception as e:
                print(f"\nERROR: Cannot open file: {e}")
                return
            
            # Find which sheets correspond to which days
            sheets_by_day = {}
            for sh in sheetnames:
                for sheet_key, day_name in sheet_to_day:
                    if sh == sheet_key and day_name not in sheets_by_day:
                        sheets_by_day[day_name] = sh
                        break
            
            if not sheets_by_day:
                # Fallback: treat whole file as one day (e.g. active sheet)
                day_name = determine_day_from_filename(args.file) or 'Monday'
                print(f"\nNo day sheets found; processing as single day: {day_name}")
                df = process_single_file(file_path, args.week, day_name, output_dir, save_day_report=args.daily)
                if df is not None:
                    all_day_dataframes.append(df)
            else:
                print(f"\nProcessing all days from {file_path.name} ({len(sheets_by_day)} sheets)")
                for day_name in day_order:
                    if day_name not in sheets_by_day:
                        continue
                    df = process_single_file(file_path, args.week, day_name, output_dir, save_day_report=args.daily)
                    if df is not None:
                        all_day_dataframes.append(df)
                        stats = generate_summary_stats_pallet_lines(df)
                        all_stats.append({'day': day_name, 'stats': stats})
                        print(f"  ✓ Successfully processed {day_name}: {len(df)} records")
                    else:
                        print(f"  ✗ Failed to process {day_name}")
    else:
        # Process all days of the week — search Reference_Data and LoadingSlipsPaths (e.g. 2024 Reports)
        print("\nFinding all loading slips for Week {}...".format(args.week))
        search_dirs = _get_loading_slip_dirs()
        week_files = {}
        exclude_patterns = ['palletlines', 'traceability', 'master', 'inventory']
        _, iso_to_cal_2026 = _get_2026_slip_week_mapping()
        try:
            slip_week_num = int(args.week)
        except (TypeError, ValueError):
            slip_week_num = None

        # Best path: process the single full-week slip workbook for this week so Mon-Fri + NFLD
        # all come from the same correctly matched source file.
        best_full_week_file, best_sheets_by_day = _find_best_full_week_slip(args.week, search_dirs)
        if best_full_week_file is not None and len(best_sheets_by_day) >= 2:
            multi_sheet_processed = True
            print(f"\nSingle full-week slip detected: {best_full_week_file.name}")
            print(f"Processing all sheets: {', '.join(sorted(best_sheets_by_day.keys()))}")
            day_order_list = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'NFLD', 'Saturday', 'Sunday']
            for day_name in day_order_list:
                if day_name not in best_sheets_by_day:
                    continue
                print(f"\n{'='*70}")
                print(f"Processing {day_name} sheet: {best_full_week_file.name}")
                print(f"{'='*70}")
                df = process_single_file(best_full_week_file, args.week, day_name, output_dir, save_day_report=args.daily)
                if df is not None:
                    all_day_dataframes.append(df)
                    stats = generate_summary_stats_pallet_lines(df)
                    all_stats.append({'day': day_name, 'stats': stats})
                    print(f"  ✓ Successfully processed {day_name}: {len(df)} records")
                else:
                    print(f"  ✗ Failed to process {day_name}")
        else:
            multi_sheet_processed = False
        if not multi_sheet_processed:
            for data_dir in search_dirs:
                found = find_week_files(args.week, data_dir)
                for day, file_path in found.items():
                    current = week_files.get(day)
                    if current is None or _loading_slip_candidate_score(file_path, args.week) > _loading_slip_candidate_score(current, args.week):
                        week_files[day] = file_path
            if not week_files:
                print(f"\nERROR: No loading slip files found for Week {args.week}")
                print("Searched in:", [str(d.relative_to(BASE_DIR)) if BASE_DIR in d.parents else str(d) for d in search_dirs])
                for data_dir in search_dirs:
                    for f in sorted(data_dir.glob("*.xlsx"))[:15]:
                        print(f"  - {f.relative_to(BASE_DIR) if BASE_DIR in f.parents else f}")
                return

            print(f"\nFound {len(week_files)} file(s) to process:")
            for day, file_path in sorted(week_files.items()):
                print(f"  {day}: {file_path.name}")
            for day_name, file_path in sorted(week_files.items()):
                print(f"\n{'='*70}")
                print(f"Processing {day_name} file: {file_path.name}")
                print(f"{'='*70}")
                df = process_single_file(file_path, args.week, day_name, output_dir, save_day_report=args.daily)
                if df is not None:
                    all_day_dataframes.append(df)
                    stats = generate_summary_stats_pallet_lines(df)
                    all_stats.append({'day': day_name, 'stats': stats})
                    print(f"  ✓ Successfully processed {day_name}: {len(df)} records")
                else:
                    print(f"  ✗ Failed to process {day_name}")
    
    if not all_day_dataframes:
        print("\nERROR: No data processed!")
        return
    
    # Create consolidated weekly report
    print("\n" + "=" * 70)
    print("CREATING CONSOLIDATED WEEKLY REPORT")
    print("=" * 70)
    
    consolidated_df = consolidate_weekly_data(all_day_dataframes)
    
    # Generate consolidated statistics
    consolidated_stats = generate_summary_stats_pallet_lines(consolidated_df)
    
    # Save consolidated report
    consolidated_filepath = save_pallet_lines_report(
        consolidated_df, 
        args.week, 
        "AllDays", 
        consolidated_stats, 
        output_dir
    )
    
    print(f"\nConsolidated weekly report saved: {consolidated_filepath.name}")
    
    # Update master inventory if specified
    if args.master_inventory:
        master_path = Path(args.master_inventory)
        print(f"\nUpdating master inventory: {master_path}")
        total_rows = append_to_master_inventory(consolidated_df, master_path)
        print(f"  Master inventory now contains {total_rows:,} total records")
    elif not args.master_inventory and output_dir:
        # Auto-create master inventory in output directory
        master_path = output_dir / "Master_Inventory_All_Weeks.xlsx"
        print(f"\nUpdating master inventory: {master_path.name}")
        total_rows = append_to_master_inventory(consolidated_df, master_path)
        print(f"  Master inventory now contains {total_rows:,} total records")
    
    # Final summary
    print("\n" + "=" * 70)
    print("WEEKLY PROCESSING COMPLETE!")
    print("=" * 70)
    print(f"\nWeek {args.week} Summary:")
    print(f"  - Days processed: {len(all_day_dataframes)}")
    print(f"  - Total stops: {consolidated_stats['total_stops']}")
    print(f"  - Total SKUs: {consolidated_stats['total_skus']}")
    print(f"  - Total pallet lines: {consolidated_stats['total_lines']:,}")
    print(f"  - Total boxes: {consolidated_stats['total_boxes']:,}")
    print(f"    - Regular: {consolidated_stats['total_regular_boxes']:,}")
    print(f"    - OD: {consolidated_stats['total_od_boxes']:,}")
    
    # Per-day breakdown
    if len(all_stats) > 1:
        print(f"\nPer-Day Breakdown:")
        for day_info in all_stats:
            day = day_info['day']
            stats = day_info['stats']
            print(f"  {day}: {stats['total_boxes']:,} boxes ({stats['total_stops']} stops, {stats['total_skus']} SKUs)")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nProcess cancelled by user.")
        sys.exit(0)
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
